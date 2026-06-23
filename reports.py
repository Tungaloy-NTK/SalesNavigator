import io
import streamlit as st
import pandas as pd
from datetime import date, datetime
import database as db

# ── PDF via reportlab ────────────────────────────────────────────────────────────
def _build_pdf(df: pd.DataFrame, title: str, subtitle: str) -> bytes:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=1*cm, rightMargin=1*cm,
        topMargin=1.5*cm, bottomMargin=1.5*cm,
    )
    styles = getSampleStyleSheet()
    brand_red   = colors.HexColor("#c0392b")
    brand_dark  = colors.HexColor("#1a1a2e")
    header_grey = colors.HexColor("#f2f2f2")

    story = []

    # Title block
    title_style = ParagraphStyle("title", parent=styles["Title"],
                                 textColor=brand_dark, fontSize=16, spaceAfter=4)
    sub_style   = ParagraphStyle("sub",   parent=styles["Normal"],
                                 textColor=colors.grey, fontSize=9, spaceAfter=8)
    story.append(Paragraph("Tungaloy-NTK Sales Navigator", title_style))
    story.append(Paragraph(title, ParagraphStyle("t2", parent=styles["Heading2"],
                                                  textColor=brand_red, fontSize=12, spaceAfter=2)))
    story.append(Paragraph(subtitle, sub_style))
    story.append(Spacer(1, 0.3*cm))

    if df.empty:
        story.append(Paragraph("No data for the selected filters.", styles["Normal"]))
        doc.build(story)
        return buf.getvalue()

    # Currency / pct columns
    money_cols = {"Sales £", "Cost £", "GP £"}
    pct_cols   = {"GP %"}
    qty_cols   = {"Qty"}

    def fmt_cell(col, val):
        if val is None or (isinstance(val, float) and pd.isna(val)):
            return "—"
        if col in money_cols:
            try: return f"£{float(val):,.0f}"
            except: return str(val)
        if col in pct_cols:
            try: return f"{float(val):.1f}%"
            except: return str(val)
        if col in qty_cols:
            try: return f"{float(val):,.0f}"
            except: return str(val)
        return str(val)

    # Build table data
    headers = list(df.columns)
    table_data = [headers]
    for _, row in df.iterrows():
        table_data.append([fmt_cell(col, row[col]) for col in headers])

    # Column widths — distribute across landscape A4 (~27cm usable)
    page_w = landscape(A4)[0] - 2*cm
    col_weights = {
        "Date": 2.0, "Cust Code": 1.4, "Customer": 3.0, "Ship To Address": 4.0,
        "Salesman": 2.2, "SM Name": 1.8, "Item Code": 1.8, "Description": 3.2,
        "Application": 2.5, "Qty": 1.0, "Sales £": 1.6, "Cost £": 1.6,
        "GP £": 1.6, "GP %": 1.2, "Invoice": 1.8,
    }
    total_w = sum(col_weights.get(c, 2.0) for c in headers)
    col_widths = [(col_weights.get(c, 2.0) / total_w) * page_w for c in headers]

    tbl = Table(table_data, colWidths=col_widths, repeatRows=1)
    tbl.setStyle(TableStyle([
        # Header row
        ("BACKGROUND",    (0, 0), (-1, 0),  brand_dark),
        ("TEXTCOLOR",     (0, 0), (-1, 0),  colors.white),
        ("FONTNAME",      (0, 0), (-1, 0),  "Helvetica-Bold"),
        ("FONTSIZE",      (0, 0), (-1, 0),  7),
        ("BOTTOMPADDING", (0, 0), (-1, 0),  5),
        ("TOPPADDING",    (0, 0), (-1, 0),  5),
        # Data rows
        ("FONTNAME",      (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE",      (0, 1), (-1, -1), 7),
        ("TOPPADDING",    (0, 1), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 1), (-1, -1), 3),
        ("ROWBACKGROUNDS",(0, 1), (-1, -1), [colors.white, header_grey]),
        ("GRID",          (0, 0), (-1, -1), 0.4, colors.HexColor("#dddddd")),
        ("ALIGN",         (0, 0), (-1, -1), "LEFT"),
        # Right-align numeric columns
        *[("ALIGN", (headers.index(c), 0), (headers.index(c), -1), "RIGHT")
          for c in money_cols | pct_cols | qty_cols if c in headers],
        # Red line under header
        ("LINEBELOW",     (0, 0), (-1, 0),  1.5, brand_red),
    ]))
    story.append(tbl)

    # Footer note
    story.append(Spacer(1, 0.4*cm))
    story.append(Paragraph(
        f"Generated {datetime.now().strftime('%d %b %Y %H:%M')} | Tungaloy-NTK Sales Navigator",
        ParagraphStyle("footer", parent=styles["Normal"], textColor=colors.grey, fontSize=7)
    ))

    doc.build(story)
    return buf.getvalue()


# ── Excel via openpyxl ────────────────────────────────────────────────────────────
def _build_excel(df: pd.DataFrame, title: str, subtitle: str) -> bytes:
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        # Write data starting at row 3 (leave space for title rows)
        df.to_excel(writer, sheet_name="Sales Report", index=False, startrow=3)
        ws = writer.sheets["Sales Report"]

        from openpyxl.styles import (Font, PatternFill, Alignment,
                                      Border, Side, numbers)
        from openpyxl.utils import get_column_letter

        brand_dark = "1A1A2E"
        brand_red  = "C0392B"
        light_grey = "F2F2F2"
        thin = Side(style="thin", color="DDDDDD")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Title rows
        ws["A1"] = "Tungaloy-NTK Sales Navigator"
        ws["A1"].font = Font(bold=True, size=14, color=brand_dark)
        ws["A2"] = f"{title} | {subtitle}"
        ws["A2"].font = Font(size=10, color=brand_red)

        # Style header row (row 4 = startrow+1)
        header_row = 4
        for cell in ws[header_row]:
            cell.font      = Font(bold=True, color="FFFFFF", size=10)
            cell.fill      = PatternFill("solid", fgColor=brand_dark)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = border
        ws.row_dimensions[header_row].height = 30

        # Style data rows
        money_cols = {"Sales £", "Cost £", "GP £"}
        pct_cols   = {"GP %"}
        qty_cols   = {"Qty"}
        cols = list(df.columns)

        for row_idx, row in enumerate(ws.iter_rows(min_row=5, max_row=ws.max_row), start=0):
            fill_colour = "FFFFFF" if row_idx % 2 == 0 else light_grey
            for cell in row:
                cell.fill      = PatternFill("solid", fgColor=fill_colour)
                cell.border    = border
                cell.alignment = Alignment(vertical="center")

        # Number formats
        for col_idx, col_name in enumerate(cols, start=1):
            col_letter = get_column_letter(col_idx)
            if col_name in money_cols:
                for cell in ws[col_letter][4:]:
                    cell.number_format = '£#,##0'
                    cell.alignment = Alignment(horizontal="right")
            elif col_name in pct_cols:
                for cell in ws[col_letter][4:]:
                    cell.number_format = '0.0"%"'
                    cell.alignment = Alignment(horizontal="right")
            elif col_name in qty_cols:
                for cell in ws[col_letter][4:]:
                    cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal="right")

        # Auto-fit column widths
        for col_idx, col_name in enumerate(cols, start=1):
            col_letter = get_column_letter(col_idx)
            max_len = max(
                len(str(col_name)),
                *[len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(5, ws.max_row + 1)]
            ) if ws.max_row >= 5 else len(str(col_name))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

        # Freeze header
        ws.freeze_panes = "A5"

        # Summary totals row
        if not df.empty:
            total_row = ws.max_row + 1
            ws.cell(total_row, 1, "TOTAL")
            ws.cell(total_row, 1).font = Font(bold=True)
            for col_idx, col_name in enumerate(cols, start=1):
                if col_name in money_cols | qty_cols:
                    col_letter = get_column_letter(col_idx)
                    cell = ws.cell(total_row, col_idx,
                                   df[col_name].sum() if col_name in df else 0)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill("solid", fgColor="E8E8E8")
                    cell.border = border
                    if col_name in money_cols:
                        cell.number_format = '£#,##0'
                        cell.alignment = Alignment(horizontal="right")
                    else:
                        cell.number_format = '#,##0'
                        cell.alignment = Alignment(horizontal="right")

    return buf.getvalue()


# ── Customer Dashboard PDF ───────────────────────────────────────────────────────
def _make_chart_image(fig, width_inch, height_inch) -> bytes:
    """Render a matplotlib figure to PNG bytes at exact figsize dimensions.
    No bbox_inches='tight' — caller must use subplots_adjust so all content
    fits within the figure bounds. This guarantees the PNG aspect ratio matches
    width_inch:height_inch, preventing distortion when displayed in the PDF."""
    import matplotlib
    matplotlib.use("Agg")
    buf = io.BytesIO()
    fig.set_size_inches(width_inch, height_inch)
    fig.savefig(buf, format="png", dpi=150, facecolor=fig.get_facecolor())
    buf.seek(0)
    return buf.read()


def _build_dashboard_pdf(df: pd.DataFrame, customer_name: str,
                          date_from, date_to, prepared_by: str,
                          hide_costs: bool = False,
                          df_py: pd.DataFrame = None,
                          consignment_data: dict = None) -> bytes:
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import matplotlib.patches as mpatches
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                    Table, TableStyle, Image, HRFlowable,
                                    PageBreak, KeepTogether)
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT

    brand_red  = "#c0392b"
    brand_dark = "#1a1a2e"
    MRED  = colors.HexColor(brand_red)
    MDARK = colors.HexColor(brand_dark)
    MGREY = colors.HexColor("#f5f5f5")
    plt.rcParams.update({"font.family": "sans-serif", "font.size": 9})

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles  = getSampleStyleSheet()
    story   = []

    def style(name, **kw):
        return ParagraphStyle(name, parent=styles["Normal"], **kw)

    # ── Prepare data ──────────────────────────────────────────────────────────
    df = df.copy()
    for c in ["Sales £", "Cost £", "GP £", "Qty"]:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)

    total_sales  = df["Sales £"].sum()
    total_gp     = df["GP £"].sum() if "GP £" in df.columns else 0
    total_qty    = df["Qty"].sum()
    num_products = df["Item Code"].nunique() if "Item Code" in df.columns else 0
    gp_pct       = total_gp / total_sales * 100 if total_sales else 0

    # Prior year same period
    py_from = date_from.replace(year=date_from.year - 1)
    py_to   = date_to.replace(year=date_to.year - 1)

    # Monthly trend from current data
    if "Date" in df.columns:
        df["_month"] = pd.to_datetime(df["Date"], errors="coerce").dt.to_period("M")
        monthly = df.groupby("_month")["Sales £"].sum().sort_index().tail(12)
    else:
        monthly = pd.Series(dtype=float)

    # Prior year monthly trend (shifted forward 1 year for alignment)
    monthly_py = pd.Series(dtype=float)
    if df_py is not None and not df_py.empty and "Date" in df_py.columns:
        df_py = df_py.copy()
        df_py["Sales £"] = pd.to_numeric(df_py["Sales £"], errors="coerce").fillna(0)
        df_py["_month"] = pd.to_datetime(df_py["Date"], errors="coerce").dt.to_period("M")
        py_raw = df_py.groupby("_month")["Sales £"].sum().sort_index()
        # Shift prior year periods forward 1 year so they align with current year x-axis
        monthly_py = pd.Series(
            py_raw.values,
            index=py_raw.index.map(lambda p: p + 12),
        ).reindex(monthly.index, fill_value=0)

    # Top items by sales (bottom computed later once py lookup is available)
    if "Item Code" in df.columns:
        _all_items = (df.groupby(["Item Code", "Description"])["Sales £"]
                      .sum().reset_index()
                      .sort_values("Sales £", ascending=False))
        top_items = _all_items.head(10)
    else:
        top_items = pd.DataFrame()

    # Applications
    if "Application" in df.columns:
        app_data = (df[df["Application"].notna() & (df["Application"] != "")]
                    .groupby("Application")["Sales £"].sum()
                    .sort_values(ascending=False).head(8))
    else:
        app_data = pd.Series(dtype=float)

    # ── Page 1: Header + metrics + trend ─────────────────────────────────────

    # Branded header bar — logo on right, title text on left
    import os as _os
    _logo_path = _os.path.join(_os.path.dirname(__file__), "logo.png")
    _logo_cell = Image(_logo_path, width=4.5*cm, height=1.1*cm) if _os.path.exists(_logo_path) else Paragraph("", style("hdr", fontSize=9))
    header_data = [[
        Paragraph(
            f'<font color="white"><b>TUNGALOY-NTK</b>  |  Customer Account Report</font>',
            style("hdr", fontSize=13, textColor=colors.white)),
        _logo_cell,
    ]]
    header_tbl = Table(header_data, colWidths=[doc.width - 5*cm, 5*cm])
    header_tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0,0), (-1,-1), MDARK),
        ("TOPPADDING",    (0,0), (-1,-1), 8),
        ("BOTTOMPADDING", (0,0), (-1,-1), 8),
        ("LEFTPADDING",   (0,0), (-1,-1), 12),
        ("RIGHTPADDING",  (0,0), (-1,-1), 8),
        ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
        ("ALIGN",         (1,0), (1,0),   "RIGHT"),
    ]))
    story.append(header_tbl)
    story.append(Spacer(1, 0.4*cm))

    story.append(Paragraph(customer_name,
                            style("cname", fontSize=18, fontName="Helvetica-Bold",
                                  textColor=MDARK, spaceAfter=6, leading=22)))
    period_str = f"{date_from.strftime('%d %b %Y')} – {date_to.strftime('%d %b %Y')}"
    story.append(Paragraph(
        f"Period: {period_str}  |  Prepared by: {prepared_by}  |  {datetime.now().strftime('%d %b %Y')}",
        style("sub", fontSize=9, textColor=colors.grey, spaceAfter=10)
    ))
    story.append(HRFlowable(width="100%", thickness=2, color=MRED, spaceAfter=14))

    # Summary metric cards
    def metric_cell(label, value, note=""):
        inner = [[Paragraph(f'<font color="#888888" size="8">{label}</font>', styles["Normal"])],
                 [Paragraph(f'<b>{value}</b>', style("mv", fontSize=16, textColor=MDARK))]]
        if note:
            inner.append([Paragraph(f'<font color="#888" size="7">{note}</font>', styles["Normal"])])
        t = Table(inner, colWidths=[4.2*cm])
        t.setStyle(TableStyle([
            ("BACKGROUND",    (0,0), (-1,-1), MGREY),
            ("ROUNDEDCORNERS",(0,0), (-1,-1), 4),
            ("TOPPADDING",    (0,0), (-1,-1), 8),
            ("BOTTOMPADDING", (0,0), (-1,-1), 8),
            ("LEFTPADDING",   (0,0), (-1,-1), 10),
            ("LINEBELOW",     (0,0), (-1,0),  1.5, MRED),
        ]))
        return t

    # ── Trend chart (left) + metric cards stacked (right) ───────────────────────
    # Page height budget (A4 portrait, 26.7 cm usable):
    #   Header + title + HR  ≈  3.5 cm
    #   Trend chart          ≈  5.5 cm  + 0.35 spacer
    #   Brand chart          ≈  5.0 cm  + 0.35 spacer
    #   Application chart    ≈  6.5 cm  + 0.35 spacer
    #   Product Cat chart    ≈  4.6 cm  + 0.35 spacer
    #   ──────────────────────────────────────────────
    #   Total                ≈ 26.5 cm  ✓ fills page 1
    CHART_COL  = 13.0 * cm
    STATS_COL  =  5.0 * cm
    CHART_W_IN = 13.0 / 2.54   # inches for matplotlib
    CHART_H_IN =  5.5 / 2.54   # taller to fill more of the page

    # Build trend chart image
    trend_img = None
    if not monthly.empty:
        import calendar as _cal
        import matplotlib.patches as mpatches
        fig, ax = plt.subplots(facecolor="white")
        months = [_cal.month_abbr[p.month] for p in monthly.index]
        x      = range(len(months))
        has_py = not monthly_py.empty and monthly_py.sum() > 0

        def _bar_labels(ax, bars, color="white"):
            for bar in bars:
                h = bar.get_height()
                if h > 0:
                    ax.text(bar.get_x() + bar.get_width() / 2, h / 2,
                            f"£{h:,.0f}", ha="center", va="center",
                            rotation=90, fontsize=6.0, color=color,
                            fontweight="bold", zorder=5)

        if has_py:
            import numpy as np
            x_arr = np.arange(len(months))
            w = 0.38
            py_vals  = [monthly_py.get(p, 0) for p in monthly.index]
            py_bars  = ax.bar(x_arr - w/2, py_vals,        width=w, color="#aab4be", zorder=2)
            cur_bars = ax.bar(x_arr + w/2, monthly.values, width=w, color=brand_dark, zorder=2)
            _bar_labels(ax, py_bars,  color=brand_dark)
            _bar_labels(ax, cur_bars, color="white")
            ax.set_xticks(x_arr)
            ax.set_xticklabels(months, rotation=0, ha="center", fontsize=8)
            legend_handles = [
                mpatches.Patch(color="#aab4be", label=str(date_from.year - 1)),
                mpatches.Patch(color=brand_dark,  label=str(date_from.year)),
            ]
            ax.legend(handles=legend_handles, fontsize=7, frameon=False, loc="upper left")
            ax.set_title("Monthly Sales — Current vs Prior Year",
                         fontsize=10, fontweight="bold", color=brand_dark)
        else:
            bars = ax.bar(x, monthly.values, color=brand_dark, width=0.6)
            _bar_labels(ax, bars, color="white")
            ax.set_xticks(list(x))
            ax.set_xticklabels(months, rotation=0, ha="center", fontsize=8)
            ax.set_title("Monthly Sales Trend", fontsize=10, fontweight="bold", color=brand_dark)

        ax.set_ylabel("Sales £", fontsize=8)
        ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda v, _: f"£{v:,.0f}"))
        ax.spines[["top","right"]].set_visible(False)
        ax.grid(axis="y", alpha=0.3, zorder=0)
        fig.subplots_adjust(left=0.13, right=0.97, top=0.88, bottom=0.18)
        chart_bytes = _make_chart_image(fig, CHART_W_IN, CHART_H_IN)
        plt.close(fig)
        trend_img = Image(io.BytesIO(chart_bytes), width=CHART_COL, height=4.5*cm)

    # Build stacked metric cards — right column alongside trend chart
    # Cards are 25% smaller than original and must fit within CHART_H (5.5 cm)
    def small_metric(label, value):
        inner = [
            [Paragraph(f'<font color="#888888" size="6">{label}</font>', styles["Normal"])],
            [Paragraph(f'<b>{value}</b>',
                       style("smv", fontSize=8, textColor=MDARK, leading=10))],
        ]
        t = Table(inner, colWidths=[STATS_COL - 0.3*cm])
        t.setStyle(TableStyle([
            ("BACKGROUND",    (0,0), (-1,-1), MGREY),
            ("TOPPADDING",    (0,0), (-1,-1), 4),
            ("BOTTOMPADDING", (0,0), (-1,-1), 4),
            ("LEFTPADDING",   (0,0), (-1,-1), 8),
            ("LINEBELOW",     (0,0), (-1,0),  1.2, MRED),
        ]))
        return t

    # Prior-year totals and variance for metric cards
    py_total_sales_card = (
        df_py["Sales £"].sum()
        if df_py is not None and not df_py.empty and "Sales £" in df_py.columns
        else None
    )
    var_sales_card = (total_sales - py_total_sales_card) if py_total_sales_card is not None else None

    cur_yr_label = str(date_from.year)
    pri_yr_label = str(date_from.year - 1)

    stats_stack = [
        [small_metric(f"Sales {cur_yr_label}",  f"£{total_sales:,.0f}")],
        [Spacer(1, 0.2*cm)],
        [small_metric(
            f"Sales {pri_yr_label}",
            f"£{py_total_sales_card:,.0f}" if py_total_sales_card is not None else "—"
        )],
        [Spacer(1, 0.2*cm)],
        [small_metric(
            "Var £",
            f"£{var_sales_card:+,.0f}" if var_sales_card is not None else "—"
        )],
    ]
    stats_tbl = Table(stats_stack, colWidths=[STATS_COL])
    stats_tbl.setStyle(TableStyle([
        ("LEFTPADDING",   (0,0), (-1,-1), 0),
        ("RIGHTPADDING",  (0,0), (-1,-1), 0),
        ("TOPPADDING",    (0,0), (-1,-1), 0),
        ("BOTTOMPADDING", (0,0), (-1,-1), 0),
        ("VALIGN",        (0,0), (-1,-1), "TOP"),
    ]))

    # Combine trend chart + metrics into one row
    left_cell = trend_img if trend_img else Spacer(CHART_COL, 5.5*cm)
    combo_row = Table([[left_cell, stats_tbl]], colWidths=[CHART_COL, STATS_COL])
    combo_row.setStyle(TableStyle([
        ("VALIGN",        (0,0), (-1,-1), "TOP"),
        ("LEFTPADDING",   (0,0), (-1,-1), 0),
        ("RIGHTPADDING",  (0,0), (-1,-1), 0),
        ("TOPPADDING",    (0,0), (-1,-1), 0),
        ("BOTTOMPADDING", (0,0), (-1,-1), 0),
    ]))
    story.append(combo_row)
    story.append(Spacer(1, 0.35*cm))

    # ── Colour palettes ───────────────────────────────────────────────────────────
    # Brand-specific fixed colours
    BRAND_COLOURS = {
        "NTK":      "#27ae60",   # green
        "INNO":     "#5dade2",   # light blue
        "TUK":      "#95a5a6",   # grey
        "TUNGALOY": "#c0392b",   # brand red
    }
    # Rich varied palette for Application / Product Category
    RICH_PALETTE = [
        "#c0392b",  # red
        "#27ae60",  # green
        "#5dade2",  # light blue
        "#f39c12",  # amber
        "#8e44ad",  # purple
        "#16a085",  # teal
        "#e67e22",  # orange
        "#2980b9",  # dark blue
        "#d35400",  # burnt orange
        "#1abc9c",  # turquoise
        "#95a5a6",  # grey
        "#2ecc71",  # light green
    ]

    # ── Helper: donut chart ───────────────────────────────────────────────────────
    # Donut figure: width=2.36 in, height=2.83 in (20% taller than wide).
    # For a SQUARE axes within this non-square figure:
    #   axes height fraction = top - bottom = 0.87 - 0.30 = 0.57
    #   axes width fraction  = (0.57 × 2.83) / 2.36 = 0.684  →  left=0.158, right=0.842
    # → axes is 1.614 in wide × 1.613 in tall → essentially square → perfect circle ✓
    # Legend anchor at (0.5, -0.42) in axes coords puts legend bottom at
    #   figure y = 0.30 + (−0.42 × 0.57) = 0.061  →  just above figure edge, no overlap.
    _D_W = 2.36   # figure width  in inches (= 6.0 cm / 2.54)
    _D_H = 2.83   # figure height in inches (= 7.2 cm / 2.54, 20% bigger)

    def _donut(series, title, colour_map=None):
        if colour_map:
            pie_colours = [colour_map.get(k, RICH_PALETTE[i % len(RICH_PALETTE)])
                           for i, k in enumerate(series.index)]
        else:
            pie_colours = [RICH_PALETTE[i % len(RICH_PALETTE)] for i in range(len(series))]
        # Truncate labels so they never overflow the cell width
        labels_trunc = [str(l)[:13] for l in series.index]
        fig, ax = plt.subplots(figsize=(_D_W, _D_H), facecolor="white")
        wedges, _, autotexts = ax.pie(
            series.values, labels=None, colors=pie_colours,
            autopct="%1.0f%%", startangle=90, pctdistance=0.75,
            wedgeprops=dict(width=0.55))
        for at in autotexts:
            at.set_fontsize(7)
        # Square axes (computed above) → perfect circle
        fig.subplots_adjust(top=0.87, bottom=0.30, left=0.158, right=0.842)
        # Anchor well below the axes so even 4-row legends never overlap the pie
        ax.legend(labels_trunc, loc="lower center", bbox_to_anchor=(0.5, -0.42),
                  ncol=2, fontsize=6, frameon=False)
        ax.set_title(title, fontsize=9, fontweight="bold", color=brand_dark, pad=4)
        b = _make_chart_image(fig, _D_W, _D_H)
        plt.close(fig)
        return b

    # ── Prepare additional chart data ────────────────────────────────────────────
    brand_data = pd.Series(dtype=float)
    if "Brand" in df.columns:
        brand_data = (df[df["Brand"].notna() & (df["Brand"] != "")]
                      .groupby("Brand")["Sales £"].sum()
                      .sort_values(ascending=False))

    # Prior-year brand totals (only populated when df_py is provided)
    brand_data_py = pd.Series(dtype=float)
    if df_py is not None and not df_py.empty and "Brand" in df_py.columns:
        df_py["Sales £"] = pd.to_numeric(df_py.get("Sales £", 0), errors="coerce").fillna(0)
        brand_data_py = (df_py[df_py["Brand"].notna() & (df_py["Brand"] != "")]
                         .groupby("Brand")["Sales £"].sum())

    itype_data = pd.Series(dtype=float)
    if "Item Type" in df.columns:
        itype_data = (df[df["Item Type"].notna() & (df["Item Type"] != "")]
                      .groupby("Item Type")["Sales £"].sum()
                      .sort_values(ascending=False))

    # Prior-year application + product category totals
    app_data_py   = pd.Series(dtype=float)
    itype_data_py = pd.Series(dtype=float)
    if df_py is not None and not df_py.empty:
        _dpy = df_py.copy()
        _dpy["Sales £"] = pd.to_numeric(_dpy.get("Sales £", 0), errors="coerce").fillna(0)
        if "Application" in _dpy.columns:
            app_data_py = (_dpy[_dpy["Application"].notna() & (_dpy["Application"] != "")]
                           .groupby("Application")["Sales £"].sum())
        if "Item Type" in _dpy.columns:
            itype_data_py = (_dpy[_dpy["Item Type"].notna() & (_dpy["Item Type"] != "")]
                             .groupby("Item Type")["Sales £"].sum())

    # Top accounts (meaningful when multiple customers in data)
    multi_account = "Customer" in df.columns and df["Customer"].nunique() > 1
    top_accounts = pd.DataFrame()
    if multi_account:
        top_accounts = (df.groupby("Customer")["Sales £"].sum()
                        .sort_values(ascending=False).head(10).reset_index())

    # Salesman breakdown (meaningful when multiple reps)
    salesman_data = pd.Series(dtype=float)
    if "Salesman" in df.columns and df["Salesman"].nunique() > 1:
        salesman_data = (df[df["Salesman"].notna()]
                         .groupby("Salesman")["Sales £"].sum()
                         .sort_values(ascending=False))

    # ── Helper: lay a list of chart Images into a single-row table ───────────────
    def _chart_row(charts, col_widths):
        if not charts:
            return None
        while len(charts) < len(col_widths):
            charts.append(Spacer(col_widths[len(charts)], col_widths[0]))
        t = Table([charts], colWidths=col_widths)
        t.setStyle(TableStyle([
            ("VALIGN",       (0,0), (-1,-1), "TOP"),
            ("LEFTPADDING",  (0,0), (-1,-1), 0),
            ("RIGHTPADDING", (0,0), (-1,-1), 0),
        ]))
        return t

    # ── Page 1 continued: 3 uniform full-width bar charts ───────────────────────
    #
    # All dimensions kept as explicit Python floats so units never get confused.
    # matplotlib uses INCHES; reportlab Image() uses CM (via *cm).
    #
    # Page 1 height budget (A4 portrait, 26.7 cm usable):
    #   Header + title + HR  ≈ 3.5 cm
    #   Trend chart + stats  ≈ 4.5 cm  + 0.35 spacer
    #   Brand chart          ≈ 3.5 cm  + 0.3  spacer
    #   Application chart    ≈ 3.8 cm  + 0.3  spacer
    #   Product Cat chart    ≈ 3.0 cm  + 0.3  spacer
    #   ─────────────────────────────────────────────
    #   Total                ≈ 19.6 cm  ✓ fits on page 1

    # Fixed figure dimensions in INCHES (matplotlib)
    # Heights chosen so all 3 charts + trend fill the full page 1
    _FW  = 7.09   # full chart width  (18.0 cm / 2.54)
    _BH  = 1.97   # brand chart       (5.0 cm / 2.54)  — 4 bars
    _AH  = 2.56   # app chart         (6.5 cm / 2.54)  — up to 8 bars
    _TH  = 1.81   # type chart        (4.6 cm / 2.54)  — 3-5 bars

    # Corresponding display heights in reportlab units
    _BH_D = 5.0 * cm
    _AH_D = 6.5 * cm
    _TH_D = 4.6 * cm

    def _hbar_full(data, title, palette, fig_h_in, disp_h, left_m=0.12,
                   top_cap=None, data_py=None, py_year=None, cur_year=None):
        """Build a full-width horizontal bar chart Image for the PDF.
        When data_py is provided, renders grouped current vs prior-year bars."""
        import numpy as np

        s = data.sort_values()
        if top_cap:
            s = s.tail(top_cap)
        labels  = s.index.tolist()
        n       = len(labels)

        if isinstance(palette, dict):
            colours = [palette.get(l, RICH_PALETTE[i % len(RICH_PALETTE)])
                       for i, l in enumerate(labels)]
        else:
            colours = [RICH_PALETTE[i % len(RICH_PALETTE)] for i in range(n)]

        fig, ax = plt.subplots(figsize=(_FW, fig_h_in), facecolor="white")

        if data_py is not None and not data_py.empty:
            # Grouped bars: prior year (grey) left, current year (brand colour) right
            y_pos = np.arange(n)
            h = 0.38
            py_vals  = [data_py.get(l, 0) for l in labels]
            cur_vals = s.values
            ax.barh(y_pos - h/2, py_vals,  height=h, color="#aab4be", label=str(py_year))
            ax.barh(y_pos + h/2, cur_vals, height=h, color=colours,   label=str(cur_year))
            ax.set_yticks(y_pos)
            ax.set_yticklabels(labels, fontsize=8)
            all_max = max(data.max(), data_py.max() if not data_py.empty else 0)
            # Value labels on current bars only (right side)
            for i, v in enumerate(cur_vals):
                ax.text(v + all_max * 0.012, y_pos[i] + h/2,
                        f"£{v:,.0f}", va="center", ha="left", fontsize=7, color=brand_dark)
            ax.set_xlim(0, all_max * 1.22)
        else:
            # Single-year bars
            ax.barh(labels, s.values, color=colours, height=0.55)
            for bar in ax.patches:
                w = bar.get_width()
                ax.text(w + data.max() * 0.012,
                        bar.get_y() + bar.get_height() / 2,
                        f"£{w:,.0f}", va="center", ha="left", fontsize=7.5, color=brand_dark)
            ax.set_xlim(0, data.max() * 1.22)

        ax.xaxis.set_major_formatter(plt.FuncFormatter(lambda v, _: f"£{v:,.0f}"))
        ax.xaxis.set_major_locator(plt.MaxNLocator(nbins=5))
        ax.set_title(title, fontsize=10, fontweight="bold", color=brand_dark)
        ax.spines[["top", "right"]].set_visible(False)
        ax.tick_params(labelsize=8)
        ax.grid(axis="x", alpha=0.3)
        fig.subplots_adjust(left=left_m, right=0.97, top=0.87, bottom=0.14)
        b = _make_chart_image(fig, _FW, fig_h_in)
        plt.close(fig)
        return Image(io.BytesIO(b), width=doc.width, height=disp_h)

    # ── Sales by Brand ───────────────────────────────────────────────────────────
    if not brand_data.empty and brand_data.sum() > 0:
        story.append(Spacer(1, 0.3*cm))
        story.append(_hbar_full(
            brand_data, "Sales by Brand", BRAND_COLOURS, _BH, _BH_D, left_m=0.10,
            data_py=brand_data_py if not brand_data_py.empty else None,
            py_year=date_from.year - 1, cur_year=date_from.year,
        ))

    # ── Sales by Application ─────────────────────────────────────────────────────
    if not app_data.empty and app_data.sum() > 0:
        story.append(Spacer(1, 0.3*cm))
        story.append(_hbar_full(
            app_data, "Sales by Application",
            RICH_PALETTE, _AH, _AH_D, left_m=0.22, top_cap=8,
            data_py=app_data_py if not app_data_py.empty else None,
            py_year=date_from.year - 1, cur_year=date_from.year,
        ))

    # ── Sales by Product Category ────────────────────────────────────────────────
    if not itype_data.empty and itype_data.sum() > 0:
        story.append(Spacer(1, 0.3*cm))
        story.append(_hbar_full(
            itype_data, "Sales by Product Category",
            RICH_PALETTE, _TH, _TH_D, left_m=0.10,
            data_py=itype_data_py if not itype_data_py.empty else None,
            py_year=date_from.year - 1, cur_year=date_from.year,
        ))

    story.append(PageBreak())

    # ── Page 2: Top Accounts + Salesman + Products table ────────────────────────
    def page_header(page_num):
        story.append(Paragraph("TUNGALOY-NTK  |  Customer Account Report",
                                style("ph", fontSize=9, textColor=colors.grey, spaceAfter=2)))
        story.append(Paragraph(f"{customer_name}  ·  p{page_num}",
                                style("ph2", fontSize=12, fontName="Helvetica-Bold",
                                      textColor=MDARK, spaceAfter=4)))
        story.append(HRFlowable(width="100%", thickness=1, color=MRED, spaceAfter=10))

    page_header(2)

    p3_charts = []

    # Page 2 chart cell: same width as page 1, taller to show more bars
    P2_W = doc.width       # full content width = 18.0 cm
    P2_H = 5.0 * cm       # balanced — leaves room for 2 × 10-row tables
    P2_FIG_W = 7.09       # inches  (18.0 cm / 2.54)
    P2_FIG_H = 1.97       # inches  (5.0 cm / 2.54)

    # Top Accounts horizontal bar (with optional prior-year grey bars)
    if not top_accounts.empty:
        import numpy as np
        acct_labels = [str(n)[:22] for n in top_accounts["Customer"].tolist()[::-1]]
        acct_vals   = top_accounts["Sales £"].tolist()[::-1]
        colours     = [brand_red if i == len(acct_labels)-1 else brand_dark
                       for i in range(len(acct_labels))]

        # Build prior-year lookup by customer name
        py_acct_lookup = {}
        if df_py is not None and not df_py.empty and "Customer" in df_py.columns and "Sales £" in df_py.columns:
            _dpy = df_py.copy()
            _dpy["Sales £"] = pd.to_numeric(_dpy["Sales £"], errors="coerce").fillna(0)
            py_acct_lookup = _dpy.groupby("Customer")["Sales £"].sum().to_dict()

        fig4, ax4 = plt.subplots(figsize=(P2_FIG_W, P2_FIG_H), facecolor="white")

        if py_acct_lookup:
            # Grouped bars: grey (prior year) behind, coloured (current year) in front
            n     = len(acct_labels)
            y_pos = np.arange(n)
            h     = 0.38
            py_vals = [py_acct_lookup.get(lbl, 0) for lbl in acct_labels]
            ax4.barh(y_pos - h/2, py_vals,   height=h, color="#aab4be", label="Prior Year")
            ax4.barh(y_pos + h/2, acct_vals,  height=h, color=colours)
            ax4.set_yticks(y_pos)
            ax4.set_yticklabels(acct_labels)
            all_max = max(max(acct_vals), max(py_vals)) if py_vals else max(acct_vals)
            ax4.set_xlim(0, all_max * 1.15)
        else:
            ax4.barh(acct_labels, acct_vals, color=colours, height=0.6)

        ax4.xaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f"£{x:,.0f}"))
        ax4.set_title("")
        ax4.spines[["top","right"]].set_visible(False)
        ax4.tick_params(labelsize=7)
        ax4.grid(axis="x", alpha=0.3)
        ax4.xaxis.set_major_locator(plt.MaxNLocator(nbins=5))
        fig4.subplots_adjust(left=0.16, right=0.98, top=0.97, bottom=0.22)
        b = _make_chart_image(fig4, P2_FIG_W, P2_FIG_H)
        plt.close(fig4)
        p3_charts.append(Image(io.BytesIO(b), width=P2_W, height=P2_H))

    if p3_charts:
        # Top Accounts spans the full content width on its own
        t = _chart_row(p3_charts, [doc.width])
        if t:
            story.append(t)
            story.append(Spacer(1, 0.4*cm))

    # ── Prior-year lookups (used by helpers below) ───────────────────────────────
    py_sales_lookup = {}
    py_qty_lookup   = {}
    if df_py is not None and not df_py.empty and "Item Code" in df_py.columns:
        df_py["Sales £"] = pd.to_numeric(df_py.get("Sales £", 0), errors="coerce").fillna(0)
        py_sales_lookup  = df_py.groupby("Item Code")["Sales £"].sum().to_dict()
        if "Qty" in df_py.columns:
            df_py["Qty"] = pd.to_numeric(df_py.get("Qty", 0), errors="coerce").fillna(0)
            py_qty_lookup = df_py.groupby("Item Code")["Qty"].sum().to_dict()

    # ── Reusable table helpers (accessible for both page 2 and page 3) ───────────

    def _product_table(items_df, sort_by="sales"):
        """Return a styled product Table (does not append to story)."""
        has_py = bool(py_sales_lookup) if sort_by == "sales" else bool(py_qty_lookup)
        cur_yr = date_from.year
        pri_yr = date_from.year - 1

        if has_py:
            if sort_by == "sales":
                hdr = ["Item Code", "Description", "Brand",
                       f"Sales £ ({cur_yr})", f"Prior £ ({pri_yr})", "Var £", "Var %"]
            else:
                hdr = ["Item Code", "Description", "Brand",
                       f"Qty ({cur_yr})", f"Prior Qty ({pri_yr})", "Var Qty", "Var %"]
            col_w      = [2.2*cm, 5.3*cm, 1.8*cm, 2.5*cm, 2.5*cm, 2.0*cm, 1.7*cm]
            right_cols = (3, 6)
            var_col    = 6
        else:
            hdr        = ["Item Code", "Description", "Brand", "Type", "Sales £", "Qty"]
            col_w      = [2.5*cm, 6.5*cm, 2*cm, 2*cm, 2.5*cm, 2*cm]
            right_cols = (4, 5)
            var_col    = None

        rows      = [hdr]
        green_idx = []
        red_idx   = []

        for i, (_, row) in enumerate(items_df.iterrows(), start=1):
            ic        = row["Item Code"]
            brand_val = df[df["Item Code"] == ic]["Brand"].iloc[0]    if "Brand"     in df.columns and len(df[df["Item Code"] == ic]) else ""
            type_val  = df[df["Item Code"] == ic]["Item Type"].iloc[0] if "Item Type" in df.columns and len(df[df["Item Code"] == ic]) else ""
            qty_val   = df[df["Item Code"] == ic]["Qty"].sum()         if "Qty"       in df.columns else 0
            sales_val = row["Sales £"]

            if has_py and sort_by == "sales":
                py_v  = py_sales_lookup.get(ic, 0)
                var_v = sales_val - py_v
                var_p = (var_v / py_v * 100) if py_v else None
                var_s = f"{var_p:+.1f}%" if var_p is not None else "NEW"
                (green_idx if (var_p or 0) >= 0 else red_idx).append(i)
                rows.append([ic, str(row.get("Description",""))[:32],
                             str(brand_val) if pd.notna(brand_val) else "",
                             f"£{sales_val:,.0f}",
                             f"£{py_v:,.0f}" if py_v else "—",
                             f"£{var_v:+,.0f}", var_s])
            elif has_py and sort_by == "qty":
                cur_q = int(qty_val)
                py_q  = int(py_qty_lookup.get(ic, 0))
                var_q = cur_q - py_q
                var_p = (var_q / py_q * 100) if py_q else None
                var_s = f"{var_p:+.1f}%" if var_p is not None else "NEW"
                (green_idx if (var_p or 0) >= 0 else red_idx).append(i)
                rows.append([ic, str(row.get("Description",""))[:32],
                             str(brand_val) if pd.notna(brand_val) else "",
                             f"{cur_q:,}",
                             f"{py_q:,}" if py_q else "—",
                             f"{var_q:+,}", var_s])
            else:
                rows.append([ic, str(row.get("Description",""))[:38],
                             str(brand_val) if pd.notna(brand_val) else "",
                             str(type_val)  if pd.notna(type_val)  else "",
                             f"£{sales_val:,.0f}", f"{qty_val:,.0f}"])

        tbl_style = [
            ("BACKGROUND",    (0,0), (-1,0),  MDARK),
            ("TEXTCOLOR",     (0,0), (-1,0),  colors.white),
            ("FONTNAME",      (0,0), (-1,0),  "Helvetica-Bold"),
            ("FONTSIZE",      (0,0), (-1,-1), 8),
            ("ROWBACKGROUNDS",(0,1), (-1,-1), [colors.white, colors.HexColor("#f9f9f9")]),
            ("GRID",          (0,0), (-1,-1), 0.4, colors.HexColor("#dddddd")),
            ("ALIGN",         (right_cols[0],0), (right_cols[1],-1), "RIGHT"),
            ("TOPPADDING",    (0,0), (-1,-1), 4),
            ("BOTTOMPADDING", (0,0), (-1,-1), 4),
            ("LINEBELOW",     (0,0), (-1,0),  1.5, MRED),
        ]
        if var_col is not None:
            for r in green_idx:
                tbl_style.append(("TEXTCOLOR", (var_col,r), (var_col,r), colors.HexColor("#27ae60")))
            for r in red_idx:
                tbl_style.append(("TEXTCOLOR", (var_col,r), (var_col,r), colors.HexColor("#c0392b")))

        t = Table(rows, colWidths=col_w, splitByRow=0)
        t.setStyle(TableStyle(tbl_style))
        return t

    def _product_block(title, items_df, sort_by="sales", spacer_before=False):
        """Append a heading + table as a KeepTogether block (never splits across pages)."""
        tbl     = _product_table(items_df, sort_by=sort_by)
        heading = Paragraph(title, style("sh2", fontSize=11, fontName="Helvetica-Bold",
                                         textColor=MDARK, spaceBefore=4, spaceAfter=4))
        if spacer_before:
            story.append(Spacer(1, 0.5*cm))
        story.append(KeepTogether([heading, tbl]))

    # ── Initialise bottom datasets (populated inside if-block below) ─────────────
    bottom_items      = pd.DataFrame()
    bottom_by_qty     = pd.DataFrame()
    bottom_sales_title = "Bottom Products by Sales"
    bottom_qty_title   = "Bottom Products by Quantity"
    top_by_qty         = pd.DataFrame()

    # Top items detail table
    if not top_items.empty:

        # ── Build qty data ───────────────────────────────────────────────────────
        _all_by_qty = (df.groupby(["Item Code", "Description"])
                       .agg(**{"Sales £": ("Sales £", "sum"), "Qty": ("Qty", "sum")})
                       .reset_index()
                       .sort_values("Qty", ascending=False))
        top_by_qty = _all_by_qty.head(10) if "Qty" in df.columns else pd.DataFrame()

        # ── Build bottom (decline) datasets ──────────────────────────────────────
        if py_sales_lookup:
            _all_items["var_sales"] = _all_items["Item Code"].map(
                lambda ic: _all_items.loc[_all_items["Item Code"] == ic, "Sales £"].values[0]
                           - py_sales_lookup.get(ic, 0)
            )
            bottom_items = (_all_items[_all_items["Item Code"].map(py_sales_lookup).gt(0)]
                            .sort_values("var_sales", ascending=True)
                            .head(10)
                            .drop(columns=["var_sales"]))
            bottom_sales_title = "Biggest Sales Declines vs Prior Year"
        else:
            bottom_items = (_all_items[_all_items["Sales £"] > 0]
                            .tail(10).iloc[::-1].reset_index(drop=True))
            bottom_sales_title = "Bottom Products by Sales"

        if py_qty_lookup and not top_by_qty.empty:
            _all_by_qty["var_qty"] = _all_by_qty["Item Code"].map(
                lambda ic: _all_by_qty.loc[_all_by_qty["Item Code"] == ic, "Qty"].values[0]
                           - py_qty_lookup.get(ic, 0)
            )
            bottom_by_qty = (_all_by_qty[_all_by_qty["Item Code"].map(py_qty_lookup).gt(0)]
                             .sort_values("var_qty", ascending=True)
                             .head(10)
                             .drop(columns=["var_qty"]))
            bottom_qty_title = "Biggest Quantity Declines vs Prior Year"
        else:
            bottom_by_qty = (pd.DataFrame() if top_by_qty.empty else
                             _all_by_qty[_all_by_qty["Qty"] > 0]
                             .tail(10).iloc[::-1].reset_index(drop=True))
            bottom_qty_title = "Bottom Products by Quantity"

        # ── PAGE 2: Top Products by Sales + Top Products by Quantity ─────────────
        _product_block("Top Products by Sales vs Prior Year", top_items, sort_by="sales")

        if not top_by_qty.empty:
            _product_block("Top Products by Quantity vs Prior Year",
                           top_by_qty, sort_by="qty", spacer_before=True)

    # Footer helper
    def _append_footer():
        story.append(Spacer(1, 0.6*cm))
        story.append(HRFlowable(width="100%", thickness=0.5, color=colors.grey))
        fd = [[
            Paragraph(
                f"Prepared by Tungaloy-NTK  |  {datetime.now().strftime('%d %b %Y')}",
                style("ftl", fontSize=7, textColor=colors.grey, spaceBefore=4)),
            Paragraph(
                "<b>Tungaloy-NTK UK Ltd.</b>  |  Suite 3 Pioneer House, Mill Street, Cannock, WS11 0EF, UK  "
                "|  Tel: +44 121 4000 231  |  salesinfo@tungaloyuk.co.uk",
                style("ftr", fontSize=7, textColor=colors.grey,
                      alignment=TA_RIGHT, spaceBefore=4)),
        ]]
        ft = Table(fd, colWidths=[doc.width * 0.35, doc.width * 0.65])
        ft.setStyle(TableStyle([
            ("VALIGN",        (0,0), (-1,-1), "TOP"),
            ("LEFTPADDING",   (0,0), (-1,-1), 0),
            ("RIGHTPADDING",  (0,0), (-1,-1), 0),
            ("TOPPADDING",    (0,0), (-1,-1), 3),
            ("BOTTOMPADDING", (0,0), (-1,-1), 0),
        ]))
        story.append(ft)

    # ── PAGE 3: Biggest Sales Declines + Biggest Quantity Declines ───────────────
    _has_page3 = not top_items.empty and (not bottom_items.empty or not bottom_by_qty.empty)

    if _has_page3:
        story.append(PageBreak())
        page_header(3)

        if not bottom_items.empty:
            _product_block(bottom_sales_title, bottom_items, sort_by="sales")

        if not bottom_by_qty.empty:
            _product_block(bottom_qty_title, bottom_by_qty, sort_by="qty", spacer_before=True)

        _append_footer()
    else:
        _append_footer()

    # ── Page 4: Consignment Stock Analysis (only when consignment filter active) ─
    if consignment_data:
        import numpy as np
        _cd = consignment_data
        story.append(PageBreak())
        _pg = 4 if _has_page3 else 3
        page_header(_pg)

        story.append(Paragraph("Consignment Stock Analysis",
                                style("cons_title", fontSize=14, fontName="Helvetica-Bold",
                                      textColor=MDARK, spaceAfter=10)))

        # ── Sales comparison: Consignment vs Non-Consignment bar chart ──────
        _cons_s = _cd['cons_sales']
        _non_s  = _cd['non_sales']
        _tot_s  = _cd['total_sales']
        _cons_pct = (_cons_s / _tot_s * 100) if _tot_s else 0

        fig_c, ax_c = plt.subplots(figsize=(7.09, 1.6), facecolor="white")
        bars = ax_c.barh(
            ["Non-Consignment", "Consignment"],
            [_non_s, _cons_s],
            color=["#aab4be", brand_red], height=0.5
        )
        for bar in bars:
            w = bar.get_width()
            pct = (w / _tot_s * 100) if _tot_s else 0
            ax_c.text(w + _tot_s * 0.01, bar.get_y() + bar.get_height() / 2,
                      f"£{w:,.0f}  ({pct:.1f}%)", va="center", fontsize=9,
                      fontweight="bold", color=brand_dark)
        ax_c.set_xlim(0, max(_cons_s, _non_s) * 1.35)
        ax_c.xaxis.set_major_formatter(plt.FuncFormatter(lambda v, _: f"£{v:,.0f}"))
        ax_c.set_title("Consignment vs Non-Consignment Sales", fontsize=11,
                        fontweight="bold", color=brand_dark)
        ax_c.spines[["top", "right"]].set_visible(False)
        ax_c.grid(axis="x", alpha=0.3)
        ax_c.tick_params(labelsize=9)
        fig_c.subplots_adjust(left=0.18, right=0.97, top=0.82, bottom=0.18)
        _comp_bytes = _make_chart_image(fig_c, 7.09, 1.6)
        plt.close(fig_c)
        story.append(Image(io.BytesIO(_comp_bytes), width=doc.width, height=4.0*cm))
        story.append(Spacer(1, 0.3*cm))

        # Summary line
        story.append(Paragraph(
            f"<b>{_cd['cons_custs']}</b> consignment accounts out of "
            f"<b>{_cd['total_custs']}</b> total  |  "
            f"Consignment = <b>{_cons_pct:.1f}%</b> of all sales",
            style("cons_sum", fontSize=9, textColor=colors.grey, spaceAfter=10)
        ))

        # ── Top 10 Consignment Items ─────────────────────────────────────────
        _CW = {'1B','1C','BM','C2','C3','D5','D8','D9','DP','E1','E5','E7','E8','E9','EB','S1','SW'}
        _df_cw  = df[df["WHS"].isin(_CW)]  if "WHS" in df.columns else df
        _df_ncw = df[~df["WHS"].isin(_CW)] if "WHS" in df.columns else pd.DataFrame()

        _TBL_STYLE = TableStyle([
            ("BACKGROUND",    (0,0), (-1,0), MDARK),
            ("TEXTCOLOR",     (0,0), (-1,0), colors.white),
            ("FONTSIZE",      (0,0), (-1,-1), 8),
            ("FONTNAME",      (0,0), (-1,0), "Helvetica-Bold"),
            ("BOTTOMPADDING", (0,0), (-1,0), 6),
            ("TOPPADDING",    (0,0), (-1,0), 6),
            ("BOTTOMPADDING", (0,1), (-1,-1), 3),
            ("TOPPADDING",    (0,1), (-1,-1), 3),
            ("ROWBACKGROUNDS",(0,1), (-1,-1), [colors.white, MGREY]),
            ("ALIGN",         (3,0), (-1,-1), "RIGHT"),
            ("ALIGN",         (0,0), (0,-1), "CENTER"),
            ("LINEBELOW",     (0,0), (-1,0), 1.5, MRED),
            ("LINEBELOW",     (0,-1),(-1,-1), 0.5, colors.grey),
        ])
        _TBL_WIDTHS = [0.6*cm, 2.5*cm, 6.5*cm, 3.0*cm, 2.5*cm, 2.8*cm]

        if "Item Code" in _df_cw.columns and "Sales £" in _df_cw.columns:
            _top_val = (_df_cw.groupby(["Item Code", "Description"])
                        .agg({"Sales £": "sum", "Qty": "sum", "Invoice": "count"})
                        .rename(columns={"Invoice": "Trans"})
                        .reset_index()
                        .sort_values("Sales £", ascending=False)
                        .head(10))
            if not _top_val.empty:
                story.append(Paragraph("Top 10 Consignment Items",
                                        style("ti_hdr", fontSize=11, fontName="Helvetica-Bold",
                                              textColor=MDARK, spaceAfter=6)))
                _hdr = ["#", "Item Code", "Description", "Sales £", "Qty", "Trans"]
                _rows = [_hdr]
                for i, (_, r) in enumerate(_top_val.iterrows(), 1):
                    _rows.append([
                        str(i),
                        str(r["Item Code"]),
                        str(r["Description"])[:32],
                        f"£{r['Sales £']:,.0f}",
                        f"{r['Qty']:,.0f}",
                        f"{int(r['Trans']):,}",
                    ])
                _t = Table(_rows, colWidths=_TBL_WIDTHS)
                _t.setStyle(_TBL_STYLE)
                story.append(_t)
                story.append(Spacer(1, 0.5*cm))

        # ── Top 10 Non-Consignment Items ────────────────────────────────────
        if not _df_ncw.empty and "Item Code" in _df_ncw.columns and "Sales £" in _df_ncw.columns:
            _top_non = (_df_ncw.groupby(["Item Code", "Description"])
                        .agg({"Sales £": "sum", "Qty": "sum", "Invoice": "count"})
                        .rename(columns={"Invoice": "Trans"})
                        .reset_index()
                        .sort_values("Sales £", ascending=False)
                        .head(10))
            if not _top_non.empty:
                story.append(Paragraph("Top 10 Non-Consignment Items",
                                        style("tn_hdr", fontSize=11, fontName="Helvetica-Bold",
                                              textColor=MDARK, spaceAfter=6)))
                _hdr2 = ["#", "Item Code", "Description", "Sales £", "Qty", "Trans"]
                _rows2 = [_hdr2]
                for i, (_, r) in enumerate(_top_non.iterrows(), 1):
                    _rows2.append([
                        str(i),
                        str(r["Item Code"]),
                        str(r["Description"])[:32],
                        f"£{r['Sales £']:,.0f}",
                        f"{r['Qty']:,.0f}",
                        f"{int(r['Trans']):,}",
                    ])
                _t2 = Table(_rows2, colWidths=_TBL_WIDTHS)
                _t2.setStyle(_TBL_STYLE)
                story.append(_t2)

        story.append(Spacer(1, 1*cm))
        _append_footer()

    doc.build(story)
    return buf.getvalue()


# ── Main reports page ────────────────────────────────────────────────────────────
def page_reports():
    st.markdown('<div class="section-header">Sales Reports</div>', unsafe_allow_html=True)

    user_id = st.session_state["user_id"]
    role    = st.session_state["role"]

    opts = db.get_report_filter_options(user_id, role)

    # ── Filter panel — always visible ────────────────────────────────────────────
    st.markdown("#### 🔍 Filters")

    db_min, db_max = db.get_league_date_range()
    default_from = date.fromisoformat(db_min[:10]) if db_min else date(2025, 1, 1)
    default_to   = date.fromisoformat(db_max[:10]) if db_max else date.today()

    col1, col2, col3, col4 = st.columns([2, 2, 2, 2])
    date_from   = col1.date_input("Date From", value=default_from, key="rpt_date_from")
    date_to     = col2.date_input("Date To",   value=default_to,   key="rpt_date_to")
    item_search = col3.text_input("Item (code or description)", placeholder="Search…", key="rpt_item")

    # Customer multiselect — tick the accounts you want
    cust_options = {f"{c['name']} ({c['code']})": c["code"] for c in opts["customers"]}

    # When consignment filter is active, restrict dropdown to consignment accounts only
    _cons_active = st.session_state.get("rpt_consignment", False)
    if _cons_active:
        try:
            with db.get_conn() as conn:
                _cons_code_set = {r['customer_code'] for r in conn.execute(
                    "SELECT DISTINCT customer_code FROM consignment_warehouses "
                    "WHERE is_active=1 AND customer_code IS NOT NULL"
                ).fetchall()}
        except Exception:
            _cons_code_set = set()
        if _cons_code_set:
            cust_options = {k: v for k, v in cust_options.items() if v in _cons_code_set}
            # Clear any previously selected accounts that are no longer valid
            if "rpt_custs_multi" in st.session_state:
                _valid = set(cust_options.keys())
                st.session_state["rpt_custs_multi"] = [
                    l for l in st.session_state["rpt_custs_multi"] if l in _valid
                ]

    # When "No spend in 2026" is active, restrict dropdown to zero-2026 accounts
    _no_spend_active = st.session_state.get("rpt_no_spend_2026", False)
    if _no_spend_active:
        try:
            with db.get_conn() as conn:
                _spent_code_set = {r['customer_code'] for r in conn.execute(
                    "SELECT DISTINCT customer_code FROM sales WHERE year=2026 AND sales_val > 0"
                ).fetchall()}
        except Exception:
            _spent_code_set = set()
        cust_options = {k: v for k, v in cust_options.items() if v not in _spent_code_set}
        if "rpt_custs_multi" in st.session_state:
            _valid = set(cust_options.keys())
            st.session_state["rpt_custs_multi"] = [
                l for l in st.session_state["rpt_custs_multi"] if l in _valid
            ]

    selected_cust_labels = col4.multiselect(
        "Accounts (consignment only)" if _cons_active else "Accounts (leave blank = all)",
        options=list(cust_options.keys()),
        key="rpt_custs_multi",
        placeholder="Select accounts…",
    )
    selected_customer_codes = [cust_options[l] for l in selected_cust_labels] or None

    def _chk_ms(container, label, key_chk, key_ms, options):
        """Checkbox toggles the multiselect. Empty multiselect = all data for that field."""
        enabled = container.checkbox(label, key=key_chk)
        if enabled and options:
            sel = container.multiselect(
                "_" + key_ms, options,
                placeholder="Select to filter, or leave blank for all",
                label_visibility="collapsed",
                key=key_ms,
            )
            return sel if sel else None
        return None

    # ── Row 2: Application / Salesman / SM Name ─────────────────────────────────
    col5, col6, col7, col8 = st.columns([2, 2, 2, 2])

    app_filter_val = _chk_ms(col5, "Application", "use_app", "rpt_app", opts["applications"])

    selected_sm_names = None
    rep_sel = []
    sm_sel  = []

    if role == "rep":
        col6.info("Showing your sales only")
        col7.info("Your SM names only")
    else:
        # Build name → [sm_names] map so reps with multiple territory codes appear once
        rep_sm_map_multi = {}
        for r in opts["reps"]:
            rep_sm_map_multi.setdefault(r["full_name"], []).append(r["sm_name"])
        rep_names    = sorted(rep_sm_map_multi.keys())
        sm_name_opts = sorted(opts["sm_names"])

        rep_sel = _chk_ms(col6, "Salesman", "use_rep", "rpt_rep",     rep_names)
        sm_sel  = _chk_ms(col7, "SM Name",  "use_sm",  "rpt_sm_name", sm_name_opts)

        if sm_sel:
            selected_sm_names = sm_sel
        elif rep_sel:
            # Collect ALL sm_names for each selected rep (handles multi-territory reps)
            selected_sm_names = [sm for n in rep_sel for sm in rep_sm_map_multi.get(n, [])]

    run_btn = col8.button("🔎 Run Report", type="primary", use_container_width=True)

    # ── Row 3: Customer enrichment filters ───────────────────────────────────────
    col9, col10, col11, col12 = st.columns([2, 2, 2, 2])

    region_filter_val      = _chk_ms(col9,  "Region",        "use_region",      "rpt_region",      opts.get("regions", []))
    cust_type_filter_val   = _chk_ms(col10, "Customer Type", "use_cust_type",   "rpt_cust_type",   opts.get("customer_types", []))
    post_area_filter_val   = _chk_ms(col11, "Post Area",     "use_post_area",   "rpt_post_area",   opts.get("post_areas", []))
    distributor_filter_val = _chk_ms(col12, "Distributor",   "use_distributor", "rpt_distributor", opts.get("distributors", []))

    # ── Row 4: Item enrichment filters ───────────────────────────────────────────
    col13, col14, col15, col16 = st.columns([2, 2, 3, 3])

    brand_filter_val     = _chk_ms(col13, "Brand",     "use_brand",     "rpt_brand",     opts.get("brands", []))
    item_type_filter_val = _chk_ms(col14, "Item Type", "use_item_type", "rpt_item_type", opts.get("item_types", []))
    consignment_only     = col15.checkbox("📦 Consignment Accounts Only", key="rpt_consignment",
                                          help="Filter to consignment warehouse accounts and show comparison stats")
    no_spend_2026        = col16.checkbox("❄️ No spend in 2026", key="rpt_no_spend_2026",
                                          help="Show only customers with zero sales recorded in 2026")

    chk1, chk2, chk3 = st.columns([2, 2, 2])
    hide_costs      = chk1.checkbox("🙈 Hide costs & margins (Cost £, GP £, GP %)", key="rpt_hide_costs")
    hide_address    = chk2.checkbox("📦 Hide Ship To Address", key="rpt_hide_address")
    compare_py      = chk3.checkbox("📅 Compare to same period last year", key="rpt_compare_py")

    if compare_py:
        py_from = date_from.replace(year=date_from.year - 1)
        py_to   = date_to.replace(year=date_to.year - 1)
        chk3.caption(f"vs {py_from.strftime('%d %b %Y')} – {py_to.strftime('%d %b %Y')}")

    # ── Apply consignment warehouse filter if ticked ─────────────────────────────
    _eff_customer_codes = selected_customer_codes
    if consignment_only:
        try:
            with db.get_conn() as conn:
                _cons_rows = conn.execute(
                    "SELECT DISTINCT customer_code FROM consignment_warehouses "
                    "WHERE is_active=1 AND customer_code IS NOT NULL"
                ).fetchall()
            _cons_codes = [r['customer_code'] for r in _cons_rows]
        except Exception:
            _cons_codes = []
        if _cons_codes:
            if _eff_customer_codes:
                _eff_customer_codes = [c for c in _eff_customer_codes if c in _cons_codes]
            else:
                _eff_customer_codes = _cons_codes
        else:
            st.warning("No consignment warehouse accounts configured.")

    # ── Apply "No spend in 2026" filter if ticked ───────────────────────────────
    if no_spend_2026:
        try:
            with db.get_conn() as conn:
                # Customers who DID spend in 2026 (any positive sale)
                _spent_rows = conn.execute(
                    "SELECT DISTINCT customer_code FROM sales WHERE year=2026 AND sales_val > 0"
                ).fetchall()
                _spent_2026 = {r['customer_code'] for r in _spent_rows}

                # All customer codes visible to this user
                _all_rows = conn.execute(
                    "SELECT DISTINCT customer_code FROM customers"
                ).fetchall()
                _all_codes = {r['customer_code'] for r in _all_rows}

            # Customers with NO 2026 spend = all - spent
            _no_spend_codes = list(_all_codes - _spent_2026)

            if _eff_customer_codes:
                _eff_customer_codes = [c for c in _eff_customer_codes if c in _no_spend_codes]
            else:
                _eff_customer_codes = _no_spend_codes

            if not _eff_customer_codes:
                st.warning("No customers match — either none have zero 2026 spend, "
                           "or your other filters exclude them all.")
        except Exception as e:
            st.error(f"Could not apply 'No spend in 2026' filter: {e}")

    st.divider()

    # ── Run query ────────────────────────────────────────────────────────────────
    if "rpt_results" not in st.session_state:
        st.session_state["rpt_results"] = None

    if run_btn:
        with st.spinner("Loading data…"):
            _query_kwargs = dict(
                user_id=user_id, role=role,
                sm_names=selected_sm_names,
                customer_codes=_eff_customer_codes,
                item_search=item_search.strip() or None,
                app_filter=app_filter_val,
                region_filter=region_filter_val,
                customer_type_filter=cust_type_filter_val,
                post_area_filter=post_area_filter_val,
                distributor_filter=distributor_filter_val,
                brand_filter=brand_filter_val,
                item_type_filter=item_type_filter_val,
            )
            rows = db.get_report_data(date_from=date_from, date_to=date_to, **_query_kwargs)
            # Consignment comparison: also fetch ALL accounts (same filters minus consignment)
            if consignment_only:
                _all_kwargs = dict(_query_kwargs)
                _all_kwargs['customer_codes'] = selected_customer_codes  # original, no consignment filter
                all_rows = db.get_report_data(date_from=date_from, date_to=date_to, **_all_kwargs)
            else:
                all_rows = None
            # Prior year data
            if compare_py:
                py_from = date_from.replace(year=date_from.year - 1)
                py_to   = date_to.replace(year=date_to.year - 1)
                py_rows = db.get_report_data(date_from=py_from, date_to=py_to, **_query_kwargs)
            else:
                py_rows = None
        st.session_state["rpt_results"]    = rows
        st.session_state["rpt_py_results"] = py_rows
        st.session_state["rpt_all_results"] = all_rows
        parts = [f"{date_from.strftime('%d %b %Y')} – {date_to.strftime('%d %b %Y')}"]
        if role != "rep":
            if selected_sm_names:
                parts.append(f"SM: {', '.join(selected_sm_names)}")
        if selected_cust_labels:
            if len(selected_cust_labels) <= 3:
                parts.append(", ".join(selected_cust_labels))
            else:
                parts.append(f"{len(selected_cust_labels)} accounts selected")
        def _fmt(v):
            return ", ".join(v) if isinstance(v, list) else str(v)

        if app_filter_val:
            parts.append(_fmt(app_filter_val))
        if region_filter_val:
            parts.append(f"Region: {_fmt(region_filter_val)}")
        if cust_type_filter_val:
            parts.append(f"Type: {_fmt(cust_type_filter_val)}")
        if post_area_filter_val:
            parts.append(f"Post Area: {_fmt(post_area_filter_val)}")
        if distributor_filter_val:
            parts.append(f"Distributor: {_fmt(distributor_filter_val)}")
        if brand_filter_val:
            parts.append(f"Brand: {_fmt(brand_filter_val)}")
        if item_type_filter_val:
            parts.append(f"Item Type: {_fmt(item_type_filter_val)}")
        if consignment_only:
            parts.append("📦 Consignment Accounts Only")
        if no_spend_2026:
            parts.append("❄️ No spend in 2026")
        st.session_state["rpt_subtitle"] = " | ".join(parts)

    rows = st.session_state.get("rpt_results")
    if rows is None:
        st.info("Set your filters above and click **Run Report**.")
        return

    if not rows:
        st.warning("No data found for the selected filters.")
        return

    df = pd.DataFrame(rows)
    for c in ["Sales £", "Cost £", "GP £", "Qty"]:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)
    if "GP %" in df.columns:
        df["GP %"] = pd.to_numeric(df["GP %"], errors="coerce")

    py_rows = st.session_state.get("rpt_py_results")

    # Auto-fetch prior year data if compare is ticked but data not yet loaded
    if compare_py and not py_rows:
        py_from = date_from.replace(year=date_from.year - 1)
        py_to   = date_to.replace(year=date_to.year - 1)
        with st.spinner("Loading prior year data…"):
            py_rows = db.get_report_data(
                user_id=user_id, role=role,
                date_from=py_from, date_to=py_to,
                sm_names=selected_sm_names,
                customer_codes=_eff_customer_codes,
                item_search=item_search.strip() or None,
                app_filter=app_filter_val,
                region_filter=region_filter_val,
                customer_type_filter=cust_type_filter_val,
                post_area_filter=post_area_filter_val,
                distributor_filter=distributor_filter_val,
                brand_filter=brand_filter_val,
                item_type_filter=item_type_filter_val,
            )
            st.session_state["rpt_py_results"] = py_rows

    df_py = None
    if compare_py and py_rows:
        df_py = pd.DataFrame(py_rows)
        for c in ["Sales £", "GP £", "Qty"]:
            if c in df_py.columns:
                df_py[c] = pd.to_numeric(df_py[c], errors="coerce").fillna(0)

    # ── View mode ────────────────────────────────────────────────────────────────
    st.markdown("---")
    view_mode = st.radio(
        "Report view",
        ["📋 By Transaction", "🏢 By Account (Total Sales)", "🔩 By Item"],
        horizontal=True, key="rpt_view_mode",
    )

    def _recalc_gp_pct(d):
        d = d.copy()
        d["GP %"] = d.apply(
            lambda r: round(r["GP £"] / r["Sales £"] * 100, 1) if r["Sales £"] else None, axis=1
        )
        return d

    def _add_comparison(current: pd.DataFrame, prior: pd.DataFrame,
                         join_cols: list) -> pd.DataFrame:
        """Merge current and prior year totals, add Growth £ and Growth % columns."""
        py_agg = prior.groupby(join_cols, dropna=False).agg({"Sales £": "sum"}).reset_index()
        py_agg = py_agg.rename(columns={"Sales £": "Prior Year £"})
        merged = current.merge(py_agg, on=join_cols, how="left")
        merged["Prior Year £"] = merged["Prior Year £"].fillna(0)
        merged["Growth £"] = merged["Sales £"] - merged["Prior Year £"]
        merged["Growth %"] = merged.apply(
            lambda r: round(r["Growth £"] / r["Prior Year £"] * 100, 1)
            if r["Prior Year £"] else None, axis=1
        )
        return merged

    if view_mode == "🏢 By Account (Total Sales)":
        group_cols = ["Cust Code", "Customer", "Salesman", "SM Name"]
        if not hide_address:
            group_cols.insert(2, "Ship To Address")
        group_cols = [c for c in group_cols if c in df.columns]
        agg_cols = {c: "sum" for c in ["Sales £", "Cost £", "GP £", "Qty"] if c in df.columns}
        display_df = df.groupby(group_cols, dropna=False).agg(agg_cols).reset_index()
        if "GP £" in display_df.columns:
            display_df = _recalc_gp_pct(display_df)
        if df_py is not None:
            join_cols = [c for c in ["Cust Code", "Customer"] if c in group_cols]
            display_df = _add_comparison(display_df, df_py, join_cols)
        display_df = display_df.sort_values("Sales £", ascending=False)
        report_title = "Sales Report — By Account"

    elif view_mode == "🔩 By Item":
        group_cols = ["Item Code", "Description", "Application"]
        group_cols = [c for c in group_cols if c in df.columns]
        agg_cols = {c: "sum" for c in ["Sales £", "Cost £", "GP £", "Qty"] if c in df.columns}
        display_df = df.groupby(group_cols, dropna=False).agg(agg_cols).reset_index()
        if "GP £" in display_df.columns:
            display_df = _recalc_gp_pct(display_df)
        if df_py is not None:
            join_cols = [c for c in ["Item Code"] if c in group_cols]
            display_df = _add_comparison(display_df, df_py, join_cols)
        display_df = display_df.sort_values("Sales £", ascending=False)
        report_title = "Sales Report — By Item"

    else:
        display_df = df.copy()
        report_title = "Sales Report — By Transaction"

    # ── Strip columns based on options ───────────────────────────────────────────
    COST_COLS = ["Cost £", "GP £", "GP %"]
    total_gp    = display_df["GP £"].sum() if "GP £" in display_df.columns else 0
    total_sales = display_df["Sales £"].sum()

    if hide_costs:
        display_df = display_df.drop(columns=[c for c in COST_COLS if c in display_df.columns])
        report_title += " (Sales Only)"
    if hide_address and "Ship To Address" in display_df.columns:
        display_df = display_df.drop(columns=["Ship To Address"])

    # ── Hide columns whose checkbox is unticked ───────────────────────────────────
    chk_col_map = {
        "use_app":         "Application",
        "use_rep":         "Salesman",
        "use_sm":          "SM Name",
        "use_region":      "Region",
        "use_cust_type":   "Customer Type",
        "use_distributor": "Distributor",
        "use_brand":       "Brand",
        "use_item_type":   "Item Type",
    }
    for chk_key, col_name in chk_col_map.items():
        if not st.session_state.get(chk_key, False) and col_name in display_df.columns:
            display_df = display_df.drop(columns=[col_name])

    # ── Summary metrics ──────────────────────────────────────────────────────────
    py_total_sales = df_py["Sales £"].sum() if df_py is not None and not df_py.empty else None
    growth_val = (total_sales - py_total_sales) if py_total_sales is not None else None
    growth_pct = (growth_val / py_total_sales * 100) if py_total_sales else None

    if hide_costs:
        if py_total_sales is not None:
            m1, m2, m3, m4, m5 = st.columns(5)
            m1.metric("Rows",            f"{len(display_df):,}")
            m2.metric("Total Sales",     f"£{total_sales:,.0f}")
            m3.metric("Prior Year Sales",f"£{py_total_sales:,.0f}")
            m4.metric("Growth £",        f"£{growth_val:,.0f}",
                      delta=f"{growth_pct:+.1f}%" if growth_pct is not None else None)
            m5.metric("Customers",       df["Cust Code"].nunique())
        else:
            m1, m2, m3 = st.columns(3)
            m1.metric("Rows",        f"{len(display_df):,}")
            m2.metric("Total Sales", f"£{total_sales:,.0f}")
            m3.metric("Customers",   df["Cust Code"].nunique())
    else:
        if py_total_sales is not None:
            m1, m2, m3, m4, m5, m6 = st.columns(6)
            m1.metric("Total Sales",     f"£{total_sales:,.0f}")
            m2.metric("Prior Year Sales",f"£{py_total_sales:,.0f}")
            m3.metric("Growth £",        f"£{growth_val:,.0f}",
                      delta=f"{growth_pct:+.1f}%" if growth_pct is not None else None)
            m4.metric("Total GP",        f"£{total_gp:,.0f}")
            m5.metric("Overall GP %",    f"{total_gp/total_sales*100:.1f}%" if total_sales else "—")
            m6.metric("Customers",       df["Cust Code"].nunique())
        else:
            m1, m2, m3, m4, m5 = st.columns(5)
            m1.metric("Rows",        f"{len(display_df):,}")
            m2.metric("Total Sales", f"£{total_sales:,.0f}")
            m3.metric("Total GP",    f"£{total_gp:,.0f}")
            m4.metric("Overall GP %", f"{total_gp/total_sales*100:.1f}%" if total_sales else "—")
            m5.metric("Customers",   df["Cust Code"].nunique())

    # ── Consignment comparison panel (split by whs_code) ────────────────────────
    _CONS_WHS = {'1B','1C','BM','C2','C3','D5','D8','D9','DP','E1','E5','E7','E8','E9','EB','S1','SW'}
    if consignment_only and "WHS" in df.columns:
        _df_cons = df[df["WHS"].isin(_CONS_WHS)]
        _df_non  = df[~df["WHS"].isin(_CONS_WHS)]

        _cons_sales = _df_cons["Sales £"].sum()
        _non_sales  = _df_non["Sales £"].sum()
        _tot        = _cons_sales + _non_sales
        _cons_pct   = (_cons_sales / _tot * 100) if _tot else 0
        _cons_items = _df_cons["Item Code"].nunique() if "Item Code" in _df_cons.columns else 0
        _non_items  = _df_non["Item Code"].nunique() if "Item Code" in _df_non.columns else 0

        st.markdown("#### Consignment vs Non-Consignment Sales")
        cx1, cx2, cx3, cx4 = st.columns(4)
        cx1.metric("Consignment WHS Sales",     f"£{_cons_sales:,.0f}",
                   delta=f"{_cons_pct:.1f}% of total")
        cx2.metric("Non-Consignment Sales",     f"£{_non_sales:,.0f}",
                   delta=f"{100 - _cons_pct:.1f}% of total")
        cx3.metric("Total Sales",               f"£{_tot:,.0f}")
        cx4.metric("Unique Items",              f"{_cons_items} consign / {_non_items} other")
        st.markdown("---")

    # ── Download buttons ─────────────────────────────────────────────────────────
    subtitle = st.session_state.get("rpt_subtitle", "")

    # Build a descriptive filename from the active filters
    import re as _re, calendar as _cal
    def _strip_code(lbl):
        """Remove only the trailing (CUSTOMER_CODE) from a label."""
        return _re.sub(r'\s*\([^)]*\)\s*$', '', lbl).strip()

    _m_from = _cal.month_abbr[date_from.month].upper()
    _m_to   = _cal.month_abbr[date_to.month].upper()
    _date_part = f"{_m_from}-{_m_to}" if _m_from != _m_to else _m_from
    _yr_part = f"{date_from.year % 100}"
    if compare_py:
        _yr_part = f"{date_from.year % 100}v{(date_from.year - 1) % 100}"

    if selected_cust_labels and len(selected_cust_labels) <= 3:
        _subj = "_".join(_strip_code(l).replace(" ", "_") for l in selected_cust_labels)
    elif role != "rep" and rep_sel:
        _subj = "_".join(r.replace(" ", "_") for r in rep_sel[:2])
    elif role != "rep" and sm_sel:
        _subj = "_".join(s.replace(" ", "_") for s in sm_sel[:2])
    elif consignment_only:
        _subj = "Consignment_Accounts"
    elif no_spend_2026:
        _subj = "No_Spend_2026"
    elif region_filter_val:
        _subj = "Region_" + ("_".join(region_filter_val) if isinstance(region_filter_val, list) else str(region_filter_val))
    else:
        _subj = "All_Accounts"
    _subj = _re.sub(r'[^\w\s_-]', '', _subj).strip()[:60]

    fname_base = f"Customer_sales_report_{_subj}_{_date_part}_{_yr_part}"

    dl1, dl2, dl3, _ = st.columns([1, 1, 1, 3])

    with dl1:
        excel_bytes = _build_excel(display_df, report_title, subtitle)
        st.download_button(
            "📥 Download Excel",
            data=excel_bytes,
            file_name=f"{fname_base}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
        )

    with dl2:
        try:
            pdf_bytes = _build_pdf(display_df, report_title, subtitle)
            st.download_button(
                "📄 Download PDF",
                data=pdf_bytes,
                file_name=f"{fname_base}.pdf",
                mime="application/pdf",
            )
        except ImportError:
            st.warning("PDF requires `reportlab` — run: `pip install reportlab`")

    with dl3:
        try:
            if selected_cust_labels and len(selected_cust_labels) == 1:
                # Single account — use account name only
                cust_display = _strip_code(selected_cust_labels[0])
            elif selected_cust_labels and len(selected_cust_labels) <= 5:
                # 2–5 accounts — list the names directly
                cust_display = " / ".join(_strip_code(lbl) for lbl in selected_cust_labels)
            else:
                # No specific accounts, or 6+ — describe the active filters
                filter_parts = []
                # Salesman / SM Name filters — shown first so the rep name is prominent
                if role != "rep":
                    if rep_sel:
                        filter_parts.append(", ".join(rep_sel))
                    elif sm_sel:
                        filter_parts.append(", ".join(sm_sel))
                if region_filter_val:
                    filter_parts.append(f"Region: {region_filter_val}")
                if cust_type_filter_val:
                    filter_parts.append(cust_type_filter_val)
                if post_area_filter_val:
                    filter_parts.append(f"Post Area: {post_area_filter_val}")
                if distributor_filter_val:
                    filter_parts.append(f"Distributor: {distributor_filter_val}")
                if brand_filter_val:
                    filter_parts.append(f"Brand: {brand_filter_val}")
                if item_type_filter_val:
                    filter_parts.append(f"Item Type: {item_type_filter_val}")
                if selected_cust_labels and len(selected_cust_labels) > 5:
                    filter_parts.insert(0, f"{len(selected_cust_labels)} Selected Accounts")
                if filter_parts:
                    n = display_df["Customer"].nunique() if "Customer" in display_df.columns else ""
                    cust_display = "  |  ".join(filter_parts) + (f" ({n} customers)" if n else "")
                else:
                    n = display_df["Customer"].nunique() if "Customer" in display_df.columns else ""
                    cust_display = f"All Accounts Summary" + (f" ({n} customers)" if n else "")

            rep = db.get_user_by_id(user_id)
            prepared_by = rep["full_name"] if rep else "Tungaloy-NTK"
            # Build consignment comparison data for the PDF (whs_code level)
            _cons_pdf_data = None
            if consignment_only and "WHS" in df.columns:
                _pdf_cons = df[df["WHS"].isin(_CONS_WHS)]
                _pdf_non  = df[~df["WHS"].isin(_CONS_WHS)]
                _cons_pdf_data = {
                    "cons_sales":  _pdf_cons["Sales £"].sum(),
                    "non_sales":   _pdf_non["Sales £"].sum(),
                    "total_sales": total_sales,
                    "cons_custs":  _pdf_cons["Cust Code"].nunique() if "Cust Code" in df.columns else 0,
                    "total_custs": df["Cust Code"].nunique() if "Cust Code" in df.columns else 0,
                }
            dash_bytes = _build_dashboard_pdf(
                df, cust_display, date_from, date_to, prepared_by,
                hide_costs=hide_costs,
                df_py=df_py if compare_py else None,
                consignment_data=_cons_pdf_data,
            )
            st.download_button(
                "📊 Dashboard PDF",
                data=dash_bytes,
                file_name=f"{fname_base}.pdf",
                mime="application/pdf",
            )
        except Exception as e:
            st.warning(f"Dashboard error: {e}")

    # ── Data table ───────────────────────────────────────────────────────────────
    st.markdown(f"**{len(display_df):,} records** | {subtitle}")

    col_config = {
        "Sales £":     st.column_config.NumberColumn("Sales £",     format="£%.0f"),
        "Cost £":      st.column_config.NumberColumn("Cost £",      format="£%.0f"),
        "GP £":        st.column_config.NumberColumn("GP £",        format="£%.0f"),
        "GP %":        st.column_config.NumberColumn("GP %",        format="%.1f%%"),
        "Qty":         st.column_config.NumberColumn("Qty",         format="%,.0f"),
        "Prior Year £":st.column_config.NumberColumn("Prior Year £",format="£%.0f"),
        "Growth £":    st.column_config.NumberColumn("Growth £",    format="£%.0f"),
        "Growth %":    st.column_config.NumberColumn("Growth %",    format="%.1f%%"),
    }
    st.dataframe(display_df, use_container_width=True, hide_index=True,
                 column_config=col_config)


def render_page():
    page_reports()
