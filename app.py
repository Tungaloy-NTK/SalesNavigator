import streamlit as st
import pandas as pd
import plotly.graph_objects as go
from datetime import date, datetime
import database as db
import auth
import data_import as di
import alert_engine as ae
import test_tools
import marketing_leads
import visit_planner
import news_board
import league_tables
import crm
import email_campaigns
import reports
import expenses
import insert_converter

st.set_page_config(
    page_title="Tungaloy-NTK Sales Navigator",
    page_icon="🎯",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── Custom CSS ──────────────────────────────────────────────────────────────────
st.markdown("""
<style>
[data-testid="stSidebar"] {
    background: #1a1a2e !important;
}
[data-testid="stSidebar"] * {
    color: #ffffff !important;
}
[data-testid="stSidebar"] .stRadio label {
    color: #ffffff !important;
}
.metric-card {
    background: #f8f9fa;
    border-radius: 8px;
    padding: 16px 20px;
    border-left: 4px solid #c0392b;
    margin-bottom: 8px;
}
.alert-red    { color: #c0392b; font-weight: bold; }
.alert-orange { color: #e67e22; font-weight: bold; }
.alert-yellow { color: #d4a017; font-weight: bold; }
.alert-green  { color: #27ae60; font-weight: bold; }
.status-badge {
    display: inline-block;
    padding: 3px 10px;
    border-radius: 12px;
    font-size: 12px;
    font-weight: bold;
    color: white;
}
.badge-red    { background:#c0392b; }
.badge-orange { background:#e67e22; }
.badge-yellow { background:#d4a017; }
.badge-green  { background:#27ae60; }
.section-header {
    font-size: 22px;
    font-weight: 700;
    color: #1a1a2e;
    border-bottom: 2px solid #c0392b;
    padding-bottom: 6px;
    margin-bottom: 16px;
}
/* Sidebar nav buttons — all grey by default */
[data-testid="stSidebar"] .stButton > button {
    background-color: #2e2e42 !important;
    border: 1px solid #3d3d57 !important;
    color: #ffffff !important;
}
[data-testid="stSidebar"] .stButton > button:hover {
    background-color: #3d3d57 !important;
    border-color: #4d4d67 !important;
}
/* Active tab button — red */
[data-testid="stSidebar"] .stButton > button[kind="primary"],
[data-testid="stSidebar"] .stButton > button[data-testid="baseButton-primary"] {
    background-color: #c0392b !important;
    border-color: #c0392b !important;
}
[data-testid="stSidebar"] .stButton > button[kind="primary"]:hover,
[data-testid="stSidebar"] .stButton > button[data-testid="baseButton-primary"]:hover {
    background-color: #a93226 !important;
    border-color: #a93226 !important;
}
/* Delete buttons — subtle grey, red on hover */
button[kind="secondary"]:has(> div > p:only-child) {
    font-size: 13px;
}
/* Consistent form save buttons */
.stFormSubmitButton > button {
    background-color: #1a1a2e !important;
    color: #ffffff !important;
    border-radius: 6px !important;
}
.stFormSubmitButton > button:hover {
    background-color: #c0392b !important;
}
/* Popover (delete confirmations) */
[data-testid="stPopover"] button[kind="primary"] {
    background-color: #c0392b !important;
    border-color: #c0392b !important;
    font-size: 12px !important;
}
/* Tooltip help icons */
.stTooltipIcon { opacity: 0.5; }
.stTooltipIcon:hover { opacity: 1; }
/* Download buttons — subtle */
.stDownloadButton > button {
    background-color: #f0f2f6 !important;
    color: #1a1a2e !important;
    border: 1px solid #ddd !important;
    font-size: 13px !important;
}
.stDownloadButton > button:hover {
    background-color: #e2e4e8 !important;
}
/* Mobile-friendly adjustments */
@media (max-width: 768px) {
    .stColumns > div { min-width: 100% !important; }
    .section-header { font-size: 18px; }
    [data-testid="stSidebar"] { width: 200px !important; }
}
</style>
""", unsafe_allow_html=True)

# ── Initialise (once per server session, not on every rerun) ────────────────────
@st.cache_resource
def _initialise():
    db.init_db()
    auth.seed_users()

_initialise()

# ── Helpers ─────────────────────────────────────────────────────────────────────
def nav(page, **kwargs):
    st.session_state["page"] = page
    for k, v in kwargs.items():
        st.session_state[k] = v
    st.rerun()

def level_badge(level):
    if level >= 60: return '<span class="status-badge badge-red">CRITICAL</span>'
    if level >= 45: return '<span class="status-badge badge-orange">WARNING</span>'
    if level >= 30: return '<span class="status-badge badge-yellow">REMINDER</span>'
    return '<span class="status-badge badge-green">OK</span>'

def trend_icon(trend):
    if trend == "declining": return "📉"
    if trend == "growing":   return "📈"
    return "➡️"

def fmt_date(d):
    if not d: return "Never"
    try: return datetime.strptime(d[:10], "%Y-%m-%d").strftime("%d %b %Y")
    except: return d

def days_label(n):
    if n is None: return "—"
    return f"{n} days ago"

# ── Login page ──────────────────────────────────────────────────────────────────
def page_login():
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        import os as _os
        _logo = _os.path.join(_os.path.dirname(__file__), "logo.png")
        st.markdown("<div style='padding:40px 0 20px'>", unsafe_allow_html=True)
        if _os.path.exists(_logo):
            st.image(_logo, use_container_width=True)
        else:
            st.markdown("""
            <div style="background:#1a1a2e;padding:24px;border-radius:12px;margin-bottom:24px;text-align:center">
                <h1 style="color:#ffffff;margin:0;font-size:28px">Tungaloy-NTK</h1>
                <p style="color:#c0392b;margin:4px 0 0;font-size:16px;font-weight:600">Sales Navigator</p>
            </div>
            """, unsafe_allow_html=True)
        st.markdown("""
        <p style="text-align:center;font-size:16px;font-weight:600;color:#c0392b;margin:8px 0 20px">
            Sales Navigator
        </p>
        </div>
        """, unsafe_allow_html=True)

        with st.form("login_form"):
            username = st.text_input("Username", placeholder="e.g. simon.turnock")
            password = st.text_input("Password", type="password")
            submitted = st.form_submit_button("Sign In", use_container_width=True, type="primary")

        if submitted:
            user, error_msg = auth.login(username, password)
            if user:
                auth.set_session(user)
                try:
                    db.log_audit(user["id"], user["full_name"], "login", "user",
                                 entity_label=user["full_name"])
                except Exception:
                    pass
                st.rerun()
            else:
                st.error(error_msg or "Invalid username or password.")

        st.markdown("""
        <p style="text-align:center;font-size:12px;color:#888;margin-top:16px">
            Default password: <code>Tungaloy2024!</code><br>
            You will be prompted to change it on first login.
        </p>
        """, unsafe_allow_html=True)

# ── Change password page ────────────────────────────────────────────────────────
def page_change_password():
    st.markdown('<div class="section-header">Change Your Password</div>', unsafe_allow_html=True)
    st.info("Please set a new password before continuing.")
    with st.form("change_pw"):
        new_pw  = st.text_input("New password", type="password")
        confirm = st.text_input("Confirm password", type="password")
        sub = st.form_submit_button("Update Password", type="primary")
    if sub:
        errors = auth.validate_password_strength(new_pw)
        if errors:
            st.error("Password must include: " + ", ".join(errors))
        elif new_pw != confirm:
            st.error("Passwords do not match.")
        else:
            uid = st.session_state["user_id"]
            db.update_password(uid, auth.hash_password(new_pw))
            db.update_password_changed_at(uid)
            st.session_state["must_change_password"] = False
            st.success("Password updated! Redirecting…")
            st.rerun()

# ── Sidebar ──────────────────────────────────────────────────────────────────────
def render_sidebar():
    with st.sidebar:
        import os as _os
        _logo = _os.path.join(_os.path.dirname(__file__), "logo.png")
        if _os.path.exists(_logo):
            st.image(_logo, use_container_width=True)
        else:
            st.markdown("""
            <div style="padding:16px 0 8px">
                <div style="font-size:20px;font-weight:700;color:#fff">Tungaloy-NTK</div>
                <div style="font-size:13px;color:#c0392b;font-weight:600">Sales Navigator</div>
            </div>
            """, unsafe_allow_html=True)
        st.markdown("<hr style='border-color:#333;margin:8px 0 16px'/>", unsafe_allow_html=True)

        role = st.session_state.get("role","")
        name = st.session_state.get("full_name","")
        st.markdown(f"<div style='font-size:13px;color:#aaa'>Signed in as</div>"
                    f"<div style='font-size:15px;font-weight:600'>{name}</div>"
                    f"<div style='font-size:12px;color:#888;margin-bottom:16px'>{role.replace('_',' ').title()}</div>",
                    unsafe_allow_html=True)

        pages = [("🏠", "Home",          "hub"),
                 ("👥", "Customers",     "customers"),
                 ("📅", "Visit Planner", "visit_planner"),
                 ("📋", "CRM",           "crm"),
                 ("📧", "Campaigns",     "email_campaigns"),
                 ("🔔", "Alerts",        "alerts"),
                 ("🔧", "Test Tools",    "test_tools"),
                 ("📢", "Mktg Leads",    "marketing_leads"),
                 ("🏆", "League Tables", "league_tables"),
                 ("📋", "Reports",       "reports"),
                 ("🧾", "Expenses",      "expenses"),
                 ("🔄", "Insert Conv.",  "insert_converter")]
        if role in ("admin", "marketing"):
            pages.append(("📤", "Upload Data",  "upload"))
            pages.append(("⚙️",  "Admin",        "admin"))

        current = st.session_state.get("page","dashboard")
        for icon, label, key in pages:
            active = "background:#c0392b;border-radius:6px;" if key == current else ""
            if st.button(f"{icon}  {label}", key=f"nav_{key}",
                         use_container_width=True,
                         type="primary" if key == current else "secondary"):
                nav(key)

        st.markdown("<hr style='border-color:#333;margin:16px 0'/>", unsafe_allow_html=True)
        if st.button("❓  Help", key="nav_help", use_container_width=True):
            nav("help")
        if st.button("🚪  Sign Out", use_container_width=True):
            try:
                db.log_audit(st.session_state.get("user_id", 0),
                             st.session_state.get("full_name", ""),
                             "logout", "user",
                             entity_label=st.session_state.get("full_name", ""))
            except Exception:
                pass
            auth.logout()
            st.rerun()

# ── Hub landing page ─────────────────────────────────────────────────────────────
# Force app restart to clear any cached data (updated with fresh GP report)
def page_hub():
    role    = st.session_state["role"]
    user_id = st.session_state["user_id"]
    name    = st.session_state.get("full_name", "")

    # ── Welcome banner ───────────────────────────────────────────────────────
    from datetime import date as _date, datetime as _dt
    _yr = _date.today().year
    _today_str = _date.today().strftime("%A %d %B %Y")

    st.markdown(f"""
    <div style="background:linear-gradient(135deg,#1a1a2e 0%,#16213e 100%);
                padding:28px 32px;border-radius:14px;margin-bottom:20px">
        <h1 style="color:#fff;margin:0;font-size:26px;letter-spacing:-0.3px">
            Welcome back, {name.split()[0]}</h1>
        <p style="color:#8899aa;margin:6px 0 0;font-size:13px">{_today_str}</p>
    </div>
    """, unsafe_allow_html=True)

    # ── Gather data ──────────────────────────────────────────────────────────
    hub = db.get_hub_counts(user_id, role)
    alert_count   = hub["alert_count"]
    total_custs   = hub["total_customers"]
    pending_approvals = len(db.get_test_requests_pending_approval()) if role in ("admin","regional_manager","marketing") else 0
    overdue_tests = len(db.get_test_requests_needing_followup())
    pending_leads = db.get_pending_marketing_leads_count(user_id if role == "rep" else None)

    # YTD sales from database
    try:
        with db.get_conn() as conn:
            if role == "rep":
                _sm_names = [r["sm_name"] for r in conn.execute(
                    "SELECT sm_name FROM sm_name_mapping WHERE user_id=?", (user_id,)).fetchall()]
                if _sm_names:
                    _ph = ",".join("?" * len(_sm_names))
                    _ytd = conn.execute(f"SELECT COALESCE(SUM(sales_val),0) AS s FROM sales WHERE year=? AND sm_name IN ({_ph})", [_yr]+_sm_names).fetchone()
                    _py  = conn.execute(f"SELECT COALESCE(SUM(sales_val),0) AS s FROM sales WHERE year=? AND sm_name IN ({_ph})", [_yr-1]+_sm_names).fetchone()
                else:
                    _ytd = {"s": 0}; _py = {"s": 0}
            else:
                _ytd = conn.execute("SELECT COALESCE(SUM(sales_val),0) AS s FROM sales WHERE year=?", [_yr]).fetchone()
                _py  = conn.execute("SELECT COALESCE(SUM(sales_val),0) AS s FROM sales WHERE year=?", [_yr-1]).fetchone()
            ytd_sales = _ytd["s"]
            py_sales  = _py["s"]
            ytd_growth = ytd_sales - py_sales
            ytd_gpct   = (ytd_growth / py_sales * 100) if py_sales else 0
    except Exception:
        ytd_sales = 0; py_sales = 0; ytd_growth = 0; ytd_gpct = 0

    # ── KPI Metrics Row ──────────────────────────────────────────────────────
    k1, k2, k3, k4, k5 = st.columns(5)
    k1.metric("YTD Sales", f"£{ytd_sales:,.0f}",
              delta=f"{ytd_gpct:+.1f}% vs {_yr-1}")
    k2.metric("Customers", f"{total_custs:,}")
    k3.metric("Active Alerts", f"{alert_count:,}",
              delta=f"{alert_count} need attention" if alert_count else "All clear",
              delta_color="inverse" if alert_count else "off")
    k4.metric("Test Tools", f"{pending_approvals + overdue_tests}",
              delta=f"{pending_approvals} awaiting approval" if pending_approvals else "None pending",
              delta_color="off")
    k5.metric("Mktg Leads", f"{pending_leads}",
              delta="pending action" if pending_leads else "None pending",
              delta_color="off")

    st.markdown("")

    # ── Quick-link tile helper ───────────────────────────────────────────────
    def _tile(icon, title, subtitle, accent):
        st.markdown(f"""
        <div style="background:#fff;border-radius:10px;padding:16px 18px;
                    border-left:4px solid {accent};min-height:80px;
                    box-shadow:0 1px 4px rgba(0,0,0,0.06)">
            <div style="display:flex;align-items:center;gap:10px;margin-bottom:6px">
                <span style="font-size:20px">{icon}</span>
                <span style="font-size:15px;font-weight:700;color:#1a1a2e">{title}</span>
            </div>
            <div style="font-size:11px;color:#888">{subtitle}</div>
        </div>""", unsafe_allow_html=True)

    # ── Navigation Tiles — Row 1 ────────────────────────────────────────────
    c1, c2, c3, c4 = st.columns(4)
    with c1:
        _tile("👥", "Customers", f"{total_custs:,} accounts", "#1a1a2e")
        if st.button("Open", key="hub_customers", use_container_width=True):
            nav("customers")
    with c2:
        _tile("📈", "Reports", "Sales reports & analysis", "#27ae60")
        if st.button("Open", key="hub_reports", use_container_width=True):
            nav("reports")
    with c3:
        _tile("📋", "CRM", "Customer 360 & tasks", "#2980b9")
        if st.button("Open", key="hub_crm", use_container_width=True):
            nav("crm")
    with c4:
        _tile("📅", "Visit Planner", "Schedule customer visits", "#16a085")
        if st.button("Open", key="hub_planner", use_container_width=True):
            nav("visit_planner")

    # ── Navigation Tiles — Row 2 ────────────────────────────────────────────
    c1, c2, c3, c4 = st.columns(4)
    with c1:
        _tile("🔔", "Alerts", f"{alert_count} active", "#c0392b")
        if st.button("Open", key="hub_alerts", use_container_width=True):
            nav("alerts")
    with c2:
        _tile("🏆", "League Tables", "Rankings & performance", "#f39c12")
        if st.button("Open", key="hub_league", use_container_width=True):
            nav("league_tables")
    with c3:
        _tile("🔧", "Test Tools", f"{pending_approvals + overdue_tests} pending", "#8e44ad")
        if st.button("Open", key="hub_test_tools", use_container_width=True):
            nav("test_tools")
    with c4:
        _tile("📢", "Mktg Leads", f"{pending_leads} pending", "#3498db")
        if st.button("Open", key="hub_leads", use_container_width=True):
            nav("marketing_leads")

    # ── Navigation Tiles — Row 3 (Expenses + Admin) ─────────────────────────
    if role in ("admin", "marketing"):
        c1, c2, c3, c4 = st.columns(4)
        with c1:
            _tile("🧾", "Expenses", "Submit & review expenses", "#e67e22")
            if st.button("Open", key="hub_expenses", use_container_width=True):
                nav("expenses")
        with c2:
            _tile("📤", "Upload Data", "Import monthly GP report", "#34495e")
            if st.button("Open", key="hub_upload", use_container_width=True):
                nav("upload")
        with c3:
            _tile("⚙️", "Admin", "Users, settings & config", "#95a5a6")
            if st.button("Open", key="hub_admin", use_container_width=True):
                nav("admin")
        with c4:
            pass
    else:
        c1, c2, _, _ = st.columns(4)
        with c1:
            _tile("🧾", "Expenses", "Submit & review expenses", "#e67e22")
            if st.button("Open", key="hub_expenses", use_container_width=True):
                nav("expenses")

    # ── Consignment vs Non-Consignment (YTD) — split by whs_code ────────────
    _CONS_WHS = ('1B','1C','BM','C2','C3','D5','D8','D9','DP','E1','E5','E7','E8','E9','EB','S1','SW')
    try:
        with db.get_conn() as conn:
            _wph = ",".join("?" * len(_CONS_WHS))
            _cons_codes = [r['customer_code'] for r in conn.execute(
                "SELECT DISTINCT customer_code FROM consignment_warehouses "
                "WHERE is_active=1 AND customer_code IS NOT NULL"
            ).fetchall()]
            if _cons_codes:
                _cph = ",".join("?" * len(_cons_codes))
                _cons = conn.execute(f"""
                    SELECT COALESCE(SUM(sales_val),0) AS sales,
                           COUNT(DISTINCT customer_code) AS custs
                    FROM sales WHERE year=? AND customer_code IN ({_cph}) AND whs_code IN ({_wph})
                """, [_yr] + _cons_codes + list(_CONS_WHS)).fetchone()
                _non = conn.execute(f"""
                    SELECT COALESCE(SUM(sales_val),0) AS sales
                    FROM sales WHERE year=? AND customer_code IN ({_cph})
                          AND (whs_code NOT IN ({_wph}) OR whs_code IS NULL)
                """, [_yr] + _cons_codes + list(_CONS_WHS)).fetchone()
                _cons_py = conn.execute(f"""
                    SELECT COALESCE(SUM(sales_val),0) AS sales
                    FROM sales WHERE year=? AND customer_code IN ({_cph}) AND whs_code IN ({_wph})
                """, [_yr - 1] + _cons_codes + list(_CONS_WHS)).fetchone()

                _cs   = _cons['sales']
                _ns   = _non['sales']
                _ts   = _cs + _ns
                _cpct = (_cs / _ts * 100) if _ts else 0
                _cpy  = _cons_py['sales']
                _grow = _cs - _cpy
                _gpct = (_grow / _cpy * 100) if _cpy else 0

                st.markdown("---")
                st.markdown("#### 📦 Consignment vs Non-Consignment (YTD)")
                c1, c2, c3, c4, c5 = st.columns(5)
                c1.metric("Consignment WHS Sales", f"£{_cs:,.0f}",
                          delta=f"{_cpct:.1f}% of total")
                c2.metric("Non-Consignment Sales", f"£{_ns:,.0f}",
                          delta=f"{100 - _cpct:.1f}% of total")
                c3.metric("Total (Consign Accounts)", f"£{_ts:,.0f}")
                c4.metric("Consignment Accounts", f"{_cons['custs']}",
                          delta=f"of {len(_cons_codes)} with consign stock")
                c5.metric("vs Last Year", f"£{_grow:+,.0f}",
                          delta=f"{_gpct:+.1f}%", delta_color="normal")
    except Exception:
        pass

# ── Customer list ────────────────────────────────────────────────────────────────
def page_customers():
    st.markdown('<div class="section-header">Customers</div>', unsafe_allow_html=True)
    role    = st.session_state["role"]
    user_id = st.session_state["user_id"]

    with st.spinner("Loading customers…"):
        if role == "rep":
            customers = db.get_customers_for_user(user_id)
        elif role == "regional_manager":
            customers = db.get_customers_for_team(user_id)
        else:
            customers = db.get_all_customers()

    if not customers:
        st.info("No customers found. Upload a GP report to populate customer data.")
        return

    # SM Name list for filter
    sm_names_set = sorted(set(c["sm_name"] for c in customers if c.get("sm_name")))

    # Filters
    col1, col2, col3, col4 = st.columns([3, 2, 2, 2])
    search     = col1.text_input("Search customer", placeholder="Name or code...")
    filter_sm  = col2.selectbox("SM Name", ["All SM Names"] + sm_names_set, key="cust_sm_filter")
    filter_lvl = col3.selectbox("Alert status", ["All","Critical","Warning","Reminder","OK"])
    filter_hv  = col4.checkbox("High-value only (>£2,000/mo)")

    hv_threshold = float(db.get_setting("high_value_monthly_gbp") or 2000)

    def alert_level(purchase_days, visit_days):
        def lvl(d):
            if d is None: return 0
            for t in [60, 45, 30]:
                if d >= t: return t
            return 0
        return max(lvl(purchase_days), lvl(visit_days))

    rows = []
    filtered = []
    for c in customers:
        pd_ = c.get("purchase_days")
        vd_ = c.get("visit_days")
        lvl = alert_level(pd_, vd_)
        last_month_sales = c.get("last_month_sales") or 0

        if search and search.lower() not in (c["customer_name"] or "").lower() \
                   and search.lower() not in (c["customer_code"] or "").lower():
            continue
        if filter_sm != "All SM Names" and c.get("sm_name") != filter_sm:
            continue
        if filter_hv and last_month_sales < hv_threshold:
            continue
        if filter_lvl == "Critical" and lvl < 60: continue
        if filter_lvl == "Warning"  and (lvl < 45 or lvl >= 60): continue
        if filter_lvl == "Reminder" and (lvl < 30 or lvl >= 45): continue
        if filter_lvl == "OK"       and lvl >= 30: continue

        filtered.append(c)
        rows.append({
            "_code":            c["customer_code"],
            "Customer":         c["customer_name"],
            "Rep":              c.get("rep_name") or "",
            "Last Visit":       fmt_date(c.get("last_visit")),
            "Days Since Visit":  vd_ or 999,
            "Last Purchase":    fmt_date(c.get("last_sale")),
            "Days Since Order":  pd_ or 999,
            "Alert Level":      lvl,
            "Trend":            None,
            "_raw_sales":       last_month_sales,
            "Monthly Sales":    f"£{last_month_sales:,.0f}" if last_month_sales else "—",
        })

    if not rows:
        st.info("No customers match the current filters.")
        return

    # Sort state
    sort_col = st.session_state.get("cust_sort_col", "Alert Level")
    sort_asc  = st.session_state.get("cust_sort_asc", False)

    SORT_KEYS = {
        "Customer":       lambda r: (r["Customer"] or "").lower(),
        "Last Visit":     lambda r: r["Days Since Visit"],
        "Last Purchase":  lambda r: r["Days Since Order"],
        "Monthly Sales":  lambda r: r.get("_raw_sales") or 0,
        "Trend":          lambda r: r["Trend"] or "",
        "Alert Level":    lambda r: r["Alert Level"],
    }

    # Store raw sales value for sorting before formatting
    for r in rows:
        pass  # already have Days Since Order/Visit as numeric

    rows.sort(key=SORT_KEYS.get(sort_col, SORT_KEYS["Alert Level"]), reverse=not sort_asc)

    def sort_btn(label, key):
        icon = ""
        if sort_col == key:
            icon = " ▲" if sort_asc else " ▼"
        col_label = f"**{label}{icon}**"
        return col_label

    def on_sort(key):
        if st.session_state.get("cust_sort_col") == key:
            st.session_state["cust_sort_asc"] = not st.session_state.get("cust_sort_asc", False)
        else:
            st.session_state["cust_sort_col"] = key
            st.session_state["cust_sort_asc"] = key == "Customer"  # A-Z for name, desc for others
        st.rerun()

    # Header row with sort buttons
    h1, h2, h3, h4, h5, h6, h7 = st.columns([3, 2, 2, 2, 1, 1, 1])
    if h1.button(sort_btn("Customer", "Customer"),       key="sh_customer",  use_container_width=True): on_sort("Customer")
    if h2.button(sort_btn("Last Visit", "Last Visit"),   key="sh_visit",     use_container_width=True): on_sort("Last Visit")
    if h3.button(sort_btn("Last Purchase","Last Purchase"),key="sh_purchase", use_container_width=True): on_sort("Last Purchase")
    if h4.button(sort_btn("Monthly Sales","Monthly Sales"),key="sh_sales",   use_container_width=True): on_sort("Monthly Sales")
    if h5.button(sort_btn("Trend", "Trend"),             key="sh_trend",     use_container_width=True): on_sort("Trend")
    if h6.button(sort_btn("Status", "Alert Level"),      key="sh_status",    use_container_width=True): on_sort("Alert Level")
    h7.markdown("")
    st.markdown("<hr style='margin:4px 0 8px'/>", unsafe_allow_html=True)

    for row in rows:
        c1, c2, c3, c4, c5, c6, c7 = st.columns([3, 2, 2, 2, 1, 1, 1])
        c1.write(f"**{row['Customer']}**" + (f"\n_{row['Rep']}_" if row.get("Rep") else ""))
        c2.write(row["Last Visit"])
        c3.write(row["Last Purchase"])
        c4.write(row["Monthly Sales"])
        c5.write(trend_icon(row["Trend"]))
        c6.markdown(level_badge(row["Alert Level"]), unsafe_allow_html=True)
        if c7.button("View", key=f"cust_view_{row['_code']}"):
            nav("customer_detail", selected_customer=row["_code"])

    # ── CSV Export ──
    if filtered:
        _csv_df = pd.DataFrame([{
            "Customer Code": c["customer_code"],
            "Customer Name": c["customer_name"],
            "SM Name": c.get("sm_name", ""),
            "Rep": c.get("rep_name", ""),
            "Last Purchase": c.get("last_sale", ""),
            "Last Visit": c.get("last_visit", ""),
            "Purchase Days": c.get("purchase_days", ""),
            "Visit Days": c.get("visit_days", ""),
            "Last Month Sales": c.get("last_month_sales", 0),
        } for c in filtered])
        st.download_button("📥 Export Customers CSV", data=_csv_df.to_csv(index=False),
                           file_name="customers_export.csv", mime="text/csv", key="dl_customers")

# ── Customer Dashboard PDF ───────────────────────────────────────────────────────
def _build_sales_chart(monthly, show_gp=True):
    """Return a BytesIO PNG of the monthly sales bar+line chart."""
    import io
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import matplotlib.ticker as mticker
    import numpy as np

    rows = sorted(monthly, key=lambda x: (x['year'], x['month']))
    labels = [f"{r['month']:02d}/{str(r['year'])[2:]}" for r in rows]
    sales  = [float(r['total_sales'] or 0) for r in rows]
    gp_val = [float(r['total_gp']    or 0) for r in rows]
    gp_pct = [round(g/s*100, 1) if s else 0 for s, g in zip(sales, gp_val)]

    x     = np.arange(len(labels))
    width = 0.4 if show_gp else 0.65

    fig, ax1 = plt.subplots(figsize=(13, 2.6))
    ax1.set_facecolor("#ffffff")
    fig.patch.set_facecolor("#ffffff")

    ax1.bar(x - width/2 if show_gp else x, sales,
            width, label="Sales £", color="#1a1a2e", zorder=3)
    if show_gp:
        ax1.bar(x + width/2, gp_val, width, label="GP £", color="#c0392b", zorder=3)

    ax1.set_ylabel("£", fontsize=7, color="#333")
    ax1.yaxis.set_major_formatter(mticker.FuncFormatter(lambda v, _: f"£{v:,.0f}"))
    ax1.set_xticks(x)
    ax1.set_xticklabels(labels, fontsize=6.5, rotation=45, ha="right")
    ax1.tick_params(axis="y", labelsize=7)
    ax1.grid(axis="y", linestyle="--", alpha=0.35, zorder=0)
    ax1.set_axisbelow(True)

    if show_gp:
        ax2 = ax1.twinx()
        ax2.plot(x, gp_pct, color="#e67e22", linewidth=1.5,
                 marker="o", markersize=3, label="GP %", zorder=4)
        ax2.set_ylabel("GP %", fontsize=7, color="#e67e22")
        ax2.yaxis.set_major_formatter(mticker.FuncFormatter(lambda v, _: f"{v:.0f}%"))
        ax2.tick_params(axis="y", labelcolor="#e67e22", labelsize=7)
        lines1, lbl1 = ax1.get_legend_handles_labels()
        lines2, lbl2 = ax2.get_legend_handles_labels()
        ax1.legend(lines1 + lines2, lbl1 + lbl2,
                   loc="upper left", fontsize=7, framealpha=0.7)
    else:
        ax1.legend(loc="upper left", fontsize=7, framealpha=0.7)

    ax1.set_title("Monthly Sales" + (" & GP" if show_gp else ""),
                  fontsize=9, color="#1a1a2e", pad=4)
    plt.tight_layout(pad=0.3)

    buf_img = io.BytesIO()
    fig.savefig(buf_img, format="png", dpi=150, bbox_inches="tight")
    plt.close(fig)
    buf_img.seek(0)
    return buf_img


def _make_table(data, col_widths, brand_dark, mid_grey, lt_grey,
                font_size=7, right_cols=None):
    """Build a compact styled reportlab Table."""
    from reportlab.platypus import Table, TableStyle
    from reportlab.lib import colors
    right_cols = right_cols or []
    tbl = Table(data, colWidths=col_widths, repeatRows=1)
    style = [
        ("BACKGROUND",    (0, 0), (-1, 0),  brand_dark),
        ("TEXTCOLOR",     (0, 0), (-1, 0),  colors.white),
        ("FONTNAME",      (0, 0), (-1, 0),  "Helvetica-Bold"),
        ("FONTSIZE",      (0, 0), (-1, -1), font_size),
        ("TOPPADDING",    (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ("LEFTPADDING",   (0, 0), (-1, -1), 3),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 3),
        ("ROWBACKGROUNDS",(0, 1), (-1, -1), [colors.white, lt_grey]),
        ("GRID",          (0, 0), (-1, -1), 0.3, mid_grey),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
    ]
    for col in right_cols:
        style.append(("ALIGN", (col, 0), (col, -1), "RIGHT"))
    tbl.setStyle(TableStyle(style))
    return tbl


def _build_customer_pdf(code, cust, status, monthly, visits, items, notes, show_gp=True):
    import io
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer, HRFlowable, Image)

    PAGE  = landscape(A4)
    MARG  = 0.9 * cm
    PW    = PAGE[0] - 2 * MARG   # usable width  ~26.9 cm
    PH    = PAGE[1] - 2 * MARG   # usable height ~18.1 cm

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=PAGE,
        leftMargin=MARG, rightMargin=MARG,
        topMargin=MARG,  bottomMargin=MARG,
    )

    styles    = getSampleStyleSheet()
    brand_red  = colors.HexColor("#c0392b")
    brand_dark = colors.HexColor("#1a1a2e")
    lt_grey    = colors.HexColor("#f2f2f2")
    mid_grey   = colors.HexColor("#dddddd")

    h1  = ParagraphStyle("h1",  parent=styles["Title"],   textColor=brand_dark, fontSize=14, spaceAfter=1, spaceBefore=0)
    h2  = ParagraphStyle("h2",  parent=styles["Heading3"],textColor=brand_red,  fontSize=8,  spaceAfter=2, spaceBefore=3)
    sub = ParagraphStyle("sub", parent=styles["Normal"],  textColor=colors.grey,fontSize=7,  spaceAfter=2)
    sm  = ParagraphStyle("sm",  parent=styles["Normal"],  fontSize=7,           spaceAfter=1)
    xs  = ParagraphStyle("xs",  parent=styles["Normal"],  textColor=colors.grey,fontSize=6.5,spaceAfter=1)

    story = []

    # ════════════════════════════════════════════════════════════════════
    # ROW 1 — header left | status right
    # ════════════════════════════════════════════════════════════════════
    lvl_labels = {"ok": "OK ✓", "reminder": "Reminder", "warning": "Warning ⚠", "critical": "CRITICAL !", None: "—"}

    hdr_left = [
        Paragraph("Tungaloy-NTK Sales Navigator", sub),
        Paragraph(cust['customer_name'], h1),
        Paragraph(
            f"Code: {code}  ·  Territory: {cust.get('sm_name','—')}  ·  {date.today().strftime('%d %b %Y')}",
            sub
        ),
    ]

    visit_str = fmt_date(status["last_visit"])
    if status.get('visit_days'):
        visit_str += f"\n{status['visit_days']} days ago"
    purch_str = fmt_date(status["last_purchase"])
    if status.get('purchase_days'):
        purch_str += f"\n{status['purchase_days']} days ago"

    status_tbl = _make_table(
        [["Last Visit", "Last Purchase", "Spend Trend", "Alert Status"],
         [visit_str, purch_str,
          (status["trend"] or "—").title(),
          lvl_labels.get(status.get("overall_level"), "—")]],
        [3.0*cm, 3.2*cm, 2.6*cm, 2.6*cm],
        brand_dark, mid_grey, lt_grey, font_size=8
    )

    hdr_row = Table(
        [[hdr_left, status_tbl]],
        colWidths=[PW * 0.52, PW * 0.48],
    )
    hdr_row.setStyle(TableStyle([
        ("VALIGN",  (0,0),(-1,-1), "BOTTOM"),
        ("LEFTPADDING",  (0,0),(-1,-1), 0),
        ("RIGHTPADDING", (0,0),(-1,-1), 0),
        ("TOPPADDING",   (0,0),(-1,-1), 0),
        ("BOTTOMPADDING",(0,0),(-1,-1), 0),
    ]))
    story.append(hdr_row)
    story.append(HRFlowable(width="100%", thickness=1.5, color=brand_red, spaceAfter=4, spaceBefore=4))

    # ════════════════════════════════════════════════════════════════════
    # ROW 2 — chart (full width)
    # ════════════════════════════════════════════════════════════════════
    if monthly:
        try:
            chart_buf = _build_sales_chart(monthly, show_gp=show_gp)
            chart_h   = PW * 0.195          # ~5.2 cm tall
            story.append(Image(chart_buf, width=PW, height=chart_h))
        except Exception:
            pass
        story.append(Spacer(1, 0.2*cm))

    # ════════════════════════════════════════════════════════════════════
    # ROW 3 — monthly table | items table | alerts+visits+notes
    # ════════════════════════════════════════════════════════════════════
    COL_L = PW * 0.22   # monthly sales
    COL_M = PW * 0.38   # items
    COL_R = PW * 0.38   # alerts / visits / notes
    GAP   = 0.15*cm

    # ── Monthly sales table ──
    # Column widths are proportional to COL_L so the inner table always fits
    if monthly:
        if show_gp:
            m_hdrs = ["Period", "Sales £", "GP £", "GP %"]
            mw = [COL_L*0.31, COL_L*0.26, COL_L*0.26, COL_L*0.17]
            m_right = [1, 2, 3]
        else:
            m_hdrs = ["Period", "Sales £"]
            mw = [COL_L*0.44, COL_L*0.56]
            m_right = [1]
        m_data = [m_hdrs]
        for r in sorted(monthly, key=lambda x: (x['year'], x['month']), reverse=True)[:16]:
            s = float(r['total_sales'] or 0)
            g = float(r['total_gp']    or 0)
            pct = round(g/s*100, 1) if s else 0
            row = [f"{r['year']}-{str(r['month']).zfill(2)}", f"£{s:,.0f}"]
            if show_gp:
                row += [f"£{g:,.0f}", f"{pct:.1f}%"]
            m_data.append(row)
        left_cell = [Paragraph("Monthly Sales", h2),
                     _make_table(m_data, mw, brand_dark, mid_grey, lt_grey,
                                 font_size=7, right_cols=m_right)]
    else:
        left_cell = [Paragraph("No sales data.", xs)]

    # ── Items table ──
    if items:
        i_data = [["Item Code", "Description", "Orders", "Last Ordered", "Sales £"]]
        for r in items[:12]:
            i_data.append([
                r['item_code'],
                (r['item_desc'] or "")[:35],
                str(r['order_count']),
                fmt_date(r['last_ordered']),
                f"£{float(r['total_sales'] or 0):,.0f}",
            ])
        iw = [2.0*cm, 4.4*cm, 1.2*cm, 2.2*cm, 1.8*cm]
        mid_cell = [Paragraph("Items Purchased", h2),
                    _make_table(i_data, iw, brand_dark, mid_grey, lt_grey,
                                font_size=7, right_cols=[2, 4])]
    else:
        mid_cell = [Paragraph("No items data.", xs)]

    # ── Alerts + Visits + Notes ──
    right_cell = []
    if status.get("item_alerts"):
        right_cell.append(Paragraph(f"Reorder Alerts ({len(status['item_alerts'])})", h2))
        a_data = [["Item Code", "Description", "Avg Interval", "Days Overdue"]]
        for ia in status["item_alerts"]:
            a_data.append([
                ia['item_code'],
                (ia['item_desc'] or "")[:32],
                f"{ia['avg_interval_days']}d",
                f"{ia['days_overdue']}d OVERDUE",
            ])
        aw = [1.9*cm, 4.0*cm, 1.8*cm, 2.2*cm]
        right_cell.append(_make_table(a_data, aw, brand_dark, mid_grey, lt_grey,
                                      font_size=7, right_cols=[2, 3]))
        right_cell.append(Spacer(1, 0.2*cm))

    if visits:
        right_cell.append(Paragraph("Visit History", h2))
        for v in visits[:4]:
            right_cell.append(Paragraph(
                f"<b>{fmt_date(v['visit_date'])}</b> · {v['contact_name'] or '—'} · {v['rep_name'] or '—'}",
                sm
            ))
            if v.get('outcome'):
                right_cell.append(Paragraph(f"  Outcome: {v['outcome']}", xs))
            if v.get('notes'):
                right_cell.append(Paragraph(f"  {v['notes']}", xs))
        right_cell.append(Spacer(1, 0.2*cm))

    if notes:
        right_cell.append(Paragraph("Management Notes", h2))
        for n in notes[:3]:
            right_cell.append(Paragraph(
                f"<b>{n['full_name']}</b> · {fmt_date(n['created_at'])}", sm))
            right_cell.append(Paragraph(n['note_text'], xs))

    if not right_cell:
        right_cell = [Paragraph("No alerts, visits or notes.", xs)]

    # ── Assemble 3-column bottom row ──
    bottom_row = Table(
        [[left_cell, mid_cell, right_cell]],
        colWidths=[COL_L, COL_M + GAP, COL_R],
        splitByRow=1,
    )
    bottom_row.setStyle(TableStyle([
        ("VALIGN",        (0,0),(-1,-1), "TOP"),
        ("LEFTPADDING",   (0,0),(-1,-1), 0),
        ("RIGHTPADDING",  (1,0),(1,-1),  4),
        ("RIGHTPADDING",  (0,0),(0,-1),  4),
        ("TOPPADDING",    (0,0),(-1,-1), 0),
        ("BOTTOMPADDING", (0,0),(-1,-1), 0),
    ]))
    story.append(bottom_row)

    doc.build(story)
    return buf.getvalue()


# ── Customer detail ──────────────────────────────────────────────────────────────
def page_customer_detail():
    code = st.session_state.get("selected_customer")
    if not code:
        nav("customers")
        return

    cust = db.get_customer(code)
    if not cust:
        st.error("Customer not found.")
        return

    back_col, gp_col, pdf_col = st.columns([5, 2, 1])
    with back_col:
        if st.button("← Back to Customers"):
            nav("customers")

    with st.spinner("Loading customer data…"):
        status  = ae.get_customer_status(code)
        monthly = db.get_monthly_sales(code)
        visits  = db.get_visits_for_customer(code)
        items   = db.get_customer_items(code)

    # Management notes (needed for PDF too)
    role = st.session_state["role"]
    user_id = st.session_state["user_id"]
    mgmt_notes = db.get_customer_notes(code) if role in ("admin", "regional_manager", "marketing") else []

    # Determine if GP should be visible
    _owns_customer = (role == "rep" and cust and cust.get("user_id") == user_id)
    _can_see_gp = role in ("admin", "regional_manager", "marketing") or _owns_customer

    with gp_col:
        if _can_see_gp:
            show_gp = st.checkbox("Include GP in PDF", value=True, key=f"pdf_gp_{code}")
        else:
            show_gp = False

    with pdf_col:
        try:
            pdf_bytes = _build_customer_pdf(
                code, cust, status, monthly, visits, items, mgmt_notes,
                show_gp=show_gp
            )
            safe_name = "".join(c if c.isalnum() else "_" for c in cust['customer_name'])
            st.download_button(
                "⬇ Overview Report PDF",
                data=pdf_bytes,
                file_name=f"{safe_name}_{date.today().strftime('%Y%m%d')}.pdf",
                mime="application/pdf",
                use_container_width=True,
            )
        except Exception as e:
            st.caption(f"PDF unavailable: {e}")

    st.markdown(f"## {cust['customer_name']}")
    st.markdown(f"*Code: {code} · Territory: {cust['sm_name']}*")

    # Status bar
    col1, col2, col3, col4 = st.columns(4)
    col1.metric("Last Visit",    fmt_date(status["last_visit"]),
                days_label(status["visit_days"]))
    col2.metric("Last Purchase", fmt_date(status["last_purchase"]),
                days_label(status["purchase_days"]))
    col3.metric("Spend Trend",   trend_icon(status["trend"]) + " " + (status["trend"] or "—").title())
    col4.markdown(f"**Alert Status**<br>{level_badge(status['overall_level'])}", unsafe_allow_html=True)

    # Declining spend warning
    if status["trend"] == "declining":
        st.warning("📉 This customer's spend has declined for 3 or more consecutive months.")

    # Item reorder alerts
    if status["item_alerts"]:
        with st.expander(f"⚠️ {len(status['item_alerts'])} item(s) overdue for reorder", expanded=True):
            for ia in status["item_alerts"]:
                st.markdown(f"- **{ia['item_code']}** — {ia['item_desc']} "
                            f"| Last ordered: {fmt_date(ia['last_ordered'])} "
                            f"| Avg interval: {ia['avg_interval_days']} days "
                            f"| **{ia['days_overdue']} days overdue**")

    st.markdown("---")
    # Customer notes (admin/regional_manager only)
    role = st.session_state["role"]
    if role in ("admin", "regional_manager", "marketing"):
        with st.expander("📝 Customer Notes (management only)", expanded=False):
            notes = db.get_customer_notes(code)
            with st.form(f"note_form_{code}"):
                new_note = st.text_area("Add a note", placeholder="Key contacts, machines, strategy notes...", key=f"nn_{code}")
                if st.form_submit_button("Save Note"):
                    if new_note.strip():
                        db.add_customer_note(code, st.session_state["user_id"], new_note.strip())
                        st.success("Note saved.")
                        st.rerun()
            if notes:
                for n in notes:
                    st.markdown(f"**{n['full_name']}** — {fmt_date(n['created_at'])}")
                    st.write(n["note_text"])
                    st.markdown("<hr style='margin:4px 0'/>", unsafe_allow_html=True)

    tab1, tab2, tab3 = st.tabs(["📅 Visit History", "💰 Sales History", "🔩 Items"])

    # ── Tab 1: Visits ──
    with tab1:
        if st.button("➕ Log a Visit", type="primary"):
            nav("log_visit", selected_customer=code)

        if visits:
            for v in visits:
                with st.container():
                    c1, c2 = st.columns([1, 4])
                    c1.markdown(f"**{fmt_date(v['visit_date'])}**")
                    c2.markdown(f"**Contact:** {v['contact_name']}")
                    if v["outcome"]:
                        c2.markdown(f"**Outcome:** {v['outcome']}")
                    if v["notes"]:
                        c2.markdown(f"**Notes:** {v['notes']}")
                    if v["next_action"]:
                        c2.markdown(f"**Next Action:** {v['next_action']}")
                    c2.markdown(f"<small>Logged by {v['rep_name']}</small>", unsafe_allow_html=True)
                    st.markdown("<hr style='margin:6px 0'/>", unsafe_allow_html=True)
        else:
            st.info("No visits logged for this customer yet.")

    # ── Tab 2: Sales ──
    with tab2:
        if monthly:
            df_m = pd.DataFrame([dict(r) for r in monthly])
            df_m["Period"]   = df_m["year"].astype(str) + "-" + df_m["month"].astype(str).str.zfill(2)
            df_m["total_sales"] = pd.to_numeric(df_m["total_sales"], errors="coerce").fillna(0)
            df_m["total_gp"]    = pd.to_numeric(df_m["total_gp"],    errors="coerce").fillna(0)
            df_m["gp_pct"]      = df_m.apply(
                lambda r: round(r["total_gp"] / r["total_sales"] * 100, 1) if r["total_sales"] else 0, axis=1
            )

            # Bar chart — Sales (& GP if permitted) with GP% line
            fig = go.Figure()
            fig.add_trace(go.Bar(
                x=df_m["Period"], y=df_m["total_sales"],
                name="Sales £", marker_color="#1a1a2e"
            ))
            if _can_see_gp:
                fig.add_trace(go.Bar(
                    x=df_m["Period"], y=df_m["total_gp"],
                    name="GP £", marker_color="#c0392b"
                ))
                fig.add_trace(go.Scatter(
                    x=df_m["Period"], y=df_m["gp_pct"],
                    name="GP %", mode="lines+markers",
                    marker=dict(color="#e67e22", size=6),
                    line=dict(color="#e67e22", width=2),
                    yaxis="y2",
                ))
            _chart_title = "Monthly Sales, GP £ & GP %" if _can_see_gp else "Monthly Sales"
            _layout_kwargs = dict(
                title=_chart_title,
                barmode="group",
                plot_bgcolor="white",
                height=340,
                margin=dict(l=0, r=50, t=40, b=0),
                legend=dict(orientation="h", yanchor="bottom", y=1.02),
                yaxis=dict(title="£"),
            )
            if _can_see_gp:
                _layout_kwargs["yaxis2"] = dict(title="GP %", overlaying="y", side="right",
                                                ticksuffix="%", showgrid=False)
            fig.update_layout(**_layout_kwargs)
            st.plotly_chart(fig, use_container_width=True)

            # Monthly totals table (mask GP columns if not permitted)
            if _can_see_gp:
                df_display = df_m[["Period","total_sales","total_gp","gp_pct"]].copy()
                df_display = df_display.sort_values("Period", ascending=False)
                st.dataframe(
                    df_display.rename(columns={
                        "total_sales": "Sales £",
                        "total_gp":    "GP £",
                        "gp_pct":      "GP %",
                    }),
                    use_container_width=True,
                    hide_index=True,
                    column_config={
                        "Sales £": st.column_config.NumberColumn("Sales £", format="£%.0f"),
                        "GP £":    st.column_config.NumberColumn("GP £",    format="£%.0f"),
                        "GP %":    st.column_config.NumberColumn("GP %",    format="%.1f%%"),
                    }
                )
            else:
                df_display = df_m[["Period","total_sales"]].copy()
                df_display = df_display.sort_values("Period", ascending=False)
                st.dataframe(
                    df_display.rename(columns={
                        "total_sales": "Sales £",
                    }),
                    use_container_width=True,
                    hide_index=True,
                    column_config={
                        "Sales £": st.column_config.NumberColumn("Sales £", format="£%.0f"),
                    }
                )
        else:
            st.info("No sales data available for this customer.")

    # ── Tab 3: Items ──
    with tab3:
        if items:
            df_i = pd.DataFrame([dict(r) for r in items])
            df_i["Last Ordered"]  = df_i["last_ordered"].apply(fmt_date)
            df_i["Total Sales £"] = pd.to_numeric(df_i["total_sales"], errors="coerce")
            df_i["Total Qty"]     = pd.to_numeric(df_i["total_qty"], errors="coerce")
            cols_show = ["item_code","item_desc","order_count","Last Ordered","Total Sales £","Total Qty"]
            rename = {"item_code":"Item Code","item_desc":"Description","order_count":"Orders"}
            st.dataframe(
                df_i[cols_show].rename(columns=rename),
                use_container_width=True,
                hide_index=True,
                column_config={
                    "Total Sales £": st.column_config.NumberColumn("Total Sales", format="£%.0f"),
                    "Total Qty":     st.column_config.NumberColumn("Qty", format="%,.0f"),
                }
            )
        else:
            st.info("No item data available for this customer.")

# ── Log visit ────────────────────────────────────────────────────────────────────
def page_log_visit():
    import lacrm as _lacrm
    code = st.session_state.get("selected_customer")
    if not code:
        nav("customers")
        return

    cust = db.get_customer(code)
    name = cust["customer_name"] if cust else code

    st.markdown(f'<div class="section-header">Log Visit — {name}</div>',
                unsafe_allow_html=True)
    if st.button("← Cancel"):
        nav("customer_detail", selected_customer=code)

    # Load known contacts for this customer
    with db.get_conn() as conn:
        known = conn.execute(
            "SELECT contact_name, email FROM customer_contacts "
            "WHERE customer_code=? ORDER BY contact_name",
            (code,)
        ).fetchall()

    known_names = [f"{r['contact_name']} ({r['email']})" if r['email'] else r['contact_name']
                   for r in known]
    contact_options = known_names + ["➕ Add a new contact"]

    with st.form("visit_form"):
        col1, col2 = st.columns(2)
        visit_date = col1.date_input("Visit Date *", value=date.today())

        if known_names:
            contact_choice = col2.selectbox(
                "Contact *",
                options=contact_options,
                help="Select the person you met, or add a new one"
            )
        else:
            contact_choice = "➕ Add a new contact"
            col2.info("No contacts on file — add one below")

        # New contact fields — shown when adding new
        new_contact_name  = ""
        new_contact_email = ""
        if contact_choice == "➕ Add a new contact" or not known_names:
            st.markdown("**New Contact Details**")
            nc1, nc2 = st.columns(2)
            new_contact_name  = nc1.text_input("Contact Name *", placeholder="Full name")
            new_contact_email = nc2.text_input("Email", placeholder="email@company.com")

        outcome     = st.text_input("Outcome", placeholder="e.g. Quoted, Demo arranged, Order placed…")
        notes       = st.text_area("Notes", placeholder="Any additional details…", height=100)
        next_action = st.text_input("Next Action", placeholder="e.g. Follow up in 2 weeks with quote")
        submitted   = st.form_submit_button("Save Visit", type="primary")

    if submitted:
        # Resolve the contact name
        if contact_choice == "➕ Add a new contact" or not known_names:
            final_contact = new_contact_name.strip()
        else:
            # Strip email from display string to get just the name
            final_contact = contact_choice.split(" (")[0].strip()

        if not final_contact:
            st.error("Contact Name is required.")
        else:
            # Save the visit
            db.log_visit(
                customer_code=code,
                customer_name=name,
                user_id=st.session_state["user_id"],
                visit_date=visit_date.strftime("%Y-%m-%d"),
                contact_name=final_contact,
                notes=notes.strip() or None,
                outcome=outcome.strip() or None,
                next_action=next_action.strip() or None,
            )

            # If new contact — save to customer_contacts and push to LACRM
            if contact_choice == "➕ Add a new contact" or not known_names:
                if new_contact_name.strip():
                    with db.get_conn() as conn:
                        conn.execute(
                            "INSERT OR IGNORE INTO customer_contacts "
                            "(customer_code, contact_name, email) VALUES (?,?,?)",
                            (code, new_contact_name.strip(),
                             new_contact_email.strip() or None)
                        )
                    # Push to LACRM if email provided
                    if new_contact_email.strip():
                        try:
                            parts = new_contact_name.strip().split(" ", 1)
                            ok, contact_id = _lacrm.create_contact(
                                first_name   = parts[0],
                                last_name    = parts[1] if len(parts) > 1 else "",
                                email        = new_contact_email.strip(),
                                company_name = name,
                                background   = f"Added via Sales Navigator visit log. Rep: {st.session_state.get('full_name','')}",
                            )
                            if ok and contact_id:
                                import requests as _req
                                _req.get("https://api.lessannoyingcrm.com", params={
                                    "UserCode":    db.get_setting("lacrm_user_code"),
                                    "APIToken":    db.get_setting("lacrm_api_key"),
                                    "Function":    "AddContactToGroup",
                                    "ContactId":   contact_id,
                                    "GroupName":   "Sales Navigator",
                                }, timeout=10)
                            st.toast("New contact added to LACRM ✅")
                        except Exception:
                            pass  # Don't block the visit save if LACRM fails

            try:
                db.log_audit(st.session_state["user_id"], st.session_state["full_name"],
                             "create", "visit", entity_label=f"{name} on {visit_date.strftime('%d %b %Y')}",
                             details=f"Contact: {final_contact}")
            except Exception:
                pass
            st.success(f"Visit logged for {name} on {visit_date.strftime('%d %b %Y')}.")
            st.balloons()
            if st.button("Back to Customer"):
                nav("customer_detail", selected_customer=code)

# ── Alerts page ──────────────────────────────────────────────────────────────────
def page_alerts():
    st.markdown('<div class="section-header">Alerts</div>', unsafe_allow_html=True)
    role    = st.session_state["role"]
    user_id = st.session_state["user_id"]

    alert_tab, reorder_tab = st.tabs(["🔔 Customer Alerts", "📦 Reorder Alerts"])

    with reorder_tab:
        st.markdown("#### Proactive Reorder Alerts")
        st.caption("Customers who usually order at a regular interval and are due (or overdue) to place their next order. "
                   "Based on historical ordering patterns — only shown for customers with 3+ orders.")
        with st.spinner("Analysing ordering patterns…"):
            reorder_data = db.get_reorder_alerts(user_id=user_id, role=role)
        if not reorder_data:
            st.success("✅ No reorder alerts — all regular customers are within their expected ordering window.")
        else:
            # Split into overdue vs upcoming
            overdue  = [r for r in reorder_data if (r["days_until_expected"] or 0) <= 0]
            upcoming = [r for r in reorder_data if (r["days_until_expected"] or 0) > 0]

            if overdue:
                st.markdown(f"<h4 style='color:#c0392b'>🔴 Overdue ({len(overdue)})</h4>", unsafe_allow_html=True)
                for r in overdue:
                    c1, c2, c3, c4, c5 = st.columns([3, 2, 2, 2, 1])
                    c1.write(f"**{r['customer_name']}**")
                    c2.write(f"Usually every **{int(r['avg_days_between_orders'])}** days")
                    c3.write(f"Last order: **{int(r['days_since_last_order'])}** days ago")
                    overdue_by = abs(int(r["days_until_expected"] or 0))
                    c4.markdown(f"<span style='color:#c0392b;font-weight:700'>⚠️ {overdue_by} days overdue</span>", unsafe_allow_html=True)
                    if c5.button("View", key=f"reorder_view_{r['customer_code']}"):
                        nav("customer_detail", selected_customer=r["customer_code"])

            if upcoming:
                st.markdown(f"<h4 style='color:#e67e22'>🟠 Due Soon ({len(upcoming)})</h4>", unsafe_allow_html=True)
                for r in upcoming:
                    c1, c2, c3, c4, c5 = st.columns([3, 2, 2, 2, 1])
                    c1.write(f"**{r['customer_name']}**")
                    c2.write(f"Usually every **{int(r['avg_days_between_orders'])}** days")
                    c3.write(f"Last order: **{int(r['days_since_last_order'])}** days ago")
                    days_left = int(r["days_until_expected"] or 0)
                    c4.markdown(f"<span style='color:#e67e22;font-weight:600'>📅 Due in ~{days_left} days</span>", unsafe_allow_html=True)
                    if c5.button("View", key=f"reorder_soon_{r['customer_code']}"):
                        nav("customer_detail", selected_customer=r["customer_code"])

            # CSV export
            _reorder_df = pd.DataFrame([{
                "Customer": r["customer_name"],
                "SM Name": r.get("sm_name", ""),
                "Rep": r.get("rep_name", ""),
                "Avg Days Between Orders": int(r["avg_days_between_orders"]),
                "Days Since Last Order": int(r["days_since_last_order"]),
                "Days Until Expected": int(r["days_until_expected"] or 0),
                "Last Order Date": r.get("last_order_date", ""),
                "Historical Orders": int(r.get("historical_orders", 0)),
            } for r in reorder_data])
            st.download_button("📥 Export Reorder Alerts CSV", data=_reorder_df.to_csv(index=False),
                               file_name="reorder_alerts.csv", mime="text/csv", key="dl_reorder")

    with alert_tab:
        col1, col2, col3 = st.columns([3, 1, 1])
        filter_type = col1.selectbox(
            "Filter by type",
            ["All", "No Purchase", "No Visit", "Declining Spend"]
        )

        emails_enabled = db.get_setting("alerts_enabled") == "true"

        if col2.button("📧 Send Alert Emails", type="primary", disabled=not emails_enabled):
            with st.spinner("Sending alert emails…"):
                sent = ae.run_alerts(send_emails=True, digest_mode=False)
            st.success(f"Emails sent for {len(sent)} alert(s).")
            st.rerun()

        if not emails_enabled:
            col3.warning("Email alerts off")

        # Always compute live — no dependency on alert_log
        with st.spinner("Checking customers…"):
            alerts = ae.compute_alerts(scope_user_id=user_id, role=role)

        type_map = {
            "No Purchase":    "no_purchase",
            "No Visit":       "no_visit",
            "Declining Spend":"declining_spend",
        }
        if filter_type != "All":
            ft = type_map.get(filter_type, "")
            alerts = [a for a in alerts if a["type"] == ft]

        if not alerts:
            st.success("✅ No alerts — all customers are within target visit and purchase windows.")
            return

        st.markdown(f"**{len(alerts)} alert(s) found**")

        # Group by level
        crit  = [a for a in alerts if a["level"] >= 60]
        warn  = [a for a in alerts if 45 <= a["level"] < 60]
        remnd = [a for a in alerts if a["level"] < 45]

        def render_alert_group(group, label, colour):
            if not group:
                return
            st.markdown(f"<h4 style='color:{colour};margin-top:16px'>{label} ({len(group)})</h4>",
                        unsafe_allow_html=True)
            for i, a in enumerate(group):
                c1, c2, c3, c4, c5 = st.columns([3, 2, 2, 2, 1])
                c1.write(f"**{a['customer_name']}**")
                c2.write(a["type_label"])
                c3.write(f"{a['days']} days" if a["days"] else "—")
                c4.write(a.get("rep_name", ""))
                if c5.button("View", key=f"alert_view_{a['customer_code']}_{a['type']}_{i}"):
                    nav("customer_detail", selected_customer=a["customer_code"])

        render_alert_group(crit,  "🔴 Critical (60+ days)", "#c0392b")
        render_alert_group(warn,  "🟠 Warning (45+ days)",  "#e67e22")
        render_alert_group(remnd, "🟡 Reminder (30+ days)", "#d4a017")

# ── Team overview ─────────────────────────────────────────────────────────────────
def page_team():
    st.markdown('<div class="section-header">Team Overview</div>', unsafe_allow_html=True)
    role    = st.session_state["role"]
    user_id = st.session_state["user_id"]

    if role == "regional_manager":
        team = [db.get_user_by_id(user_id)] + list(db.get_team_members(user_id))
    else:
        team = db.get_reps_and_managers()

    team_map = {u["id"]: u["full_name"] for u in team if u}
    selected_rep = st.selectbox(
        "Select rep",
        options=list(team_map.keys()),
        format_func=lambda uid: team_map.get(uid, str(uid))
    )

    if not selected_rep:
        return

    rep = db.get_user_by_id(selected_rep)
    customers = db.get_customers_for_user(selected_rep)   # already includes alert status columns

    # GP masking: this page only shows monthly sales, not GP columns.
    # If GP columns are added here in future, gate them with:
    #   _can_see_gp = role in ("admin", "regional_manager", "marketing")
    # Page access is already restricted to admin/regional_manager/marketing.

    st.markdown(f"### {rep['full_name'] if rep else ''} — {len(customers)} customers")

    def _alert_level(pd_, vd_):
        def lvl(d):
            if d is None: return 0
            for t in [60,45,30]:
                if d >= t: return t
            return 0
        return max(lvl(pd_), lvl(vd_))

    rows = []
    for c in customers:
        pd_ = c.get("purchase_days")
        vd_ = c.get("visit_days")
        lvl = _alert_level(pd_, vd_)
        rows.append({
            "_code":       c["customer_code"],
            "Customer":    c["customer_name"],
            "Last Visit":  fmt_date(c.get("last_visit")),
            "Last Order":  fmt_date(c.get("last_sale")),
            "_sales_raw":  c.get("last_month_sales") or 0,
            "Monthly":     f"£{c.get('last_month_sales'):,.0f}" if c.get("last_month_sales") else "—",
            "Alert":       lvl,
        })

    rows.sort(key=lambda x: x["Alert"], reverse=True)

    hdr = st.columns([3,2,2,2,1,1])
    for h, col in zip(["Customer","Last Visit","Last Order","Monthly","Status",""], hdr):
        col.markdown(f"**{h}**")
    st.markdown("<hr style='margin:4px 0 8px'/>", unsafe_allow_html=True)

    for row in rows:
        c1,c2,c3,c4,c5,c6 = st.columns([3,2,2,2,1,1])
        c1.write(row["Customer"])
        c2.write(row["Last Visit"])
        c3.write(row["Last Order"])
        c4.write(row["Monthly"])
        c5.markdown(level_badge(row["Alert"]), unsafe_allow_html=True)
        if c6.button("View", key=f"team_view_{row['_code']}"):
            nav("customer_detail", selected_customer=row["_code"])

# ── Upload data ───────────────────────────────────────────────────────────────────
def page_upload():
    st.markdown('<div class="section-header">Upload Data</div>', unsafe_allow_html=True)

    tab_gp, tab_cust, tab_item, tab_download = st.tabs(["📊 GP Report", "🏢 Customer Info", "🔩 Item Info", "📥 Download & Verify"])

    # ── GP Report tab ──────────────────────────────────────────────────────────
    with tab_gp:
        batches = db.get_recent_sales_import_batches()
        if batches:
            st.info(f"Last import: **{batches[0]['import_batch']}**")

        st.markdown("""
        Upload the monthly GP Excel report. The app reads the **Display** sheet and imports:
        - Customer and territory data
        - All sales transaction lines
        - Dates and item codes for reorder tracking
        """)

        uploaded = st.file_uploader(
            "Choose GP Report (.xlsx)",
            type=["xlsx"],
            key="gp_upload",
            help="Upload the GP report Excel file — e.g. 'GP report Jan-Mar26.xlsx'"
        )

        if uploaded:
            with st.spinner("Reading file…"):
                try:
                    df = di.parse_gp_excel(uploaded)
                    preview = di.preview_import(df)
                except Exception as e:
                    st.error(f"Error reading file: {e}")
                    return

            st.success(f"File read successfully: **{preview['total_rows']:,} rows**, "
                       f"**{preview['customers']} customers**")
            st.write(f"Date range: {preview['date_range'][0]} → {preview['date_range'][1]}")
            st.write(f"Territories: {', '.join(preview['sm_names'][:10])}")

            with st.expander("Preview first 10 rows"):
                st.dataframe(preview["sample"], use_container_width=True)

            if st.button("✅ Confirm Import", type="primary", key="gp_confirm"):
                with st.spinner("Importing…"):
                    result = di.import_to_db(df)
                msg = (
                    f"Import complete: {result['customers_updated']} customers updated, "
                    f"{result['sales_inserted']} new sales rows added, "
                    f"{result['sales_skipped']} duplicates skipped."
                )
                if result.get("address_backfilled", 0) > 0:
                    msg += f" **{result['address_backfilled']} Ship To Addresses backfilled.**"
                st.success(msg)
                try:
                    db.log_audit(st.session_state["user_id"], st.session_state["full_name"],
                                 "import", "sales_data",
                                 entity_label=uploaded.name,
                                 details=f"{result['sales_inserted']} rows, {result['customers_updated']} customers")
                except Exception:
                    pass

    # ── Customer Info tab ──────────────────────────────────────────────────────
    with tab_cust:
        st.markdown("""
        Upload your **Customer Info** spreadsheet to enrich account records with:
        - **Customer** — customer code/number
        - **Post Area** — postcode area prefix (e.g. G, S, CV)
        - **Region** — UK region (e.g. Scotland, West Midlands, North)
        - **Customer Type** — End User, Distributor, OEM, etc.

        The file must have a **Customer** column containing the customer number.
        Only customers already in the system will be updated; unknown numbers are skipped.
        """)

        col1, col2 = st.columns([3, 1])
        with col1:
            cust_file = st.file_uploader(
                "Choose Customer Info (.xlsx)",
                type=["xlsx"],
                key="cust_info_upload",
                help="Upload the customer info spreadsheet — e.g. 'Customer info 30.06.26.xlsx'"
            )
        with col2:
            import io
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment

            wb = Workbook()
            sheet = wb.active
            sheet.title = "Customer Info"

            headers = ["Customer Code", "Customer Type", "Post Area", "Region", "Salesman Name"]
            for col_num, header in enumerate(headers, 1):
                cell = sheet.cell(row=1, column=col_num)
                cell.value = header
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")

            example_data = [
                [887, "End User", "G", "Scotland", "SCOTLAND-W-HAMILTON"],
                [511, "Distributor", "S", "Scotland", "SCOTLAND-W-HAMILTON"],
                [228, "End User", "CV", "West Midlands", "TURNOCK-WMIDS"],
            ]
            for row_num, row_data in enumerate(example_data, 2):
                for col_num, value in enumerate(row_data, 1):
                    sheet.cell(row=row_num, column=col_num).value = value

            for col_num in range(1, len(headers) + 1):
                sheet.column_dimensions[chr(64 + col_num)].width = 20

            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            st.download_button(
                label="📥 Download Template",
                data=buffer.getvalue(),
                file_name="Customer_Info_Template.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="template_download"
            )

        if cust_file:
            with st.spinner("Reading file…"):
                try:
                    cust_df = di.parse_customer_info(cust_file)
                except Exception as e:
                    st.error(f"Error reading file: {e}")
                    return

            st.success(f"File read successfully: **{len(cust_df):,} customer records**")

            preview_cols = [c for c in ["customer_code","customer_type","post_area","region","salesman_name","main_distributor"] if c in cust_df.columns]
            with st.expander("Preview first 10 rows"):
                st.dataframe(cust_df[preview_cols].head(10), use_container_width=True, hide_index=True)

            if st.button("✅ Import Customer Info", type="primary", key="cust_confirm"):
                with st.spinner("Importing…"):
                    result = di.import_customer_info(cust_df)
                st.success(
                    f"Done! **{result['updated']} customers updated** with new info. "
                    f"{result['skipped']} records skipped (customer number not yet in system)."
                )

    # ── Item Info tab ──────────────────────────────────────────────────────────
    with tab_item:
        st.markdown("""
        Upload your **Item Info** spreadsheet to enrich product records with:
        - **Brand** — product brand (e.g. INNO, TUNG)
        - **Item Type** — product category (e.g. Holder, Insert, Drill)

        The file must have an **Item** column containing the 7-digit item code.
        New items are added; existing items are updated.
        """)

        item_file = st.file_uploader(
            "Choose Item Info (.xlsx)",
            type=["xlsx"],
            key="item_info_upload",
            help="Upload the item info spreadsheet — e.g. 'Item info 09.04.26.xlsx'"
        )

        if item_file:
            with st.spinner("Reading file…"):
                try:
                    item_df = di.parse_item_info(item_file)
                except Exception as e:
                    st.error(f"Error reading file: {e}")
                    return

            st.success(f"File read successfully: **{len(item_df):,} item records**")

            preview_cols = [c for c in ["item_code", "brand", "item_type"] if c in item_df.columns]
            with st.expander("Preview first 10 rows"):
                st.dataframe(item_df[preview_cols].head(10), use_container_width=True, hide_index=True)

            if st.button("✅ Import Item Info", type="primary", key="item_confirm"):
                with st.spinner("Importing…"):
                    result = di.import_item_info(item_df)
                st.success(
                    f"Done! **{result['inserted']} new items added**, "
                    f"**{result['updated']} existing items updated**."
                )

    # ── Download & Verify tab ──────────────────────────────────────────────────────
    with tab_download:
        st.markdown("### Download & Verify Customer Data")
        st.markdown("""
        Download all customer data from the system to verify accuracy, then upload corrected data back.
        This helps ensure your customer information is always up-to-date and correct.
        """)

        col1, col2 = st.columns(2)

        # ── Download section ──
        with col1:
            st.markdown("#### 📥 Download All Contact Data")
            st.caption("Customers + all contact email addresses for campaigns")
            if st.button("⬇️ Download Complete Data", type="primary", use_container_width=True):
                with st.spinner("Preparing export…"):
                    import pandas as pd
                    import io
                    from openpyxl import Workbook
                    from openpyxl.styles import Font, PatternFill, Alignment

                    with db.get_conn() as conn:
                        # Get all customers
                        customers = conn.execute(
                            "SELECT id, customer_code, customer_name, sm_name FROM customers ORDER BY customer_code"
                        ).fetchall()

                        # Get all AC contacts/prospects that were synced
                        ac_contacts = conn.execute(
                            "SELECT email, first_name, last_name, full_name, company, phone, address, city, state, postal_code, country, notes FROM customer_contacts ORDER BY full_name"
                        ).fetchall()

                    # Convert to DataFrames
                    df_customers = pd.DataFrame(customers)
                    df_ac = pd.DataFrame(ac_contacts) if ac_contacts else pd.DataFrame(columns=['email', 'first_name', 'last_name', 'full_name', 'company', 'phone', 'address', 'city', 'state', 'postal_code', 'country', 'notes'])

                    # Create Excel file with multiple sheets
                    output = io.BytesIO()
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        df_customers.to_excel(writer, sheet_name='Customers', index=False)
                        if not df_ac.empty:
                            df_ac.to_excel(writer, sheet_name='AC Contacts & Prospects', index=False)

                        # Format all sheets
                        workbook = writer.book
                        for sheet_name in workbook.sheetnames:
                            worksheet = writer.sheets[sheet_name]
                            for cell in worksheet[1]:
                                cell.font = Font(bold=True, color="FFFFFF")
                                cell.fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
                                cell.alignment = Alignment(horizontal="center")

                            # Auto-adjust column widths
                            for column in worksheet.columns:
                                max_length = 0
                                column_letter = column[0].column_letter
                                for cell in column:
                                    try:
                                        if len(str(cell.value)) > max_length:
                                            max_length = len(str(cell.value))
                                    except:
                                        pass
                                worksheet.column_dimensions[column_letter].width = min(max_length + 2, 50)

                    output.seek(0)
                    st.download_button(
                        label="📄 Download All Data",
                        data=output.getvalue(),
                        file_name=f"all_contact_data_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                    st.success(f"✅ Export ready! Includes:")
                    st.caption(f"• {len(df_customers)} customers\n• {len(df_ac)} Active Campaign contacts & prospects")

        # ── Upload section ──
        with col2:
            st.markdown("#### ⬆️ Upload Cleaned Data")
            st.caption("Upload the corrected Excel file to update contacts")

            corrected_file = st.file_uploader(
                "Choose cleaned data (.xlsx)",
                type=["xlsx"],
                key="customer_data_upload"
            )

            if corrected_file:
                with st.spinner("Reading file…"):
                    try:
                        import pandas as pd
                        excel_file = pd.ExcelFile(corrected_file)

                        # Show available sheets
                        st.info(f"Sheets found: {', '.join(excel_file.sheet_names)}")

                        # Try to read AC Contacts sheet
                        if "AC Contacts & Prospects" in excel_file.sheet_names:
                            df_ac_corrected = pd.read_excel(corrected_file, sheet_name="AC Contacts & Prospects")
                            st.success(f"AC Contacts file read: **{len(df_ac_corrected)} rows**")

                            with st.expander("Preview AC contact data"):
                                st.dataframe(df_ac_corrected.head(10), use_container_width=True)

                            if st.button("✅ Confirm Upload AC Contacts", type="primary", use_container_width=True, key="upload_ac"):
                                with st.spinner("Updating contact data…"):
                                    imported = 0
                                    with db.get_conn() as conn:
                                        for _, row in df_ac_corrected.iterrows():
                                            email = row.get("email", "").strip()
                                            if email:
                                                conn.execute("""
                                                    INSERT OR REPLACE INTO customer_contacts
                                                    (email, first_name, last_name, full_name, company, phone, address,
                                                     city, state, postal_code, country, notes)
                                                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                                                """, (
                                                    email,
                                                    row.get("first_name", ""),
                                                    row.get("last_name", ""),
                                                    row.get("full_name", ""),
                                                    row.get("company", ""),
                                                    row.get("phone", ""),
                                                    row.get("address", ""),
                                                    row.get("city", ""),
                                                    row.get("state", ""),
                                                    row.get("postal_code", ""),
                                                    row.get("country", ""),
                                                    row.get("notes", ""),
                                                ))
                                                imported += 1

                                    st.success(f"✅ Updated {imported} contacts!")
                                    try:
                                        db.log_audit(st.session_state["user_id"], st.session_state["full_name"],
                                                    "import", "ac_contacts_corrected",
                                                    entity_label=corrected_file.name,
                                                    details=f"{imported} AC contacts updated")
                                    except:
                                        pass
                        else:
                            st.warning("No 'AC Contacts & Prospects' sheet found in file")
                    except Exception as e:
                        st.error(f"Error reading file: {e}")

# ── Admin ─────────────────────────────────────────────────────────────────────────
def page_admin():
    import activecampaign as ac
    import lacrm
    st.markdown('<div class="section-header">Admin</div>', unsafe_allow_html=True)
    tab1, tab2, tab3, tab4, tab5, tab6, tab7 = st.tabs(["👥 Users", "⚙️ Settings", "📧 Email Config", "🔗 ActiveCampaign", "📋 LACRM", "📜 Audit Trail", "🔒 Security Log"])

    with tab1:
        users = db.get_all_users()
        df_users = pd.DataFrame([{
            "Username":  u["username"],
            "Name":      u["full_name"],
            "Email":     u["email"],
            "Role":      u["role"].replace("_"," ").title(),
            "Active":    "Yes" if u["is_active"] else "No",
        } for u in users])
        st.dataframe(df_users, use_container_width=True, hide_index=True)

        st.markdown("### Reset User Password")
        reset_user = st.selectbox(
            "Select user",
            [u["username"] for u in users],
            key="reset_user_select"
        )
        new_pw = st.text_input("New password", type="password", key="reset_pw_input")
        if st.button("Reset Password", key="reset_pw_btn"):
            if len(new_pw) < 8:
                st.error("Password must be at least 8 characters.")
            else:
                user = db.get_user_by_username(reset_user)
                if user:
                    db.update_password(user["id"], auth.hash_password(new_pw))
                    db.update_password_changed_at(user["id"])
                    db.reset_failed_logins(user["id"])
                    try:
                        db.log_audit(st.session_state["user_id"], st.session_state["full_name"],
                                     "update", "user", entity_id=user["id"],
                                     entity_label=reset_user, details="Password reset by admin")
                    except Exception:
                        pass
                    st.success(f"Password reset for {reset_user}.")

    with tab2:
        st.markdown("### Alert Thresholds")
        hv = st.number_input(
            "High-value customer threshold (£/month)",
            value=float(db.get_setting("high_value_monthly_gbp") or 2000),
            step=100.0
        )
        alerts_on = st.checkbox(
            "Enable automated alerts",
            value=(db.get_setting("alerts_enabled") == "true")
        )
        st.markdown("### AI Help Assistant")
        ai_key = st.text_input(
            "Anthropic API Key",
            value=db.get_setting("anthropic_api_key") or "",
            type="password",
            help="Required for the Help chat. Get one at console.anthropic.com")

        if st.button("Save Settings", key="save_settings"):
            db.set_setting("high_value_monthly_gbp", str(hv))
            db.set_setting("alerts_enabled", "true" if alerts_on else "false")
            if ai_key:
                db.set_setting("anthropic_api_key", ai_key.strip())
            try:
                db.log_audit(st.session_state["user_id"], st.session_state["full_name"],
                             "update", "setting", entity_label="Admin settings",
                             details=f"HV threshold: £{hv:.0f}, alerts: {'on' if alerts_on else 'off'}")
            except Exception:
                pass
            st.success("Settings saved.")

    with tab3:
        st.markdown("### Email Configuration")

        st.markdown("#### 📤 Campaign Sending (SendGrid / SMTP Relay)")
        st.info("Used by the Email Campaigns module to send bulk emails. "
                "We recommend SendGrid — use `smtp.sendgrid.net`, port 587, "
                "username `apikey`, password = your SendGrid API key.")

        sg1, sg2 = st.columns([3, 1])
        smtp_host = sg1.text_input("SMTP Host",
                                   value=db.get_setting("smtp_host") or "",
                                   placeholder="smtp.sendgrid.net")
        smtp_port = sg2.text_input("SMTP Port",
                                   value=db.get_setting("smtp_port") or "587")
        su1, su2 = st.columns(2)
        smtp_user = su1.text_input("SMTP Username",
                                   value=db.get_setting("smtp_username") or "apikey")
        smtp_pass = su2.text_input("SMTP Password / API Key",
                                   type="password",
                                   help="For SendGrid this is your API key.")
        site_url  = st.text_input("Site Base URL",
                                  value=db.get_setting("site_base_url") or "https://tungaloy.co.uk",
                                  help="Used to build tracking links and unsubscribe URLs in emails.")
        if st.button("💾 Save Campaign Email Settings", key="save_campaign_smtp"):
            db.set_setting("smtp_host",     smtp_host.strip())
            db.set_setting("smtp_port",     smtp_port.strip())
            db.set_setting("smtp_username", smtp_user.strip())
            if smtp_pass:
                db.set_setting("smtp_password", smtp_pass)
            db.set_setting("site_base_url", site_url.strip())
            st.success("Campaign email settings saved.")

        st.markdown("---")
        st.markdown("#### 🔔 Alert / Notification Emails")
        st.info("Configure the SMTP settings for the email account used to send alerts and notifications.")

        col1, col2 = st.columns([3, 1])
        alert_host = col1.text_input("SMTP Host",
                                    value=db.get_setting("email_smtp_host") or "",
                                    placeholder="e.g. mail.yourdomain.co.uk",
                                    key="alert_smtp_host")
        alert_port = col2.text_input("SMTP Port",
                                    value=db.get_setting("email_smtp_port") or "587",
                                    placeholder="587",
                                    key="alert_smtp_port")
        from_email = st.text_input("From email address",
                                   value=db.get_setting("email_from") or "",
                                   placeholder="e.g. alerts@tungaloyuk.co.uk")
        email_pw   = st.text_input("Email password / app password",
                                   type="password",
                                   help="Use an App Password if MFA is enabled on the account.")
        if st.button("💾 Save Alert Email Config", key="save_email"):
            db.set_setting("email_smtp_host", alert_host.strip())
            db.set_setting("email_smtp_port", alert_port.strip())
            db.set_setting("email_from", from_email.strip())
            if email_pw:
                db.set_setting("email_password", email_pw)
            st.success("Alert email configuration saved.")

        if st.button("Send Test Email", key="test_email"):
            ok, msg = ae.send_email(
                [st.session_state["email"]],
                "Sales Navigator — Test Email",
                "<h2>Test email from Tungaloy-NTK Sales Navigator</h2>"
                "<p>If you received this, email is configured correctly.</p>"
            )
            if ok: st.success("Test email sent!")
            else:  st.error(f"Failed: {msg}")

    with tab4:
        st.markdown("### ActiveCampaign Integration")

        # ── Connection settings ──
        with st.expander("🔑 API Credentials", expanded=False):
            ac_url = st.text_input("API URL",
                                   value=db.get_setting("ac_api_url") or "",
                                   placeholder="https://youraccountname.api-us1.com")
            ac_key = st.text_input("API Key", type="password",
                                   value=db.get_setting("ac_api_key") or "")
            if st.button("Save Credentials", key="save_ac_creds"):
                db.set_setting("ac_api_url", ac_url.strip())
                db.set_setting("ac_api_key", ac_key.strip())
                st.success("Credentials saved.")

        # ── Connection test ──
        col_test, _ = st.columns([2, 3])
        if col_test.button("🔌 Test Connection", key="ac_test"):
            with st.spinner("Connecting…"):
                ok, msg = ac.test_connection()
            if ok:
                st.success(f"✅ {msg}")
            else:
                st.error(f"❌ {msg}")

        st.markdown("---")

        # ── Lists & Tags overview ──
        st.markdown("#### Your ActiveCampaign Lists")
        if st.button("🔄 Load Lists & Tags", key="ac_load"):
            with st.spinner("Fetching from ActiveCampaign…"):
                lists = ac.get_lists()
                tags  = ac.get_tags()
            if lists:
                st.session_state["ac_lists"] = lists
            if tags:
                st.session_state["ac_tags"] = tags

        if "ac_lists" in st.session_state:
            lists = st.session_state["ac_lists"]
            st.dataframe(
                pd.DataFrame(lists)[["id", "name"]].rename(
                    columns={"id": "List ID", "name": "List Name"}),
                use_container_width=True, hide_index=True
            )

        if "ac_tags" in st.session_state:
            tags = st.session_state["ac_tags"]
            st.markdown(f"**Tags ({len(tags)}):** " +
                        ", ".join(f"`{t['tag']}`" for t in tags[:30]))

        st.markdown("---")

        # ── Import contacts FROM ActiveCampaign ──
        st.markdown("#### Import Contacts & Prospects FROM ActiveCampaign")
        st.info(
            "Pull all contacts (1,595) + prospects (2,611) from your ActiveCampaign database. "
            "Unsubscribed contacts are automatically excluded. "
            "This creates a searchable contact database for email campaigns."
        )
        if st.button("📥 Sync AC Contacts & Prospects to SalesNavigator", type="primary"):
            with st.spinner("Importing contacts & prospects from ActiveCampaign…"):
                imported, skipped_unsub, errors_count, errors = ac.sync_ac_contacts_to_db()
            st.success(f"✅ Imported {imported} contacts! (Skipped {skipped_unsub} unsubscribed)")
            if errors:
                with st.expander(f"⚠️ {errors_count} errors during import"):
                    for err in errors:
                        st.caption(f"• {err}")
            try:
                db.log_audit(st.session_state["user_id"], st.session_state["full_name"],
                            "import", "ac_contacts",
                            details=f"{imported} contacts imported ({skipped_unsub} unsubscribed excluded)")
            except:
                pass

        st.markdown("---")

        # ── Customer sync ──
        st.markdown("#### Sync Customers to ActiveCampaign")
        st.info(
            "Customers need an email address to sync. "
            "Only customers with an email stored will be sent to ActiveCampaign. "
            "Use the List ID from the table above to add contacts to a specific list."
        )

        col1, col2 = st.columns(2)
        ac_list_id  = col1.text_input("AC List ID (optional)", key="ac_sync_list",
                                      placeholder="e.g. 1")
        ac_tag_name = col2.text_input("Tag to apply", key="ac_sync_tag",
                                      value="Sales Navigator")

        if st.button("🚀 Sync Customers Now", key="ac_sync_btn", type="primary"):
            with st.spinner("Fetching contacts…"):
                with db.get_conn() as conn:
                    customers = conn.execute("""
                        SELECT cc.customer_code, c.customer_name, c.region,
                               c.customer_type, u.full_name as rep_name,
                               cc.contact_name, cc.email
                        FROM customer_contacts cc
                        LEFT JOIN customers c ON cc.customer_code = c.customer_code
                        LEFT JOIN users u ON c.user_id = u.id
                        WHERE cc.email IS NOT NULL AND cc.email != ''
                    """).fetchall()
                customers = [dict(r) for r in customers]

            if not customers:
                st.warning("No customer contacts found.")
            else:
                st.info(f"Found {len(customers)} customers with email addresses.")
                progress = st.progress(0)
                status   = st.empty()

                def _progress(done, total):
                    progress.progress(done / total)
                    status.text(f"Syncing {done}/{total}…")

                synced, skipped, errors = ac.sync_customers_to_ac(
                    customers,
                    list_id=ac_list_id.strip() or None,
                    tag_name=ac_tag_name.strip() or "Sales Navigator",
                    progress_callback=_progress,
                )
                progress.empty()
                status.empty()
                st.success(f"✅ Synced {synced} contacts. Skipped {skipped} (no email).")
                if errors:
                    with st.expander(f"⚠️ {len(errors)} errors"):
                        for e in errors:
                            st.text(e)

    with tab5:
        st.markdown("### Less Annoying CRM Integration")

        # Credentials
        with st.expander("🔑 API Credentials", expanded=False):
            lc_user = st.text_input("User Code",
                                    value=db.get_setting("lacrm_user_code") or "")
            lc_key  = st.text_input("API Key", type="password",
                                    value=db.get_setting("lacrm_api_key") or "")
            if st.button("Save LACRM Credentials", key="save_lacrm"):
                db.set_setting("lacrm_user_code", lc_user.strip())
                db.set_setting("lacrm_api_key",   lc_key.strip())
                st.success("Saved.")

        # Connection test
        if st.button("🔌 Test LACRM Connection", key="lacrm_test"):
            with st.spinner("Connecting…"):
                ok, msg = lacrm.test_connection()
            if ok:
                st.success(f"✅ {msg}")
            else:
                st.error(f"❌ {msg}")

        st.markdown("---")

        # Sync contacts
        st.markdown("#### Sync Contacts to LACRM")
        st.info("Pushes your 1,500+ customer contacts into Less Annoying CRM. "
                "Existing contacts are updated, new ones are created.")

        if st.button("🚀 Sync to LACRM", key="lacrm_sync", type="primary"):
            with db.get_conn() as conn:
                contacts = conn.execute("""
                    SELECT cc.customer_code, c.customer_name, c.region,
                           c.customer_type, u.full_name as rep_name,
                           cc.contact_name, cc.email
                    FROM customer_contacts cc
                    LEFT JOIN customers c ON cc.customer_code = c.customer_code
                    LEFT JOIN users u ON c.user_id = u.id
                    WHERE cc.email IS NOT NULL AND cc.email != ''
                """).fetchall()
            contacts = [dict(r) for r in contacts]

            if not contacts:
                st.warning("No contacts with email addresses found.")
            else:
                st.info(f"Syncing {len(contacts)} contacts…")
                progress = st.progress(0)
                status   = st.empty()

                def _lc_progress(done, total):
                    progress.progress(done / total)
                    status.text(f"Syncing {done}/{total}…")

                synced, skipped, errors = lacrm.sync_contacts_to_lacrm(
                    contacts, progress_callback=_lc_progress
                )
                progress.empty()
                status.empty()
                st.success(f"✅ Synced {synced} contacts. Skipped {skipped}.")
                if errors:
                    with st.expander(f"⚠️ {len(errors)} errors"):
                        for e in errors:
                            st.text(e)

        st.markdown("---")

        # Tag contacts in group
        st.markdown("#### Tag Contacts in LACRM Group")
        st.info("Adds all Sales Navigator contacts to the 'Sales Navigator' group in LACRM. "
                "Safe to re-run — contacts already in the group are skipped.")

        if st.button("🏷️ Tag All Contacts in LACRM", key="lacrm_tag", type="primary"):
            with db.get_conn() as conn:
                contacts = conn.execute(
                    "SELECT email FROM customer_contacts WHERE email IS NOT NULL AND email != ''"
                ).fetchall()
            emails = [r["email"].strip() for r in contacts]

            st.info(f"Tagging {len(emails)} contacts…")
            progress = st.progress(0)
            status   = st.empty()
            tagged   = [0]; notfound = [0]; errors = [0]
            lock     = threading.Lock()

            def _tag_one(email):
                import requests as _req
                uc = db.get_setting("lacrm_user_code")
                ak = db.get_setting("lacrm_api_key")
                r = _req.get("https://api.lessannoyingcrm.com", params={
                    "UserCode": uc, "APIToken": ak,
                    "Function": "SearchContacts", "SearchTerms": email, "Page": 1
                }, timeout=15)
                results = r.json().get("Result", [])
                contact_id = None
                for c in results:
                    for e in (c.get("Email") or []):
                        if (e.get("Text") or "").lower() == email.lower():
                            contact_id = c["ContactId"]
                            break
                    if contact_id:
                        break
                if not contact_id:
                    return "notfound"
                r2 = _req.get("https://api.lessannoyingcrm.com", params={
                    "UserCode": uc, "APIToken": ak,
                    "Function": "AddContactToGroup",
                    "ContactId": contact_id, "GroupName": "Sales Navigator"
                }, timeout=15)
                return "ok" if r2.json().get("Success") else "error"

            from concurrent.futures import ThreadPoolExecutor, as_completed
            import threading as _threading
            _lock = _threading.Lock()
            _done = [0]

            with ThreadPoolExecutor(max_workers=8) as ex:
                futures = {ex.submit(_tag_one, e): e for e in emails}
                for f in as_completed(futures):
                    result = f.result()
                    with _lock:
                        _done[0] += 1
                        if result == "ok":       tagged[0] += 1
                        elif result == "notfound": notfound[0] += 1
                        else:                    errors[0] += 1
                        progress.progress(_done[0] / len(emails))
                        status.text(f"{_done[0]}/{len(emails)} — Tagged: {tagged[0]}, Not found yet: {notfound[0]}")

            progress.empty()
            status.empty()
            st.success(f"✅ Tagged {tagged[0]} contacts in LACRM 'Sales Navigator' group.")
            if notfound[0]:
                st.caption(f"{notfound[0]} contacts not yet indexed in LACRM — re-run in a day or two to pick them up.")
            if errors[0]:
                st.warning(f"{errors[0]} errors encountered.")

        st.markdown("---")

        # Settings for auto-logging visits
        st.markdown("#### Auto-log Visits to LACRM")
        auto_log = db.get_setting("lacrm_auto_log_visits") == "true"
        new_auto  = st.toggle("Automatically push visit notes to LACRM when a rep logs a visit",
                              value=auto_log, key="lacrm_auto_log")
        if new_auto != auto_log:
            db.set_setting("lacrm_auto_log_visits", "true" if new_auto else "false")
            st.success("Setting saved.")
        if new_auto:
            st.caption("When a rep logs a visit in Sales Navigator, a note will automatically "
                       "be created on the matching LACRM contact.")

    with tab6:
        st.markdown("### Audit Trail")
        st.caption("A log of all actions taken in the system — who changed what and when.")
        af1, af2, af3 = st.columns(3)
        _au_users = db.get_all_users()
        _au_user_opts = ["All Users"] + [u["full_name"] for u in _au_users]
        _au_user_sel = af1.selectbox("Filter by user", _au_user_opts, key="audit_user_filter")
        _au_uid = None
        if _au_user_sel != "All Users":
            _au_uid = next((u["id"] for u in _au_users if u["full_name"] == _au_user_sel), None)
        _ent_opts = ["All Types", "visit", "contact", "note", "task", "test_request",
                     "marketing_lead", "planned_visit", "news", "expense", "customer",
                     "user", "setting", "campaign", "template", "sales_data", "segment", "prospect"]
        _ent_sel = af2.selectbox("Filter by type", _ent_opts, key="audit_entity_filter")
        _au_ent = None if _ent_sel == "All Types" else _ent_sel
        _act_opts = ["All Actions", "create", "update", "delete", "login", "logout", "import", "export"]
        _act_sel = af3.selectbox("Filter by action", _act_opts, key="audit_action_filter")
        _au_act = None if _act_sel == "All Actions" else _act_sel
        _au_total = db.get_audit_log_count(user_id=_au_uid, entity_type=_au_ent, action=_au_act)
        st.markdown(f"**{_au_total}** audit entries found")
        _pg_sz = 50
        _au_pg = st.number_input("Page", min_value=1, max_value=max(1, (_au_total // _pg_sz) + 1),
                                 value=1, key="audit_page")
        _au_off = (_au_pg - 1) * _pg_sz
        _au_entries = db.get_audit_log(limit=_pg_sz, offset=_au_off,
                                       user_id=_au_uid, entity_type=_au_ent, action=_au_act)
        if not _au_entries:
            st.info("No audit entries found for these filters.")
        else:
            _ai = {"create": "🟢", "update": "🔵", "delete": "🔴",
                   "login": "🔑", "logout": "🚪", "import": "📤", "export": "📥"}
            for _e in _au_entries:
                _ic = _ai.get(_e["action"], "⚪")
                _ts = _e["created_at"][:19] if _e["created_at"] else ""
                _un = _e.get("user_name") or _e.get("actor_name") or "Unknown"
                _lb = _e.get("entity_label") or ""
                _dt = _e.get("details") or ""
                _ln = f"{_ic} **{_ts}** — **{_un}** {_e['action']}d {_e['entity_type']}"
                if _lb:
                    _ln += f": *{_lb}*"
                st.markdown(_ln)
                if _dt:
                    st.caption(f"  {_dt}")
            _adf = pd.DataFrame([{
                "Timestamp": e["created_at"],
                "User": e.get("user_name") or e.get("actor_name") or "",
                "Action": e["action"], "Type": e["entity_type"],
                "Label": e.get("entity_label") or "",
                "Details": e.get("details") or "",
            } for e in _au_entries])
            st.download_button("📥 Export Audit Log CSV", data=_adf.to_csv(index=False),
                               file_name="audit_log.csv", mime="text/csv", key="dl_audit")

    with tab7:
        st.markdown("### 🔒 Security Log")
        st.caption("All login attempts — successful and failed — with timestamps and IP addresses.")

        _sec_log = db.get_login_audit(limit=500)
        if not _sec_log:
            st.info("No login attempts recorded yet.")
        else:
            _sec_df = pd.DataFrame([{
                "Timestamp":   e["timestamp"],
                "Username":    e["username"],
                "IP Address":  e["ip_address"] or "unknown",
                "Result":      "✅ Success" if e["success"] else "❌ Failed",
                "Reason":      e["failure_reason"] or "",
            } for e in _sec_log])

            # Summary stats
            total   = len(_sec_df)
            success = _sec_df["Result"].str.startswith("✅").sum()
            failed  = total - success
            locked  = _sec_df["Reason"].str.contains("lock|attempt", case=False, na=False).sum()
            mc1, mc2, mc3, mc4 = st.columns(4)
            mc1.metric("Total Attempts", total)
            mc2.metric("Successful Logins", success)
            mc3.metric("Failed Attempts", failed)
            mc4.metric("Lockout Events", locked)

            st.dataframe(
                _sec_df,
                use_container_width=True,
                hide_index=True,
                column_config={
                    "Result": st.column_config.TextColumn(width="small"),
                    "Timestamp": st.column_config.TextColumn(width="medium"),
                }
            )
            st.download_button(
                "📥 Export Security Log CSV",
                data=_sec_df.to_csv(index=False),
                file_name="security_log.csv",
                mime="text/csv",
                key="dl_security"
            )

        st.divider()
        st.markdown("### 🔐 Unlock Account")
        st.caption("If a user has been locked out, reset their failed attempts here.")
        _all_users = db.get_all_users()
        _unlock_sel = st.selectbox("Select user to unlock", [u["username"] for u in _all_users], key="unlock_user_sel")
        if st.button("Unlock Account", key="unlock_btn"):
            _u = db.get_user_by_username(_unlock_sel)
            if _u:
                db.reset_failed_logins(_u["id"])
                st.success(f"Account unlocked for {_unlock_sel}.")

# ── Main router ───────────────────────────────────────────────────────────────────
# ── Help Chat ────────────────────────────────────────────────────────────────────

HELP_SYSTEM_PROMPT = """You are the built-in help assistant for Tungaloy-NTK Sales Navigator, a Streamlit-based sales management tool for Tungaloy-UK (cutting tools manufacturer).

You know the tool inside out. Here are the features:

**Pages:**
- Home: Dashboard with KPI tiles (customers, monthly sales, visits, alerts, tasks), quick links
- Customers: Full customer list with search, filters by rep/region/alert level/trend. Click View to see customer detail
- Customer Detail: Customer profile with KPIs (last visit, last purchase, spend trend, alert status), reorder alerts, notes, visit/sales/items tabs. PDF download button
- Visit Planner: Plan and log customer visits, route planning
- CRM: Customer 360 view, contacts, notes & interactions, tasks & follow-ups, segments, prospects
- Email Campaigns: Create email templates, send campaigns, track opens/clicks, landing pages, opt-out management
- Alerts: Customer alerts for overdue visits and declining spend
- Test Tools: Log and track cutting tool tests at customer sites
- Marketing Leads: Track inbound marketing leads
- News Board: Internal team announcements
- League Tables: Reps by Sales, Reps by GP, Sales Growth, Applications, Zero-Price Holders, Free Holders — all filterable by SM Name
- Reports: Downloadable PDF reports with filters
- Expenses: Track and submit expense claims
- Team Overview: Management view of team performance (admin/regional_manager only)
- Admin: User management, settings, email config, API integrations (admin only)

**Key concepts:**
- SM Name = Sales Manager territory name (e.g. EAST MIDS-GIBSON, SCOTLAND-W-HAMILTON)
- Customers are grouped by customer_code
- Sales data comes from uploaded GP reports (Excel)
- Reps log visits which link to LACRM
- The tool has ActiveCampaign integration for importing contacts
- Prospects are potential customers imported from ActiveCampaign that aren't yet linked to existing accounts
- Segments let you group customers by criteria (rep, no-order-since, min spend)

**Navigation:** Click page names in the left sidebar. Customer detail pages have a "Back to Customers" button.

**PDF Downloads:** Available on customer detail pages (top right) and in Reports. Tick/untick "Include GP" to control GP data in the PDF.

**Login:** Each user has a username and password. First login may require a password change. Roles: admin, regional_manager, rep, marketing.

Be helpful, concise, and friendly. If you don't know the answer, say so. Don't make things up. Guide users step-by-step when explaining how to do something."""


def page_help():
    st.markdown('<div class="section-header">Help & Support</div>', unsafe_allow_html=True)
    st.caption("Ask anything about the Sales Navigator — how to use features, find data, or troubleshoot issues.")

    api_key = db.get_setting("anthropic_api_key")
    if not api_key:
        st.warning("AI Help is not configured yet. Ask your admin to add an Anthropic API key in Admin → Settings.")
        return

    # Initialise chat history
    if "help_messages" not in st.session_state:
        st.session_state.help_messages = []

    # Render existing messages
    for msg in st.session_state.help_messages:
        with st.chat_message(msg["role"], avatar="🧑" if msg["role"] == "user" else "🤖"):
            st.markdown(msg["content"])

    # Input
    if prompt := st.chat_input("What do you need help with?"):
        # Show user message
        st.session_state.help_messages.append({"role": "user", "content": prompt})
        with st.chat_message("user", avatar="🧑"):
            st.markdown(prompt)

        # Call Claude
        try:
            import anthropic
            client = anthropic.Anthropic(api_key=api_key)

            # Build message list for API
            api_messages = [{"role": m["role"], "content": m["content"]}
                           for m in st.session_state.help_messages]

            with st.chat_message("assistant", avatar="🤖"):
                with st.spinner("Thinking..."):
                    response = client.messages.create(
                        model="claude-sonnet-4-20250514",
                        max_tokens=1024,
                        system=HELP_SYSTEM_PROMPT,
                        messages=api_messages,
                    )
                    reply = response.content[0].text

                st.markdown(reply)

            st.session_state.help_messages.append({"role": "assistant", "content": reply})

        except Exception as e:
            st.error(f"Sorry, I couldn't get a response: {e}")

    # Clear chat button
    if st.session_state.help_messages:
        if st.button("🗑 Clear chat", key="clear_help_chat"):
            st.session_state.help_messages = []
            st.rerun()


def main():
    if not auth.is_logged_in():
        if st.session_state.pop("session_timed_out", False):
            st.warning("⏱️ Your session expired after inactivity. Please sign in again.")
        page_login()
        return

    # ── Session timeout check ────────────────────────────────────────────────────
    if not auth.check_session_timeout():
        auth.logout()
        st.session_state["session_timed_out"] = True
        st.rerun()
        return

    if st.session_state.get("must_change_password"):
        render_sidebar()
        page_change_password()
        return

    render_sidebar()

    page = st.session_state.get("page", "hub")
    role = st.session_state.get("role", "rep")
    user = {
        "id":        st.session_state.get("user_id"),
        "username":  st.session_state.get("username"),
        "full_name": st.session_state.get("full_name"),
        "role":      role,
        "email":     st.session_state.get("email"),
    }

    if page == "hub":
        page_hub()
    elif page == "customers":
        page_customers()
    elif page == "customer_detail":
        page_customer_detail()
    elif page == "log_visit":
        page_log_visit()
    elif page == "alerts":
        page_alerts()
    elif page in ("test_tools", "test_tools_new", "test_tools_detail", "test_tools_dashboard"):
        test_tools.render_page()
    elif page in ("marketing_leads", "marketing_leads_new", "marketing_leads_detail"):
        marketing_leads.render_page()
    elif page in ("visit_planner", "plan_visit_new"):
        visit_planner.render_page()
    elif page in ("news", "news_new"):
        news_board.render_page()
    elif page == "crm":
        crm.render_page(user)
    elif page == "email_campaigns":
        email_campaigns.render_page(user)
    elif page == "league_tables":
        league_tables.render_page()
    elif page == "reports":
        reports.render_page()
    elif page == "expenses":
        expenses.page_expenses()
    elif page == "insert_converter":
        insert_converter.page_insert_converter()
    elif page == "upload" and role in ("admin", "marketing"):
        page_upload()
    elif page == "admin" and role in ("admin", "marketing"):
        page_admin()
    elif page == "help":
        page_help()
    else:
        page_hub()

if __name__ == "__main__":
    main()
