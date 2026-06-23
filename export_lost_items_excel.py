"""
export_lost_items_excel.py

Exports the "Regular items not ordered 2026" dataset to an Excel workbook
with three sheets:
  - Summary          : one row per customer with total items + revenue at risk
  - Detail           : one row per (customer, item) with full history
  - By Sales Manager : rollup by rep (SM Name → full name)

Uses the same SQL + filters as push_lost_items_to_lacrm.py so the numbers
tie back exactly to what was pushed to LACRM.

Run: python export_lost_items_excel.py
"""

import sys
import os
from datetime import datetime
from collections import defaultdict

try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

import database as db

# Match the filter rules in push_lost_items_to_lacrm.py
NON_PRODUCT_KEYWORDS = [
    "FREIGHT", "CARRIAGE", "HANDLING", "ENGINEERING CHARGE",
    "UPLIFT", "ADMIN", "EXTRA CHARGE", "DELIVERY CHARGE",
    "RESTOCK", "SERVICE CHARGE", "MIN ORDER",
]


def _is_real_product(item_code, item_desc):
    code = (item_code or "").strip().upper()
    desc = (item_desc or "").strip().upper()
    if not code:
        return False
    if code.startswith("10000"):
        return False
    for kw in NON_PRODUCT_KEYWORDS:
        if kw in desc:
            return False
    return True


def _fetch_rows():
    sql = """
    WITH item_2025 AS (
        SELECT customer_code, item_code, item_desc,
               COUNT(DISTINCT invoice_no) AS order_count,
               SUM(qty)        AS total_qty,
               SUM(sales_val)  AS total_sales,
               AVG(unit_price) AS avg_price,
               MAX(sale_date)  AS last_order_date
        FROM sales
        WHERE year = 2025
          AND item_code IS NOT NULL AND TRIM(item_code) != ''
          AND sales_val > 0
        GROUP BY customer_code, item_code
        HAVING COUNT(DISTINCT invoice_no) >= 3
    ),
    item_2026 AS (
        SELECT DISTINCT customer_code, item_code
        FROM sales
        WHERE year = 2026 AND item_code IS NOT NULL
    )
    SELECT i.customer_code, c.customer_name,
           c.sm_name,
           COALESCE(u.full_name, '(Unassigned)') AS rep_name,
           i.item_code, i.item_desc,
           i.order_count, i.total_qty, i.total_sales,
           i.avg_price, i.last_order_date
    FROM item_2025 i
    JOIN customers c ON c.customer_code = i.customer_code
    LEFT JOIN sm_name_mapping sm ON sm.sm_name = c.sm_name
    LEFT JOIN users u ON u.id = sm.user_id
    LEFT JOIN item_2026 j
      ON i.customer_code = j.customer_code AND i.item_code = j.item_code
    WHERE j.item_code IS NULL
    ORDER BY c.customer_name, i.total_sales DESC
    """
    with db.get_conn() as conn:
        rows = conn.execute(sql).fetchall()

    # Apply real-product filter
    return [dict(r) for r in rows
            if _is_real_product(r["item_code"], r["item_desc"])]


def build_workbook(rows, out_path):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule

    wb = openpyxl.Workbook()

    # ── Colours / styling ────────────────────────────────────────────
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(border_style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    money_fmt = '£#,##0'
    money_fmt2 = '£#,##0.00'
    int_fmt = '#,##0'

    # ── Summary sheet ────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Summary"

    # Aggregate by customer
    cust_stats = defaultdict(lambda: {"name": "", "items": 0, "revenue": 0.0})
    for r in rows:
        code = r["customer_code"]
        cust_stats[code]["name"] = r["customer_name"]
        cust_stats[code]["items"] += 1
        cust_stats[code]["revenue"] += float(r["total_sales"] or 0)

    cust_sorted = sorted(
        cust_stats.items(),
        key=lambda kv: -kv[1]["revenue"]
    )

    sum_headers = ["Rank", "Customer Code", "Customer Name",
                   "Items Not Re-Ordered", "2025 Revenue at Risk"]
    ws_sum.append(sum_headers)
    for col_idx, _ in enumerate(sum_headers, 1):
        cell = ws_sum.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = border

    for rank, (code, s) in enumerate(cust_sorted, 1):
        ws_sum.append([rank, code, s["name"], s["items"], s["revenue"]])

    total_items   = sum(s["items"]   for _, s in cust_sorted)
    total_revenue = sum(s["revenue"] for _, s in cust_sorted)
    total_row_idx = ws_sum.max_row + 1
    ws_sum.cell(row=total_row_idx, column=1, value="")
    ws_sum.cell(row=total_row_idx, column=2, value="")
    ws_sum.cell(row=total_row_idx, column=3, value="TOTAL").font = Font(bold=True)
    ws_sum.cell(row=total_row_idx, column=4, value=total_items).font = Font(bold=True)
    ws_sum.cell(row=total_row_idx, column=5, value=total_revenue).font = Font(bold=True)

    # Formats + widths
    for row in ws_sum.iter_rows(min_row=2, max_row=ws_sum.max_row):
        row[3].number_format = int_fmt
        row[4].number_format = money_fmt
        for c in row:
            c.border = border

    ws_sum.column_dimensions["A"].width = 6
    ws_sum.column_dimensions["B"].width = 15
    ws_sum.column_dimensions["C"].width = 35
    ws_sum.column_dimensions["D"].width = 22
    ws_sum.column_dimensions["E"].width = 24
    ws_sum.freeze_panes = "A2"

    # Colour scale on the revenue column
    if len(cust_sorted) > 1:
        rng = f"E2:E{ws_sum.max_row - 1}"
        ws_sum.conditional_formatting.add(rng, ColorScaleRule(
            start_type="min", start_color="FFFFFF",
            end_type="max",   end_color="FF6B6B"
        ))

    # ── Detail sheet ─────────────────────────────────────────────────
    ws_det = wb.create_sheet("Detail")
    det_headers = ["Customer Code", "Customer Name", "SM Name", "Rep Name",
                   "Item Code", "Description", "2025 Order Count", "Total Qty",
                   "2025 Sales (£)", "Avg Unit Price (£)", "Last Order Date"]
    ws_det.append(det_headers)
    for col_idx in range(1, len(det_headers) + 1):
        cell = ws_det.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = border

    # Sort rows the same way the LACRM notes were built:
    # customer name ascending, then total_sales descending within each customer
    rows_sorted = sorted(
        rows,
        key=lambda r: (r["customer_name"] or "", -float(r["total_sales"] or 0))
    )

    for r in rows_sorted:
        ws_det.append([
            r["customer_code"],
            r["customer_name"],
            r["sm_name"] or "",
            r["rep_name"] or "(Unassigned)",
            r["item_code"],
            r["item_desc"] or "",
            int(r["order_count"] or 0),
            float(r["total_qty"] or 0),
            float(r["total_sales"] or 0),
            float(r["avg_price"] or 0),
            r["last_order_date"] or "",
        ])

    # Formats (new column positions: 7=OrderCount, 8=Qty, 9=Sales, 10=AvgPrice)
    for row in ws_det.iter_rows(min_row=2, max_row=ws_det.max_row):
        row[6].number_format  = int_fmt
        row[7].number_format  = int_fmt
        row[8].number_format  = money_fmt
        row[9].number_format  = money_fmt2
        for c in row:
            c.border = border

    # Column widths (11 columns now)
    widths = [15, 32, 12, 22, 18, 45, 16, 12, 16, 18, 15]
    for i, w in enumerate(widths, 1):
        ws_det.column_dimensions[get_column_letter(i)].width = w

    ws_det.freeze_panes = "A2"
    ws_det.auto_filter.ref = ws_det.dimensions

    # Colour scale on sales column (column I now)
    last_row = ws_det.max_row
    if last_row > 2:
        rng = f"I2:I{last_row}"
        ws_det.conditional_formatting.add(rng, ColorScaleRule(
            start_type="min", start_color="FFFFFF",
            end_type="max",   end_color="FFA94D"
        ))

    # ── By Sales Manager sheet ───────────────────────────────────────
    ws_sm = wb.create_sheet("By Sales Manager")

    # Aggregate
    sm_stats = defaultdict(lambda: {
        "rep_name": "", "sm_name": "",
        "customers": set(), "items": 0, "revenue": 0.0
    })
    for r in rows:
        key = (r["sm_name"] or "", r["rep_name"] or "(Unassigned)")
        s = sm_stats[key]
        s["sm_name"]  = key[0]
        s["rep_name"] = key[1]
        s["customers"].add(r["customer_code"])
        s["items"]   += 1
        s["revenue"] += float(r["total_sales"] or 0)

    sm_sorted = sorted(
        sm_stats.values(),
        key=lambda s: -s["revenue"]
    )

    sm_headers = ["Rank", "SM Name", "Rep / Sales Manager",
                  "Customers Affected", "Items Not Re-Ordered",
                  "2025 Revenue at Risk", "Avg £ / Customer"]
    ws_sm.append(sm_headers)
    for col_idx in range(1, len(sm_headers) + 1):
        cell = ws_sm.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = border

    for rank, s in enumerate(sm_sorted, 1):
        n_cust = len(s["customers"])
        avg_per_cust = (s["revenue"] / n_cust) if n_cust else 0
        ws_sm.append([
            rank,
            s["sm_name"],
            s["rep_name"],
            n_cust,
            s["items"],
            s["revenue"],
            avg_per_cust,
        ])

    # Totals row
    total_custs_unique = len({c for s in sm_sorted for c in s["customers"]})
    total_items_sm     = sum(s["items"] for s in sm_sorted)
    total_rev_sm       = sum(s["revenue"] for s in sm_sorted)
    tot_row = ws_sm.max_row + 1
    ws_sm.cell(row=tot_row, column=3, value="TOTAL").font = Font(bold=True)
    ws_sm.cell(row=tot_row, column=4, value=total_custs_unique).font = Font(bold=True)
    ws_sm.cell(row=tot_row, column=5, value=total_items_sm).font = Font(bold=True)
    ws_sm.cell(row=tot_row, column=6, value=total_rev_sm).font = Font(bold=True)

    # Formats + borders
    for row in ws_sm.iter_rows(min_row=2, max_row=ws_sm.max_row):
        if row[3].value is not None: row[3].number_format = int_fmt
        if row[4].value is not None: row[4].number_format = int_fmt
        if row[5].value is not None: row[5].number_format = money_fmt
        if row[6].value is not None: row[6].number_format = money_fmt
        for c in row:
            c.border = border

    ws_sm.column_dimensions["A"].width = 6
    ws_sm.column_dimensions["B"].width = 12
    ws_sm.column_dimensions["C"].width = 26
    ws_sm.column_dimensions["D"].width = 20
    ws_sm.column_dimensions["E"].width = 22
    ws_sm.column_dimensions["F"].width = 22
    ws_sm.column_dimensions["G"].width = 20
    ws_sm.freeze_panes = "A2"

    # Colour scale on revenue column
    if len(sm_sorted) > 1:
        rng = f"F2:F{ws_sm.max_row - 1}"
        ws_sm.conditional_formatting.add(rng, ColorScaleRule(
            start_type="min", start_color="FFFFFF",
            end_type="max",   end_color="FF6B6B"
        ))

    wb.save(out_path)


def main():
    print("=" * 60)
    print("Regular Items Not Ordered 2026 — Excel Export")
    print("=" * 60)

    rows = _fetch_rows()
    unique_custs = len({r["customer_code"] for r in rows})
    total_val = sum(float(r["total_sales"] or 0) for r in rows)
    print(f"Rows (items)      : {len(rows):,}")
    print(f"Unique customers  : {unique_custs}")
    print(f"2025 revenue      : £{total_val:,.0f}")

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_dir = r"C:\Users\rob.werhun\IMC\Tungaloy-UK - Public\Robs ondrive data\Financial\Budget\SalesNavigator"
    out_path = os.path.join(out_dir, f"Regular_items_not_ordered_2026_{ts}.xlsx")

    build_workbook(rows, out_path)
    print(f"\n[OK] Saved: {out_path}")


if __name__ == "__main__":
    main()
