"""
export_lost_items_per_rep.py

Generates one Excel file per Sales Manager containing only their customers'
"Regular items not ordered 2026" data — ready to email each rep individually.

Files go into:  Financial/Budget/SalesNavigator/rep_packs_<timestamp>/

Each file has two sheets:
  - Summary : that rep's customers, ranked by revenue at risk
  - Detail  : one row per (customer, item) — full order history, auto-filtered

Numbers tie back to push_lost_items_to_lacrm.py and export_lost_items_excel.py.

Run: python export_lost_items_per_rep.py
"""

import sys
import os
import re
from datetime import datetime
from collections import defaultdict

try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

import database as db

NON_PRODUCT_KEYWORDS = [
    "FREIGHT", "CARRIAGE", "HANDLING", "ENGINEERING CHARGE",
    "UPLIFT", "ADMIN", "EXTRA CHARGE", "DELIVERY CHARGE",
    "RESTOCK", "SERVICE CHARGE", "MIN ORDER",
]


def _is_real_product(item_code, item_desc):
    code = (item_code or "").strip().upper()
    desc = (item_desc or "").strip().upper()
    if not code:
        return False
    if code.startswith("10000"):
        return False
    for kw in NON_PRODUCT_KEYWORDS:
        if kw in desc:
            return False
    return True


def _fetch_rows():
    sql = """
    WITH item_2025 AS (
        SELECT customer_code, item_code, item_desc,
               COUNT(DISTINCT invoice_no) AS order_count,
               SUM(qty)        AS total_qty,
               SUM(sales_val)  AS total_sales,
               AVG(unit_price) AS avg_price,
               MAX(sale_date)  AS last_order_date
        FROM sales
        WHERE year = 2025
          AND item_code IS NOT NULL AND TRIM(item_code) != ''
          AND sales_val > 0
        GROUP BY customer_code, item_code
        HAVING COUNT(DISTINCT invoice_no) >= 3
    ),
    item_2026 AS (
        SELECT DISTINCT customer_code, item_code
        FROM sales
        WHERE year = 2026 AND item_code IS NOT NULL
    )
    SELECT i.customer_code, c.customer_name,
           c.sm_name,
           COALESCE(u.full_name, '(Unassigned)') AS rep_name,
           COALESCE(u.email, '')                 AS rep_email,
           i.item_code, i.item_desc,
           i.order_count, i.total_qty, i.total_sales,
           i.avg_price, i.last_order_date
    FROM item_2025 i
    JOIN customers c ON c.customer_code = i.customer_code
    LEFT JOIN sm_name_mapping sm ON sm.sm_name = c.sm_name
    LEFT JOIN users u ON u.id = sm.user_id
    LEFT JOIN item_2026 j
      ON i.customer_code = j.customer_code AND i.item_code = j.item_code
    WHERE j.item_code IS NULL
    ORDER BY c.customer_name, i.total_sales DESC
    """
    with db.get_conn() as conn:
        rows = conn.execute(sql).fetchall()

    return [dict(r) for r in rows
            if _is_real_product(r["item_code"], r["item_desc"])]


def _safe_filename(s):
    """Make a rep name safe for a filename."""
    s = re.sub(r'[^\w\s-]', '', s).strip()
    s = re.sub(r'\s+', '_', s)
    return s or "Unknown"


def build_rep_workbook(rep_name, rep_email, rep_rows, out_path):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule

    wb = openpyxl.Workbook()

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(border_style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    money_fmt  = '£#,##0'
    money_fmt2 = '£#,##0.00'
    int_fmt    = '#,##0'

    # ── Summary sheet ────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Summary"

    # Intro block
    ws_sum.cell(row=1, column=1, value="Regular Items Not Re-Ordered in 2026").font = Font(bold=True, size=14)
    ws_sum.cell(row=2, column=1, value=f"Sales Manager: {rep_name}").font = Font(bold=True)
    if rep_email:
        ws_sum.cell(row=3, column=1, value=f"Email: {rep_email}")
    ws_sum.cell(row=4, column=1, value=f"Generated: {datetime.now().strftime('%d %b %Y %H:%M')}")
    ws_sum.cell(row=5, column=1,
                value="These are items your customers ordered 3+ times in 2025 but haven't re-ordered in 2026.")

    # Aggregate by customer
    cust_stats = defaultdict(lambda: {
        "name": "", "sm_name": "",
        "items": 0, "revenue": 0.0
    })
    for r in rep_rows:
        code = r["customer_code"]
        cust_stats[code]["name"]    = r["customer_name"]
        cust_stats[code]["sm_name"] = r["sm_name"] or ""
        cust_stats[code]["items"]  += 1
        cust_stats[code]["revenue"] += float(r["total_sales"] or 0)

    cust_sorted = sorted(
        cust_stats.items(),
        key=lambda kv: -kv[1]["revenue"]
    )

    header_row = 7
    sum_headers = ["Rank", "Customer Code", "Customer Name",
                   "SM Name", "Items Not Re-Ordered", "2025 Revenue at Risk"]
    for col_idx, h in enumerate(sum_headers, 1):
        cell = ws_sum.cell(row=header_row, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = border

    data_start = header_row + 1
    for rank, (code, s) in enumerate(cust_sorted, 1):
        r_idx = data_start + rank - 1
        ws_sum.cell(row=r_idx, column=1, value=rank)
        ws_sum.cell(row=r_idx, column=2, value=code)
        ws_sum.cell(row=r_idx, column=3, value=s["name"])
        ws_sum.cell(row=r_idx, column=4, value=s["sm_name"])
        ws_sum.cell(row=r_idx, column=5, value=s["items"])
        ws_sum.cell(row=r_idx, column=6, value=s["revenue"])

    # Totals
    total_items   = sum(s["items"]   for _, s in cust_sorted)
    total_revenue = sum(s["revenue"] for _, s in cust_sorted)
    tot_row = data_start + len(cust_sorted)
    ws_sum.cell(row=tot_row, column=3, value="TOTAL").font = Font(bold=True)
    ws_sum.cell(row=tot_row, column=5, value=total_items).font   = Font(bold=True)
    ws_sum.cell(row=tot_row, column=6, value=total_revenue).font = Font(bold=True)

    # Formats
    for row in ws_sum.iter_rows(min_row=data_start, max_row=tot_row):
        if row[4].value is not None: row[4].number_format = int_fmt
        if row[5].value is not None: row[5].number_format = money_fmt
        for c in row:
            c.border = border

    ws_sum.column_dimensions["A"].width = 6
    ws_sum.column_dimensions["B"].width = 15
    ws_sum.column_dimensions["C"].width = 35
    ws_sum.column_dimensions["D"].width = 18
    ws_sum.column_dimensions["E"].width = 22
    ws_sum.column_dimensions["F"].width = 24
    ws_sum.freeze_panes = f"A{data_start}"

    if len(cust_sorted) > 1:
        rng = f"F{data_start}:F{tot_row - 1}"
        ws_sum.conditional_formatting.add(rng, ColorScaleRule(
            start_type="min", start_color="FFFFFF",
            end_type="max",   end_color="FF6B6B"
        ))

    # ── Detail sheet ─────────────────────────────────────────────────
    ws_det = wb.create_sheet("Detail")
    det_headers = ["Customer Code", "Customer Name", "SM Name",
                   "Item Code", "Description", "2025 Order Count",
                   "Total Qty", "2025 Sales (£)", "Avg Unit Price (£)",
                   "Last Order Date"]
    ws_det.append(det_headers)
    for col_idx in range(1, len(det_headers) + 1):
        cell = ws_det.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = border

    rows_sorted = sorted(
        rep_rows,
        key=lambda r: (r["customer_name"] or "", -float(r["total_sales"] or 0))
    )

    for r in rows_sorted:
        ws_det.append([
            r["customer_code"],
            r["customer_name"],
            r["sm_name"] or "",
            r["item_code"],
            r["item_desc"] or "",
            int(r["order_count"] or 0),
            float(r["total_qty"] or 0),
            float(r["total_sales"] or 0),
            float(r["avg_price"] or 0),
            r["last_order_date"] or "",
        ])

    for row in ws_det.iter_rows(min_row=2, max_row=ws_det.max_row):
        row[5].number_format = int_fmt
        row[6].number_format = int_fmt
        row[7].number_format = money_fmt
        row[8].number_format = money_fmt2
        for c in row:
            c.border = border

    widths = [15, 32, 14, 18, 45, 16, 12, 16, 18, 15]
    for i, w in enumerate(widths, 1):
        ws_det.column_dimensions[get_column_letter(i)].width = w

    ws_det.freeze_panes = "A2"
    ws_det.auto_filter.ref = ws_det.dimensions

    last_row = ws_det.max_row
    if last_row > 2:
        rng = f"H2:H{last_row}"
        ws_det.conditional_formatting.add(rng, ColorScaleRule(
            start_type="min", start_color="FFFFFF",
            end_type="max",   end_color="FFA94D"
        ))

    wb.save(out_path)


def main():
    print("=" * 60)
    print("Per-Rep Excel Export — Regular Items Not Ordered 2026")
    print("=" * 60)

    rows = _fetch_rows()
    unique_custs = len({r["customer_code"] for r in rows})
    total_val = sum(float(r["total_sales"] or 0) for r in rows)
    print(f"Rows (items)      : {len(rows):,}")
    print(f"Unique customers  : {unique_custs}")
    print(f"2025 revenue      : £{total_val:,.0f}")
    print()

    # Group by rep (NOT by sm_name territory — one file per actual person)
    by_rep = defaultdict(list)
    rep_email = {}
    for r in rows:
        rep = r["rep_name"] or "(Unassigned)"
        by_rep[rep].append(r)
        if r.get("rep_email"):
            rep_email[rep] = r["rep_email"]

    # Sort reps by revenue descending
    rep_sorted = sorted(
        by_rep.items(),
        key=lambda kv: -sum(float(r["total_sales"] or 0) for r in kv[1])
    )

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_dir = os.path.join(
        r"C:\Users\rob.werhun\IMC\Tungaloy-UK - Public\Robs ondrive data\Financial\Budget\SalesNavigator",
        f"rep_packs_{ts}"
    )
    os.makedirs(out_dir, exist_ok=True)
    print(f"Output folder: {out_dir}\n")

    for rep_name, rep_rows in rep_sorted:
        n_cust   = len({r["customer_code"] for r in rep_rows})
        n_items  = len(rep_rows)
        rev      = sum(float(r["total_sales"] or 0) for r in rep_rows)
        email    = rep_email.get(rep_name, "")

        fname = f"Lost_Items_2026_{_safe_filename(rep_name)}.xlsx"
        fpath = os.path.join(out_dir, fname)
        build_rep_workbook(rep_name, email, rep_rows, fpath)

        print(f"  [OK] {rep_name:<22} — {n_cust:>3} custs, {n_items:>4} items, £{rev:>9,.0f}  →  {fname}")

    print()
    print(f"[DONE] {len(rep_sorted)} files written to {out_dir}")


if __name__ == "__main__":
    main()
