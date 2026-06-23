"""
Insert Converter — finds Tungaloy alternatives for competitor cutting inserts.
Tabs: Insert Lookup | Bulk Upload | Grade Converter | Geometry Converter
"""
import re
import io
import functools
import os
from datetime import datetime

import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

DB_PATH = os.path.join(os.path.dirname(__file__), "insert_converter_db.xlsx")
WEBSHOP = "https://webshop.tungaloyuk.co.uk"

ISO_COLOURS = {
    "P": "#4a90d9", "M": "#f5a623", "K": "#e74c3c",
    "N": "#27ae60", "S": "#9b59b6", "H": "#f39c12",
}
ISO_LABELS = {
    "P": "P — Steel", "M": "M — Stainless", "K": "K — Cast Iron",
    "N": "N — Non-ferrous", "S": "S — Super alloys", "H": "H — Hardened",
}

# ── Data loading ──────────────────────────────────────────────────────────────

@functools.lru_cache(maxsize=1)
def _load_data():
    xl = pd.ExcelFile(DB_PATH)

    # Companies
    comp = pd.read_excel(xl, "CompanyMast", header=1)
    comp.columns = ["CompanyLangID", "CompanyID", "CompanySeq", "CompanyName", "CONCAT"]
    comp = comp[comp["CompanyLangID"] == "100 EN"][["CompanyID", "CompanyName"]].copy()
    comp["CompanyID"] = comp["CompanyID"].astype(int)

    # Grades
    grd = pd.read_excel(xl, "Grade", header=1)
    grd.columns = ["GradeCompID", "GradeAppID", "GradeIso", "GradeGrade",
                   "H1","H2","H3","H4","H5","H6","H7","H8","H9","H10"]
    grd = grd[grd["GradeCompID"] != "GradeCompID"].copy()
    grd["CompanyID"] = grd["GradeCompID"].str.extract(r"^(\d+)").astype(float).astype("Int64")
    grd["AppID"]     = grd["GradeAppID"].str.extract(r"^(\d+)").astype(float).astype("Int64")

    # Geometries
    geo = pd.read_excel(xl, "Geometry", header=1)
    geo.columns = ["GeoCompID", "GeoAppID", "GeoIso", "GeoGeometry",
                   "H1","H2","H3","H4","H5","H6","H7","H8","H9","H10"]
    geo = geo[geo["GeoCompID"] != "GeometryCompID"].copy()
    geo["CompanyID"] = geo["GeoCompID"].str.extract(r"^(\d+)").astype(float).astype("Int64")
    geo["AppID"]     = geo["GeoAppID"].str.extract(r"^(\d+)").astype(float).astype("Int64")

    return comp, grd, geo


@functools.lru_cache(maxsize=1)
def _grade_to_company():
    """Reverse lookup: competitor grade code → CompanyID."""
    _, grd, _ = _load_data()
    out = {}
    for _, row in grd.iterrows():
        g = str(row["GradeGrade"]).strip().upper()
        if g and g != "GRADEGRADE" and pd.notna(row["CompanyID"]):
            out[g] = int(row["CompanyID"])
    return out


# ── Parsing & matching ────────────────────────────────────────────────────────

def _parse_insert(text: str):
    """
    Parse an insert code like 'CCMT 060202-F3M IC807' or 'CNMG 120408-PM 4425'.
    Returns (iso_base, chipbreaker, grade)  or (None,None,None) on failure.
    iso_base  e.g. 'CCMT060202'
    chipbreaker e.g. 'F3M' or 'PM'  (may be None)
    grade     e.g. 'IC807' or '4425'  (may be None)
    """
    t = text.strip().upper()
    # Shape(4) + optional space + size(6-8) + optional -chipbreaker + optional grade
    # More flexible: allows chipbreaker and grade with numbers or letters
    m = re.match(
        r'^([A-Z]{4})\s*(\d{6,8})'           # CCMT060202 or CNMG120408
        r'(?:\s*-\s*([A-Z0-9/]+))?'           # -F3M or -PM (optional)
        r'(?:\s+([A-Z0-9]+))?$',              # IC807 or 4425 (optional, can start with letter or digit)
        t,
    )
    if not m:
        return None, None, None
    iso_base    = m.group(1) + m.group(2)
    chipbreaker = m.group(3)
    grade       = m.group(4)
    return iso_base, chipbreaker, grade


def _get_iso_material(iso_base: str) -> str | None:
    """Infer ISO material group from insert shape prefix heuristic (rough)."""
    # We actually derive it from the database matches, not from the code itself
    return None


def _insert_type(iso_base: str) -> str:
    """
    Guess Negative (1130) or Positive (1140) from the relief-angle letter (pos 1).
    N/A → negative; B/C/D/E/F/G/P → positive.
    """
    if len(iso_base) < 2:
        return "both"
    relief = iso_base[1]
    if relief in ("N", "A"):
        return "negative"
    elif relief in ("B", "C", "D", "E", "F", "G", "P"):
        return "positive"
    return "both"


def _collect_alts(row, h_cols=("H1","H2","H3","H4","H5","H6","H7","H8","H9","H10")):
    """Return list of non-null alternative values with H5 flagged as first choice."""
    alts = []
    for col in h_cols:
        v = row.get(col)
        if pd.notna(v) and str(v).strip():
            alts.append((str(v).strip(), col == "H5"))
    return alts


def _is_code(val: str) -> bool:
    """Distinguish a real chip-breaker code from a description like 'All round'."""
    return bool(val) and len(val) <= 6 and " " not in val


def _find_alternatives(iso_base, chipbreaker, grade):
    """
    Core lookup. Returns dict:
      competitor_name, grade_alts, geo_alts, part_numbers, webshop_urls
    """
    comp_df, grd_df, geo_df = _load_data()
    g2c = _grade_to_company()

    result = {
        "competitor":    None,
        "iso_material":  None,
        "grade_alts":    [],   # list of (grade_code, is_first_choice)
        "geo_alts":      [],   # list of (chipbreaker_code, is_first_choice)
        "part_numbers":  [],   # list of full Tungaloy part-number strings
        "webshop_urls":  [],   # matching search links
        "warnings":      [],
    }

    # ── Identify competitor ──
    comp_id = None
    if grade:
        g_upper = grade.upper()
        comp_id = g2c.get(g_upper)

    # If not found, try adding common prefixes (GC for Sandvik, etc.)
    if comp_id is None and grade:
        grade_upper = grade.upper()
        # Try with common prefixes if the grade is purely numeric or starts with a digit
        if grade_upper[0].isdigit():
            prefixes_to_try = ["GC", "HC", "SM"]  # Sandvik uses GC; others may use HC, SM, etc.
            for pfx in prefixes_to_try:
                comp_id = g2c.get(pfx + grade_upper)
                if comp_id is not None:
                    break

    # If still not found, try prefix fallback (first 2 chars of provided grade)
    if comp_id is None and grade:
        pfx = grade[:2].upper()
        for g, cid in g2c.items():
            if g.startswith(pfx):
                comp_id = cid
                break

    if comp_id is None:
        result["warnings"].append(f"Could not identify competitor from grade '{grade}'.")
    else:
        row = comp_df[comp_df["CompanyID"] == comp_id]
        result["competitor"] = row["CompanyName"].values[0] if not row.empty else str(comp_id)

    # ── Grade alternatives ──
    if grade and comp_id is not None:
        g_rows = grd_df[
            (grd_df["CompanyID"] == comp_id) &
            (grd_df["GradeGrade"].astype(str).str.upper() == grade.upper()) &
            (grd_df["AppID"] == 1030)          # ISO turn insert
        ]
        if g_rows.empty:
            # Try all applications
            g_rows = grd_df[
                (grd_df["CompanyID"] == comp_id) &
                (grd_df["GradeGrade"].astype(str).str.upper() == grade.upper())
            ]
        if not g_rows.empty:
            result["iso_material"] = str(g_rows.iloc[0]["GradeIso"]).strip()
            alts = _collect_alts(g_rows.iloc[0])
            result["grade_alts"] = alts
        else:
            result["warnings"].append(f"Grade '{grade}' not found for {result['competitor']}.")

    # ── Geometry alternatives ──
    if chipbreaker and comp_id is not None:
        insert_t = _insert_type(iso_base or "")
        app_ids_pref = []
        if insert_t in ("negative", "both"):
            app_ids_pref.append(1130)
        if insert_t in ("positive", "both"):
            app_ids_pref.append(1140)
        if not app_ids_pref:
            app_ids_pref = [1130, 1140]

        iso_mat = result["iso_material"]
        cb_upper = chipbreaker.upper()

        # Try progressively relaxed filters until we find a match
        geo_rows = pd.DataFrame()
        attempts = [
            # (app_ids, iso_filter)
            (app_ids_pref, iso_mat),          # preferred insert type + ISO match
            ([1130, 1140],  iso_mat),          # both insert types + ISO match
            (app_ids_pref, None),              # preferred insert type, any ISO
            ([1130, 1140],  None),             # fully relaxed
        ]
        for a_ids, i_mat in attempts:
            mask = (
                (geo_df["CompanyID"] == comp_id) &
                (geo_df["GeoGeometry"].astype(str).str.upper() == cb_upper) &
                (geo_df["AppID"].isin(a_ids))
            )
            if i_mat:
                mask &= geo_df["GeoIso"].astype(str).str.upper() == i_mat.upper()
            geo_rows = geo_df[mask]
            if not geo_rows.empty:
                break

        if not geo_rows.empty:
            alts = _collect_alts(geo_rows.iloc[0])
            result["geo_alts"] = [(v, fc) for v, fc in alts if _is_code(v)]
        else:
            result["warnings"].append(f"Geometry '{chipbreaker}' not found for {result['competitor']}.")

    # ── Construct part numbers ──
    if iso_base:
        # Prioritize first-choice alternatives (marked with ★)
        # Put all first-choice items first, then others
        first_choice_grades = [v for v, fc in result["grade_alts"] if fc]
        other_grades = [v for v, fc in result["grade_alts"] if not fc]
        grades_to_use = (first_choice_grades + other_grades)[:3]

        first_choice_geos = [v for v, fc in result["geo_alts"] if fc]
        other_geos = [v for v, fc in result["geo_alts"] if not fc]
        geo_to_use = (first_choice_geos + other_geos)[:2]

        # First choice grade + first choice geometry (up to 2 combinations)
        combos = []
        if grades_to_use and geo_to_use:
            combos.append(f"{iso_base}-{geo_to_use[0]} {grades_to_use[0]}")
            # Alt 2: same geo, next grade (or same grade, next geo)
            if len(grades_to_use) > 1:
                combos.append(f"{iso_base}-{geo_to_use[0]} {grades_to_use[1]}")
            elif len(geo_to_use) > 1:
                combos.append(f"{iso_base}-{geo_to_use[1]} {grades_to_use[0]}")
        elif grades_to_use:
            combos.append(f"{iso_base} {grades_to_use[0]}")
            if len(grades_to_use) > 1:
                combos.append(f"{iso_base} {grades_to_use[1]}")
        elif geo_to_use:
            combos.append(f"{iso_base}-{geo_to_use[0]}")

        result["part_numbers"] = combos
        result["webshop_urls"] = [
            f"{WEBSHOP}/search?q={pn.replace(' ', '+')}" for pn in combos
        ]

    return result


# ── UI helpers ─────────────────────────────────────────────────────────────────

def _alt_chips(alts, max_show=5):
    """Render grade/geometry alternative chips as HTML."""
    if not alts:
        return "<span style='color:#aaa'>—</span>"
    parts = []
    for label, is_fc in alts[:max_show]:
        if is_fc:
            parts.append(
                f"<span style='background:#1a3a5c;color:#fff;padding:3px 8px;"
                f"border-radius:4px;font-weight:700;margin:2px;display:inline-block'>"
                f"★ {label}</span>"
            )
        else:
            parts.append(
                f"<span style='background:#eef2f7;color:#333;padding:3px 8px;"
                f"border-radius:4px;margin:2px;display:inline-block'>{label}</span>"
            )
    return " ".join(parts)


def _render_result_card(label, res):
    """Render a single lookup result as a card."""
    st.markdown(f"**{label}**")
    if res["competitor"]:
        col_a, col_b = st.columns(2)
        col_a.markdown(f"**Competitor:** {res['competitor']}")
        iso = res["iso_material"]
        if iso:
            colour = ISO_COLOURS.get(iso, "#888")
            col_b.markdown(
                f"**ISO material:** <span style='background:{colour};color:#fff;"
                f"padding:2px 8px;border-radius:4px;font-weight:700'>{iso}</span>",
                unsafe_allow_html=True,
            )

    if res["grade_alts"]:
        st.markdown("**Tungaloy Grade Alternatives:**")
        st.markdown(_alt_chips(res["grade_alts"]), unsafe_allow_html=True)

    if res["geo_alts"]:
        st.markdown("**Tungaloy Chip-Breaker Alternatives:**")
        st.markdown(_alt_chips(res["geo_alts"]), unsafe_allow_html=True)

    if res["part_numbers"]:
        st.markdown("**Suggested Tungaloy Part Numbers:**")
        for i, (pn, url) in enumerate(zip(res["part_numbers"], res["webshop_urls"]), 1):
            c1, c2 = st.columns([3, 2])
            c1.markdown(f"`{pn}`")
            c2.markdown(
                f"<a href='{url}' target='_blank' rel='noopener noreferrer'>"
                f"<button style='background:#1a3a5c;color:#fff;border:none;padding:4px 12px;"
                f"border-radius:4px;cursor:pointer'>🔍 Check Webshop (Alt {i})</button></a>",
                unsafe_allow_html=True,
            )
    elif not res["warnings"]:
        st.info("No Tungaloy part number could be constructed — check grade/geometry tabs manually.")

    for w in res["warnings"]:
        st.warning(w)


# ── Main pages ────────────────────────────────────────────────────────────────

def page_insert_converter():
    st.markdown("## 🔄 Insert Converter")
    st.caption("Find Tungaloy alternatives for competitor inserts")

    try:
        comp_df, grd_df, geo_df = _load_data()
    except Exception as e:
        st.error(f"Could not load converter database: {e}")
        return

    company_map  = dict(zip(comp_df["CompanyID"], comp_df["CompanyName"]))
    company_opts = sorted(comp_df["CompanyName"].tolist())

    tab1, tab2, tab3, tab4 = st.tabs([
        "🔍 Insert Lookup",
        "📋 Bulk Upload",
        "🏷️ Grade Converter",
        "🔷 Geometry Converter",
    ])

    # ══════════════════════════════════════════════════════════════════════════
    # TAB 1 — Insert Lookup (single)
    # ══════════════════════════════════════════════════════════════════════════
    with tab1:
        st.markdown("#### Type a competitor insert and find the Tungaloy alternative")
        st.caption("Format: `CCMT 060202-F3M IC807` &nbsp;|&nbsp; `VNMG 160408-M3M IC6015` &nbsp;|&nbsp; `DNMG 110404-PP IC20`")

        code_input = st.text_input("Insert Code", placeholder="CCMT 060202-F3M IC807", key="lookup_code")

        if st.button("🔍 Find Alternative", key="lookup_btn", type="primary"):
            if not code_input.strip():
                st.warning("Please enter an insert code.")
            else:
                iso_base, chipbreaker, grade = _parse_insert(code_input)
                if iso_base is None:
                    st.error(
                        "Could not parse that insert code. "
                        "Expected format: `CCMT 060202-F3M IC807` (Shape + Size + Chipbreaker + Grade)"
                    )
                else:
                    with st.spinner("Looking up alternatives..."):
                        res = _find_alternatives(iso_base, chipbreaker, grade)

                    st.markdown("---")
                    # Show parsed components
                    pc1, pc2, pc3 = st.columns(3)
                    pc1.metric("ISO Base", iso_base)
                    pc2.metric("Chipbreaker", chipbreaker or "—")
                    pc3.metric("Grade", grade or "—")
                    st.markdown("---")

                    _render_result_card(f"Results for: **{code_input.strip().upper()}**", res)

                    if res["part_numbers"]:
                        st.info(
                            "💡 Click **Check Webshop** to verify the alternative exists and see "
                            "live stock availability on the Tungaloy UK webshop."
                        )

    # ══════════════════════════════════════════════════════════════════════════
    # TAB 2 — Bulk Upload
    # ══════════════════════════════════════════════════════════════════════════
    with tab2:
        st.markdown("#### Upload competitor insert list — get Tungaloy alternatives for all")
        st.caption(
            "Upload the template Excel file. The **Competitor** column should contain "
            "insert codes like `CCMT 060202-F3M IC807`, one per row."
        )

        uploaded = st.file_uploader("Upload template (.xlsx)", type=["xlsx"], key="bulk_upload")

        if uploaded:
            try:
                ul_df = pd.read_excel(uploaded, sheet_name=None)
                # Find the sheet with a 'Competitor' column
                target_sheet = None
                for sheet_name, sheet_df in ul_df.items():
                    if any("competitor" in str(c).lower() for c in sheet_df.columns):
                        target_sheet = sheet_name
                        break
                if target_sheet is None:
                    st.error("Could not find a 'Competitor' column in the uploaded file. Please use the template.")
                else:
                    df = ul_df[target_sheet].copy()
                    comp_col = next(c for c in df.columns if "competitor" in str(c).lower())
                    inserts = df[comp_col].dropna().astype(str).str.strip()
                    inserts = inserts[inserts != ""].tolist()

                    st.success(f"Found **{len(inserts)} inserts** to process.")

                    if st.button("⚙️ Process All Inserts", key="bulk_process", type="primary"):
                        rows = []
                        prog = st.progress(0)
                        status = st.empty()

                        for i, code in enumerate(inserts):
                            status.text(f"Processing {i+1}/{len(inserts)}: {code}")
                            iso_base, chipbreaker, grade = _parse_insert(code)

                            if iso_base is None:
                                rows.append({
                                    "Competitor Insert":   code,
                                    "Parsed":              "❌ Could not parse",
                                    "Competitor":          "",
                                    "ISO Material":        "",
                                    "Alt 1 Part No.":      "",
                                    "Alt 1 Webshop Link":  "",
                                    "Alt 2 Part No.":      "",
                                    "Alt 2 Webshop Link":  "",
                                    "Notes":               "Unrecognised format",
                                })
                                prog.progress((i + 1) / len(inserts))
                                continue

                            res = _find_alternatives(iso_base, chipbreaker, grade)

                            alt1_pn  = res["part_numbers"][0] if len(res["part_numbers"]) > 0 else ""
                            alt1_url = res["webshop_urls"][0] if len(res["webshop_urls"]) > 0 else ""
                            alt2_pn  = res["part_numbers"][1] if len(res["part_numbers"]) > 1 else ""
                            alt2_url = res["webshop_urls"][1] if len(res["webshop_urls"]) > 1 else ""

                            rows.append({
                                "Competitor Insert":   code,
                                "Parsed":              f"{iso_base} | CB:{chipbreaker} | Gr:{grade}",
                                "Competitor":          res["competitor"] or "Unknown",
                                "ISO Material":        res["iso_material"] or "",
                                "Alt 1 Part No.":      alt1_pn,
                                "Alt 1 Webshop Link":  alt1_url,
                                "Alt 2 Part No.":      alt2_pn,
                                "Alt 2 Webshop Link":  alt2_url,
                                "Notes":               "; ".join(res["warnings"]),
                            })
                            prog.progress((i + 1) / len(inserts))

                        status.text("✅ All done!")
                        out_df = pd.DataFrame(rows)

                        # ── Show results ──
                        st.markdown(f"### Results — {len(out_df)} inserts processed")
                        found    = out_df[out_df["Alt 1 Part No."] != ""]
                        no_match = out_df[out_df["Alt 1 Part No."] == ""]

                        st.markdown(
                            f"✅ **{len(found)}** matched &nbsp;&nbsp; "
                            f"⚠️ **{len(no_match)}** no match found"
                        )

                        # Display with clickable links
                        for _, row in out_df.iterrows():
                            with st.expander(f"{'✅' if row['Alt 1 Part No.'] else '⚠️'}  {row['Competitor Insert']}"):
                                cols = st.columns([2, 1, 3, 2])
                                cols[0].write(f"**Competitor:** {row['Competitor']}")
                                cols[1].write(f"**ISO:** {row['ISO Material']}")
                                if row["Alt 1 Part No."]:
                                    cols[2].write(f"**Alt 1:** `{row['Alt 1 Part No.']}`")
                                    cols[3].markdown(
                                        f"<a href='{row['Alt 1 Webshop Link']}' target='_blank' rel='noopener noreferrer'>"
                                        f"<button style='background:#1a3a5c;color:#fff;border:none;"
                                        f"padding:4px 12px;border-radius:4px;cursor:pointer'>"
                                        f"🔍 Webshop Alt 1</button></a>",
                                        unsafe_allow_html=True,
                                    )
                                if row["Alt 2 Part No."]:
                                    c2a, c2b = st.columns([3, 2])
                                    c2a.write(f"**Alt 2:** `{row['Alt 2 Part No.']}`")
                                    c2b.markdown(
                                        f"<a href='{row['Alt 2 Webshop Link']}' target='_blank' rel='noopener noreferrer'>"
                                        f"<button style='background:#2c6e9e;color:#fff;border:none;"
                                        f"padding:4px 12px;border-radius:4px;cursor:pointer'>"
                                        f"🔍 Webshop Alt 2</button></a>",
                                        unsafe_allow_html=True,
                                    )
                                if row["Notes"]:
                                    st.caption(f"⚠️ {row['Notes']}")

                        # ── Excel download (Professional export) ──
                        st.markdown("---")
                        st.markdown("#### Download Results")

                        buf = io.BytesIO()
                        wb = Workbook()
                        ws = wb.active
                        ws.title = "Results"

                        # ── Header with logo ──
                        try:
                            logo_path = os.path.join(os.path.dirname(__file__), "logo.png")
                            if os.path.exists(logo_path):
                                img = XLImage(logo_path)
                                img.width = 80
                                img.height = 80
                                ws.add_image(img, "A1")
                        except Exception:
                            pass  # Logo optional

                        # Set row height for logo
                        ws.row_dimensions[1].height = 90

                        # ── Metadata section ──
                        meta_start_row = 2
                        ws[f"C{meta_start_row}"] = "TUNGALOY-NTK"
                        ws[f"C{meta_start_row}"].font = Font(name="Calibri", size=14, bold=True, color="1a3a5c")

                        meta_start_row = 3
                        upload_name = st.session_state.get("full_name", "Unknown User")
                        ws[f"C{meta_start_row}"] = f"Generated: {datetime.now().strftime('%d %b %Y at %H:%M')}"
                        ws[f"C{meta_start_row}"].font = Font(name="Calibri", size=10, color="555555")

                        meta_start_row = 4
                        ws[f"C{meta_start_row}"] = f"Uploaded by: {upload_name}"
                        ws[f"C{meta_start_row}"].font = Font(name="Calibri", size=10, color="555555")

                        meta_start_row = 5
                        ws[f"C{meta_start_row}"] = f"Inserts processed: {len(out_df)}"
                        ws[f"C{meta_start_row}"].font = Font(name="Calibri", size=10, color="555555")

                        # ── Results table ──
                        table_start_row = 7

                        # Write headers
                        headers = [
                            "Competitor Insert",
                            "Parsed",
                            "Competitor",
                            "ISO Material",
                            "Alt 1 Part No.",
                            "Alt 1 Webshop",
                            "Alt 2 Part No.",
                            "Alt 2 Webshop",
                            "Notes"
                        ]
                        for col_idx, header in enumerate(headers, start=1):
                            cell = ws.cell(row=table_start_row, column=col_idx)
                            cell.value = header
                            cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
                            cell.fill = PatternFill(start_color="1a3a5c", end_color="1a3a5c", fill_type="solid")
                            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

                        # ── Write data rows with formatting ──
                        thin_border = Border(
                            left=Side(style="thin"),
                            right=Side(style="thin"),
                            top=Side(style="thin"),
                            bottom=Side(style="thin")
                        )

                        for row_idx, (_, row) in enumerate(out_df.iterrows(), start=table_start_row + 1):
                            # Alternate row colors
                            row_color = "F5F5F5" if (row_idx - table_start_row - 1) % 2 == 0 else "FFFFFF"

                            # Col 1: Competitor Insert
                            cell = ws.cell(row=row_idx, column=1, value=str(row["Competitor Insert"]))
                            cell.font = Font(name="Calibri", size=10)
                            cell.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type="solid")
                            cell.border = thin_border
                            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

                            # Col 2: Parsed
                            cell = ws.cell(row=row_idx, column=2, value=str(row["Parsed"]))
                            cell.font = Font(name="Calibri", size=9)
                            cell.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type="solid")
                            cell.border = thin_border
                            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

                            # Col 3: Competitor
                            cell = ws.cell(row=row_idx, column=3, value=str(row["Competitor"]))
                            cell.font = Font(name="Calibri", size=10)
                            cell.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type="solid")
                            cell.border = thin_border
                            cell.alignment = Alignment(horizontal="left", vertical="center")

                            # Col 4: ISO Material
                            iso_val = str(row["ISO Material"]) if row["ISO Material"] else ""
                            cell = ws.cell(row=row_idx, column=4, value=iso_val)
                            cell.font = Font(name="Calibri", size=10, bold=True)
                            if iso_val:
                                iso_color = ISO_COLOURS.get(iso_val, "#888888").lstrip("#")
                                cell.fill = PatternFill(start_color=iso_color, end_color=iso_color, fill_type="solid")
                                cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
                            else:
                                cell.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type="solid")
                            cell.border = thin_border
                            cell.alignment = Alignment(horizontal="center", vertical="center")

                            # Col 5: Alt 1 Part No.
                            alt1_pn = str(row["Alt 1 Part No."]) if row["Alt 1 Part No."] else ""
                            cell = ws.cell(row=row_idx, column=5, value=alt1_pn)
                            cell.font = Font(name="Calibri", size=10)
                            cell.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type="solid")
                            cell.border = thin_border
                            cell.alignment = Alignment(horizontal="left", vertical="center")

                            # Col 6: Alt 1 Webshop (HYPERLINK)
                            alt1_url = str(row["Alt 1 Webshop Link"]) if row["Alt 1 Webshop Link"] else ""
                            if alt1_url and alt1_pn:
                                cell = ws.cell(row=row_idx, column=6)
                                cell.value = alt1_pn
                                from openpyxl.worksheet.hyperlink import Hyperlink
                                cell.hyperlink = Hyperlink(ref=cell.coordinate, target=alt1_url)
                                cell.font = Font(name="Calibri", size=10, color="0563C1", underline="single")
                                cell.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type="solid")
                                cell.border = thin_border
                                cell.alignment = Alignment(horizontal="left", vertical="center")
                            else:
                                cell = ws.cell(row=row_idx, column=6, value="")
                                cell.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type="solid")
                                cell.border = thin_border

                            # Col 7: Alt 2 Part No.
                            alt2_pn = str(row["Alt 2 Part No."]) if row["Alt 2 Part No."] else ""
                            cell = ws.cell(row=row_idx, column=7, value=alt2_pn)
                            cell.font = Font(name="Calibri", size=10)
                            cell.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type="solid")
                            cell.border = thin_border
                            cell.alignment = Alignment(horizontal="left", vertical="center")

                            # Col 8: Alt 2 Webshop (HYPERLINK)
                            alt2_url = str(row["Alt 2 Webshop Link"]) if row["Alt 2 Webshop Link"] else ""
                            if alt2_url and alt2_pn:
                                cell = ws.cell(row=row_idx, column=8)
                                cell.value = alt2_pn
                                from openpyxl.worksheet.hyperlink import Hyperlink
                                cell.hyperlink = Hyperlink(ref=cell.coordinate, target=alt2_url)
                                cell.font = Font(name="Calibri", size=10, color="0563C1", underline="single")
                                cell.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type="solid")
                                cell.border = thin_border
                                cell.alignment = Alignment(horizontal="left", vertical="center")
                            else:
                                cell = ws.cell(row=row_idx, column=8, value="")
                                cell.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type="solid")
                                cell.border = thin_border

                            # Col 9: Notes
                            cell = ws.cell(row=row_idx, column=9, value=str(row["Notes"]))
                            cell.font = Font(name="Calibri", size=9, color="D9534F" if row["Notes"] else "555555")
                            cell.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type="solid")
                            cell.border = thin_border
                            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

                            # Set row height for wrapped content
                            ws.row_dimensions[row_idx].height = None  # Auto height

                        # ── Column widths ──
                        ws.column_dimensions["A"].width = 22
                        ws.column_dimensions["B"].width = 28
                        ws.column_dimensions["C"].width = 18
                        ws.column_dimensions["D"].width = 14
                        ws.column_dimensions["E"].width = 18
                        ws.column_dimensions["F"].width = 18
                        ws.column_dimensions["G"].width = 18
                        ws.column_dimensions["H"].width = 18
                        ws.column_dimensions["I"].width = 35

                        # ── Save and download ──
                        wb.save(buf)
                        buf.seek(0)

                        st.download_button(
                            "⬇️ Download Excel Results",
                            data=buf,
                            file_name=f"tungaloy_alternatives_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        )

            except Exception as e:
                st.error(f"Error processing file: {e}")

    # ══════════════════════════════════════════════════════════════════════════
    # TAB 3 — Grade Converter
    # ══════════════════════════════════════════════════════════════════════════
    with tab3:
        c1, c2, c3 = st.columns([2, 2, 1])
        with c1:
            competitor = st.selectbox("Competitor", ["— Select —"] + company_opts, key="gc_comp")
        with c2:
            search_term = st.text_input("Competitor Grade", placeholder="e.g. GC4225", key="gc_grade")
        with c3:
            iso_filter = st.multiselect("ISO", list(ISO_LABELS), format_func=lambda x: ISO_LABELS[x], key="gc_iso")

        app_options = {"ISO Turn Insert": 1030, "Milling Insert": 1010, "Grooving": 1050}
        app_sel = st.multiselect("Application", list(app_options), default=["ISO Turn Insert"], key="gc_app")

        if competitor == "— Select —":
            st.info("Select a competitor to begin.")
        else:
            comp_id = comp_df[comp_df["CompanyName"] == competitor]["CompanyID"].values[0]
            res = grd_df[grd_df["CompanyID"] == comp_id].copy()
            if app_sel:
                res = res[res["AppID"].isin([app_options[a] for a in app_sel])]
            if iso_filter:
                res = res[res["GradeIso"].isin(iso_filter)]
            if search_term.strip():
                res = res[res["GradeGrade"].astype(str).str.upper().str.contains(search_term.strip().upper(), na=False)]

            if res.empty:
                st.warning("No matches found.")
            else:
                st.markdown(f"**{len(res)} result(s)** — ★ = First Choice &nbsp;&nbsp; *Wear resistant → first choice → tougher*")
                h_cols = [f"H{i}" for i in range(1, 11)]
                cols = st.columns([1, 2, 1, 5])
                for col, lbl in zip(cols, ["ISO", "Competitor Grade", "App", "Tungaloy Alternatives"]):
                    col.markdown(f"**{lbl}**")
                st.markdown("<hr style='margin:4px 0'>", unsafe_allow_html=True)
                for _, row in res.iterrows():
                    iso = str(row["GradeIso"]).strip()
                    alts = _collect_alts(row, h_cols)
                    app_lbl = str(row["GradeAppID"]).split(" ", 1)[-1] if pd.notna(row["GradeAppID"]) else ""
                    c_iso, c_grade, c_app, c_alts = st.columns([1, 2, 1, 5])
                    c_iso.markdown(f"<span style='background:{ISO_COLOURS.get(iso,'#888')};color:#fff;padding:2px 8px;border-radius:4px;font-weight:700'>{iso}</span>", unsafe_allow_html=True)
                    c_grade.markdown(f"**{row['GradeGrade']}**")
                    c_app.markdown(f"<span style='font-size:0.8rem;color:#555'>{app_lbl}</span>", unsafe_allow_html=True)
                    c_alts.markdown(_alt_chips(alts), unsafe_allow_html=True)

    # ══════════════════════════════════════════════════════════════════════════
    # TAB 4 — Geometry Converter
    # ══════════════════════════════════════════════════════════════════════════
    with tab4:
        g1, g2, g3 = st.columns([2, 2, 1])
        with g1:
            geo_comp = st.selectbox("Competitor", ["— Select —"] + company_opts, key="geo_comp")
        with g2:
            geo_search = st.text_input("Competitor Geometry", placeholder="e.g. MF, M3M, KM", key="geo_geo")
        with g3:
            geo_iso = st.multiselect("ISO", list(ISO_LABELS), format_func=lambda x: ISO_LABELS[x], key="geo_iso")

        insert_type = st.radio("Insert Type", ["Both", "Negative", "Positive"], horizontal=True, key="geo_type")

        if geo_comp == "— Select —":
            st.info("Select a competitor to begin.")
        else:
            g_comp_id = comp_df[comp_df["CompanyName"] == geo_comp]["CompanyID"].values[0]
            gr = geo_df[geo_df["CompanyID"] == g_comp_id].copy()
            if insert_type == "Negative":
                gr = gr[gr["AppID"] == 1130]
            elif insert_type == "Positive":
                gr = gr[gr["AppID"] == 1140]
            if geo_iso:
                gr = gr[gr["GeoIso"].isin(geo_iso)]
            if geo_search.strip():
                gr = gr[gr["GeoGeometry"].astype(str).str.upper().str.contains(geo_search.strip().upper(), na=False)]

            if gr.empty:
                st.warning("No geometry matches found.")
            else:
                st.markdown(f"**{len(gr)} result(s)** — ★ = First Choice &nbsp;&nbsp; *Lower feed → first choice → higher feed*")
                h_cols = [f"H{i}" for i in range(1, 11)]
                cols = st.columns([1, 2, 1, 5])
                for col, lbl in zip(cols, ["ISO", "Competitor Geometry", "Type", "Tungaloy Alternatives"]):
                    col.markdown(f"**{lbl}**")
                st.markdown("<hr style='margin:4px 0'>", unsafe_allow_html=True)
                for _, row in gr.iterrows():
                    iso = str(row["GeoIso"]).strip()
                    alts = _collect_alts(row, h_cols)
                    code_alts = [(v, fc) for v, fc in alts if _is_code(v)]
                    app_lbl = "Negative" if row["AppID"] == 1130 else "Positive"
                    c_iso, c_geo, c_app, c_alts = st.columns([1, 2, 1, 5])
                    c_iso.markdown(f"<span style='background:{ISO_COLOURS.get(iso,'#888')};color:#fff;padding:2px 8px;border-radius:4px;font-weight:700'>{iso}</span>", unsafe_allow_html=True)
                    c_geo.markdown(f"**{row['GeoGeometry']}**")
                    c_app.markdown(f"<span style='font-size:0.8rem;color:#555'>{app_lbl}</span>", unsafe_allow_html=True)
                    c_alts.markdown(_alt_chips(code_alts if code_alts else alts), unsafe_allow_html=True)
